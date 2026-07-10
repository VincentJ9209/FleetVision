from __future__ import annotations

import subprocess
import sys
from pathlib import Path

import pandas as pd
import pytest
import yaml
from openpyxl import load_workbook

from fleetvision.data.build_pilot_review_excel import EXCEL_COLUMNS, WORKLIST_SHEET
from fleetvision.data.build_pilot_review_packages import (
    ReviewerConfig,
    PilotReviewCollaborationConfig,
    build_pilot_review_packages,
)
from fleetvision.data.build_pilot_review_worklist import WORKLIST_COLUMNS
from fleetvision.data.merge_pilot_review_packages import (
    PilotReviewCollaborationMergeConfig,
    merge_pilot_review_packages,
    write_merge_outputs,
)


REVIEWERS = (ReviewerConfig("vincent", "Vincent"), ReviewerConfig("sister", "Reviewer_Sister"))


def make_row(index: int, project_root: Path) -> dict[str, str]:
    rel_path = f"dataset/01_raw/02_claimable_damage/images/{index:03d}.jpg"
    image_path = project_root / rel_path
    image_path.parent.mkdir(parents=True, exist_ok=True)
    image_path.write_bytes(f"image-{index}".encode("utf-8"))
    return {
        "review_id": f"review-{index:03d}",
        "image_id": f"image-{index:03d}",
        "source_bucket": "02_claimable_damage",
        "original_path": rel_path,
        "filename": f"{index:03d}.jpg",
        "suggested_photo_type_review": "exterior",
        "photo_type_confidence": "0.91",
        "suggested_angle_review": "front_left",
        "angle_confidence": "0.82",
        "auto_review_notes": "auto note",
        "seed_photo_type_review": "exterior",
        "seed_angle_review": "front_left",
        "seed_is_exterior_review": "1",
        "seed_has_visible_damage_review": "1",
        "seed_severity_review": "minor",
        "seed_review_status": "pending",
        "seed_reviewer": "",
        "seed_review_notes": "seed note",
        "human_photo_type_review": "exterior",
        "human_angle_review": "front_left",
        "human_is_exterior_review": "1",
        "human_has_visible_damage_review": "1",
        "human_severity_review": "minor",
        "human_review_status": "pending",
        "human_reviewer": "",
        "human_reviewed_at": "",
        "human_review_notes": "",
    }


def make_worklist(project_root: Path, row_count: int = 4) -> pd.DataFrame:
    return pd.DataFrame([make_row(index, project_root) for index in range(1, row_count + 1)], columns=WORKLIST_COLUMNS)


def prepare_packages(tmp_path: Path) -> tuple[PilotReviewCollaborationMergeConfig, pd.DataFrame]:
    worklist = make_worklist(tmp_path)
    guide_pdf = tmp_path / "docs" / "01_phase_guides" / "FleetVision_人工審核填寫指南_Excel協作版.pdf"
    guide_pdf.parent.mkdir(parents=True, exist_ok=True)
    guide_pdf.write_bytes(b"pdf")
    source_csv = tmp_path / "dataset" / "00_catalog" / "worklist.csv"
    source_csv.parent.mkdir(parents=True, exist_ok=True)
    worklist.to_csv(source_csv, index=False, encoding="utf-8-sig")
    package_config = PilotReviewCollaborationConfig(
        source_worklist_csv=source_csv,
        guide_pdf=guide_pdf,
        output_root=tmp_path / "outputs" / "manual_review" / "collaboration",
        project_root=tmp_path,
        reviewers=REVIEWERS,
        expected_rows=4,
    )
    build_pilot_review_packages(package_config)
    merge_config = PilotReviewCollaborationMergeConfig(
        source_worklist_csv=source_csv,
        assignment_csv=package_config.output_root / "pilot500_review_assignments.csv",
        vincent_workbook=package_config.output_root / "packages" / "vincent" / "review_workbook.xlsx",
        sister_workbook=package_config.output_root / "packages" / "sister" / "review_workbook.xlsx",
        output_csv=tmp_path / "outputs" / "manual_review" / "collaboration" / "results.csv",
        summary_csv=tmp_path / "outputs" / "metadata" / "summary.csv",
        errors_csv=tmp_path / "outputs" / "metadata" / "errors.csv",
        reviewers=REVIEWERS,
        expected_rows=4,
    )
    return merge_config, worklist


def set_cell(workbook_path: Path, row: int, column_name: str, value) -> None:
    workbook = load_workbook(workbook_path)
    worksheet = workbook[WORKLIST_SHEET]
    worksheet.cell(row=row, column=EXCEL_COLUMNS.index(column_name) + 1, value=value)
    workbook.save(workbook_path)


def test_normal_two_workbooks_merge_and_pass_validator(tmp_path: Path) -> None:
    config, source = prepare_packages(tmp_path)
    set_cell(config.vincent_workbook, 2, "human_review_status", "reviewed")
    set_cell(config.vincent_workbook, 2, "human_reviewer", "Vincent")
    set_cell(config.vincent_workbook, 2, "human_reviewed_at", "2026-07-10T09:00:00")
    set_cell(config.sister_workbook, 2, "human_review_notes", "sister note")

    result = merge_pilot_review_packages(config)
    write_merge_outputs(result, config)
    merged = pd.read_csv(config.output_csv, dtype=str, keep_default_na=False, encoding="utf-8-sig")

    assert result.is_valid is True
    assert merged.columns.tolist() == WORKLIST_COLUMNS
    assert merged["review_id"].tolist() == source["review_id"].tolist()
    assert merged.loc[0, "human_review_status"] == "reviewed"
    assert merged.loc[1, "human_review_notes"] == "sister note"
    assert not any(column.endswith(("_x", "_y")) for column in merged.columns)


def test_rejects_cross_reviewer_or_assignment_mismatch(tmp_path: Path) -> None:
    config, _ = prepare_packages(tmp_path)
    set_cell(config.vincent_workbook, 2, "review_id", "review-002")

    with pytest.raises(ValueError, match="unassigned review_id|duplicate review_id"):
        merge_pilot_review_packages(config)

    config, _ = prepare_packages(tmp_path / "assignment")
    assignments = pd.read_csv(config.assignment_csv, dtype=str, keep_default_na=False, encoding="utf-8-sig")
    assignments.loc[0, "reviewer_name"] = "Wrong"
    assignments.to_csv(config.assignment_csv, index=False, encoding="utf-8-sig")

    with pytest.raises(ValueError, match="assignment reviewer mismatch"):
        merge_pilot_review_packages(config)


def test_rejects_missing_duplicate_unknown_and_overlapping_review_id(tmp_path: Path) -> None:
    config, _ = prepare_packages(tmp_path)
    set_cell(config.vincent_workbook, 2, "review_id", "")
    with pytest.raises(ValueError, match="blank review_id"):
        merge_pilot_review_packages(config)

    config, _ = prepare_packages(tmp_path / "duplicate")
    set_cell(config.vincent_workbook, 3, "review_id", "review-001")
    with pytest.raises(ValueError, match="duplicate review_id"):
        merge_pilot_review_packages(config)

    config, _ = prepare_packages(tmp_path / "unknown")
    set_cell(config.vincent_workbook, 2, "review_id", "review-999")
    with pytest.raises(ValueError, match="unassigned review_id"):
        merge_pilot_review_packages(config)

    config, _ = prepare_packages(tmp_path / "overlap")
    set_cell(config.sister_workbook, 2, "review_id", "review-001")
    with pytest.raises(ValueError, match="unassigned review_id|overlapping review_id"):
        merge_pilot_review_packages(config)


def test_rejects_image_id_original_path_and_formula_edits(tmp_path: Path) -> None:
    config, _ = prepare_packages(tmp_path)
    set_cell(config.vincent_workbook, 2, "image_id", "wrong-image")
    with pytest.raises(ValueError, match="image_id mismatch"):
        merge_pilot_review_packages(config)

    config, _ = prepare_packages(tmp_path / "path")
    set_cell(config.vincent_workbook, 2, "original_path", "wrong/path.jpg")
    with pytest.raises(ValueError, match="original_path mismatch"):
        merge_pilot_review_packages(config)

    config, _ = prepare_packages(tmp_path / "formula")
    set_cell(config.vincent_workbook, 2, "human_review_notes", "=1+1")
    with pytest.raises(ValueError, match="contains formula"):
        merge_pilot_review_packages(config)


def test_pending_is_not_auto_changed_and_failed_merge_does_not_overwrite_output(tmp_path: Path) -> None:
    config, _ = prepare_packages(tmp_path)
    result = merge_pilot_review_packages(config)
    assert set(result.merged_dataframe["human_review_status"]) == {"pending"}

    config.output_csv.parent.mkdir(parents=True, exist_ok=True)
    config.output_csv.write_text("previous success\n", encoding="utf-8")
    set_cell(config.vincent_workbook, 2, "human_review_notes", "=1+1")
    with pytest.raises(ValueError):
        merge_pilot_review_packages(config)
    assert config.output_csv.read_text(encoding="utf-8") == "previous success\n"


def test_cli_uses_tmp_paths_without_touching_real_data(tmp_path: Path) -> None:
    repo_root = Path(__file__).resolve().parents[1]
    config, _ = prepare_packages(tmp_path)
    config_yaml = tmp_path / "collaboration_config.yaml"
    config_yaml.write_text(
        yaml.safe_dump(
            {
                "source_worklist_csv": str(config.source_worklist_csv),
                "assignment_csv": str(config.assignment_csv),
                "vincent_workbook": str(config.vincent_workbook),
                "sister_workbook": str(config.sister_workbook),
                "output_csv": str(config.output_csv),
                "summary_csv": str(config.summary_csv),
                "errors_csv": str(config.errors_csv),
                "expected_rows": 4,
                "reviewers": [
                    {"reviewer_id": "vincent", "reviewer_name": "Vincent"},
                    {"reviewer_id": "sister", "reviewer_name": "Reviewer_Sister"},
                ],
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )

    completed = subprocess.run(
        [sys.executable, str(repo_root / "scripts" / "phase04_merge_pilot_review_packages.py"), "--config", str(config_yaml)],
        text=True,
        capture_output=True,
        check=False,
    )

    assert completed.returncode == 0, completed.stderr
    assert config.output_csv.exists()
    assert config.summary_csv.exists()
    assert config.errors_csv.exists()
    assert "merged_rows: 4" in completed.stdout
