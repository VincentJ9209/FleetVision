from __future__ import annotations

import subprocess
import sys
from importlib.util import find_spec
from pathlib import Path

import pandas as pd
import pytest
import yaml

from fleetvision.data.build_pilot_review_excel import (
    DEFAULT_OUTPUT_XLSX,
    DROPDOWN_COLUMNS,
    EXCEL_COLUMNS,
    OPTIONS_SHEET,
    SUMMARY_SHEET,
    WORKLIST_SHEET,
    PilotReviewExcelConfig,
    build_pilot_review_excel,
    write_workbook,
)
from fleetvision.data.build_pilot_review_worklist import WORKLIST_COLUMNS


OPENPYXL_AVAILABLE = find_spec("openpyxl") is not None


def sample_config(tmp_path: Path | None = None, expected_rows: int = 500) -> PilotReviewExcelConfig:
    root = tmp_path or Path(".")
    return PilotReviewExcelConfig(
        input_csv=root / "worklist.csv",
        output_xlsx=root / "pilot500_human_review_interface.xlsx",
        project_root=root,
        expected_rows=expected_rows,
    )


def make_row(index: int, **overrides: str) -> dict[str, str]:
    row = {
        "review_id": f"review-{index:03d}",
        "image_id": f"image-{index:03d}",
        "source_bucket": "02_claimable_damage",
        "original_path": f"dataset/01_raw/02_claimable_damage/images/{index:03d}.jpg",
        "filename": f"{index:03d}.jpg",
        "suggested_photo_type_review": "exterior",
        "photo_type_confidence": "0.91",
        "suggested_angle_review": "front_left",
        "angle_confidence": "0.82",
        "auto_review_notes": "auto note",
        "seed_photo_type_review": "exterior",
        "seed_angle_review": "front_left",
        "seed_is_exterior_review": "1",
        "seed_has_visible_damage_review": "1",
        "seed_severity_review": "minor",
        "seed_review_status": "pending",
        "seed_reviewer": "Vincent",
        "seed_review_notes": "seed note",
        "human_photo_type_review": "exterior",
        "human_angle_review": "front_left",
        "human_is_exterior_review": "1",
        "human_has_visible_damage_review": "1",
        "human_severity_review": "minor",
        "human_review_status": "pending",
        "human_reviewer": "Vincent",
        "human_reviewed_at": "",
        "human_review_notes": "",
    }
    row.update(overrides)
    return row


def make_dataframe(row_count: int = 500) -> pd.DataFrame:
    return pd.DataFrame([make_row(index) for index in range(1, row_count + 1)], columns=WORKLIST_COLUMNS)


def write_csv(path: Path, dataframe: pd.DataFrame) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    dataframe.to_csv(path, index=False, encoding="utf-8-sig")


def require_openpyxl() -> None:
    if not OPENPYXL_AVAILABLE:
        pytest.skip("openpyxl is not installed in the project environment")


def build_and_reload(tmp_path: Path, dataframe: pd.DataFrame | None = None):
    require_openpyxl()
    from openpyxl import load_workbook

    source = dataframe if dataframe is not None else make_dataframe()
    config = sample_config(tmp_path, expected_rows=len(source))
    workbook = build_pilot_review_excel(source, config)
    write_workbook(workbook, config.output_xlsx)
    return load_workbook(config.output_xlsx)


def test_workbook_can_be_reopened_and_sheet_names_are_correct(tmp_path: Path) -> None:
    workbook = build_and_reload(tmp_path)

    assert workbook.sheetnames == ["使用說明", "覆核工作表", "進度摘要", "選項清單"]


def test_worklist_has_500_rows_and_preserves_order_and_human_values(tmp_path: Path) -> None:
    dataframe = make_dataframe()
    dataframe.loc[0, "human_review_notes"] = "=not a formula"
    workbook = build_and_reload(tmp_path, dataframe)
    worksheet = workbook[WORKLIST_SHEET]

    assert worksheet.max_row == 501
    assert worksheet.cell(row=2, column=EXCEL_COLUMNS.index("image_id") + 1).value == "image-001"
    assert worksheet.cell(row=501, column=EXCEL_COLUMNS.index("image_id") + 1).value == "image-500"
    assert worksheet.cell(row=2, column=EXCEL_COLUMNS.index("human_photo_type_review") + 1).value == "exterior"
    assert worksheet.cell(row=2, column=EXCEL_COLUMNS.index("human_review_notes") + 1).data_type == "s"


def test_dropdown_validation_freeze_panes_and_autofilter_exist(tmp_path: Path) -> None:
    workbook = build_and_reload(tmp_path)
    worksheet = workbook[WORKLIST_SHEET]

    assert worksheet.freeze_panes == "G2"
    assert worksheet.auto_filter.ref is not None
    validations = list(worksheet.data_validations.dataValidation)
    assert len(validations) >= len(DROPDOWN_COLUMNS)


def test_image_hyperlink_uses_original_path(tmp_path: Path) -> None:
    image_path = tmp_path / "dataset" / "01_raw" / "02_claimable_damage" / "images" / "001.jpg"
    image_path.parent.mkdir(parents=True, exist_ok=True)
    image_path.write_bytes(b"fake image bytes")
    workbook = build_and_reload(tmp_path)
    worksheet = workbook[WORKLIST_SHEET]

    assert worksheet.cell(row=2, column=1).value == "開啟圖片"
    assert worksheet.cell(row=2, column=1).hyperlink.target == image_path.resolve().as_uri()
    assert worksheet.cell(row=2, column=1).hyperlink.target != "dataset/01_raw/02_claimable_damage/images/001.jpg"
    assert worksheet.cell(row=2, column=1).hyperlink.target.startswith("file:///")


def test_absolute_original_path_hyperlink_uses_file_uri(tmp_path: Path) -> None:
    image_path = tmp_path / "absolute" / "001.jpg"
    image_path.parent.mkdir(parents=True, exist_ok=True)
    image_path.write_bytes(b"fake image bytes")
    dataframe = make_dataframe()
    dataframe.loc[0, "original_path"] = str(image_path)

    workbook = build_and_reload(tmp_path, dataframe)
    worksheet = workbook[WORKLIST_SHEET]

    assert worksheet.cell(row=2, column=1).value == "開啟圖片"
    assert worksheet.cell(row=2, column=1).hyperlink.target == image_path.resolve().as_uri()
    assert worksheet.cell(row=2, column=1).hyperlink.target.startswith("file:///")


def test_options_sheet_is_hidden_and_summary_formulas_exist(tmp_path: Path) -> None:
    workbook = build_and_reload(tmp_path)

    assert workbook[OPTIONS_SHEET].sheet_state == "hidden"
    summary = workbook[SUMMARY_SHEET]
    formulas = [summary.cell(row=row, column=2).value for row in range(2, summary.max_row + 1)]
    assert any(str(formula).startswith("=COUNTIF") for formula in formulas)
    assert any("COUNTIFS" in str(formula) for formula in formulas)


def test_input_dataframe_is_not_modified(tmp_path: Path) -> None:
    require_openpyxl()
    dataframe = make_dataframe()
    before = dataframe.copy(deep=True)

    build_pilot_review_excel(dataframe, sample_config(tmp_path))

    pd.testing.assert_frame_equal(dataframe, before)


def test_headers_do_not_include_merge_suffix_columns(tmp_path: Path) -> None:
    workbook = build_and_reload(tmp_path)
    worksheet = workbook[WORKLIST_SHEET]
    headers = [worksheet.cell(row=1, column=column).value for column in range(1, worksheet.max_column + 1)]

    assert not any(str(header).endswith(("_x", "_y")) for header in headers)


def test_missing_required_column_fails_clearly() -> None:
    dataframe = make_dataframe().drop(columns=["human_review_status"])

    with pytest.raises(ValueError, match="missing required column"):
        build_pilot_review_excel(dataframe, sample_config())


def test_duplicate_image_id_fails_clearly() -> None:
    dataframe = make_dataframe()
    dataframe.loc[1, "image_id"] = dataframe.loc[0, "image_id"]

    with pytest.raises(ValueError, match="duplicate image_id"):
        build_pilot_review_excel(dataframe, sample_config())


def test_cli_uses_tmp_paths_without_touching_real_files(tmp_path: Path) -> None:
    require_openpyxl()
    repo_root = Path(__file__).resolve().parents[1]
    input_csv = tmp_path / "worklist.csv"
    output_xlsx = tmp_path / "out" / "review.xlsx"
    config_yaml = tmp_path / "pilot_review_excel_config.yaml"
    write_csv(input_csv, make_dataframe(row_count=3))
    config_yaml.write_text(
        yaml.safe_dump(
            {
                "input_csv": str(input_csv),
                "output_xlsx": str(output_xlsx),
                "expected_rows": 3,
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )

    completed = subprocess.run(
        [
            sys.executable,
            str(repo_root / "scripts" / "phase04_build_pilot_review_excel.py"),
            "--config",
            str(config_yaml),
            "--project-root",
            str(tmp_path),
        ],
        text=True,
        capture_output=True,
        check=False,
    )

    assert completed.returncode == 0, completed.stderr
    assert output_xlsx.exists()
    assert "rows: 3" in completed.stdout
    from openpyxl import load_workbook

    workbook = load_workbook(output_xlsx)
    worksheet = workbook[WORKLIST_SHEET]
    expected_target = (tmp_path / "dataset" / "01_raw" / "02_claimable_damage" / "images" / "001.jpg").resolve().as_uri()
    assert worksheet.cell(row=2, column=1).hyperlink.target == expected_target


def test_missing_openpyxl_reports_clear_runtime_error() -> None:
    if OPENPYXL_AVAILABLE:
        pytest.skip("openpyxl is installed")

    with pytest.raises(RuntimeError, match="openpyxl is required"):
        build_pilot_review_excel(make_dataframe(row_count=1), sample_config(expected_rows=1))


def test_generated_excel_path_is_ignored_precisely() -> None:
    gitignore = Path(__file__).resolve().parents[1] / ".gitignore"
    lines = gitignore.read_text(encoding="utf-8").splitlines()

    assert DEFAULT_OUTPUT_XLSX.as_posix() in lines
