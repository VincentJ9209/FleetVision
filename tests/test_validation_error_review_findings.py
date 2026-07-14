from __future__ import annotations

import hashlib
import json
import zipfile
from pathlib import Path
from typing import Sequence

import pandas as pd
import pytest
import yaml
from openpyxl import load_workbook

import fleetvision.data.validation_error_review_findings as findings


CONFIG_PATH = Path("configs/data/phase04_5l_completed_review_findings_config.yaml")


def _digest(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest().upper()


def _write_config(tmp_path: Path) -> Path:
    payload = yaml.safe_load(CONFIG_PATH.read_text(encoding="utf-8"))
    completed = tmp_path / "completed.xlsx"
    source = tmp_path / "source.xlsx"
    package = tmp_path / "package.zip"
    completed.write_bytes(b"completed")
    source.write_bytes(b"source")
    with zipfile.ZipFile(package, "w") as archive:
        archive.writestr("assets/original/a.jpg", b"a")
        archive.writestr("assets/overlay/a.jpg", b"b")
    payload["completed_review"]["workbook_path"] = str(completed)
    payload["completed_review"]["workbook_size_bytes"] = completed.stat().st_size
    payload["completed_review"]["workbook_sha256"] = findings.sha256_file(completed)
    payload["completed_review"]["source_workbook_path"] = str(source)
    payload["completed_review"]["source_workbook_sha256"] = findings.sha256_file(source)
    payload["completed_review"]["frozen_package_path"] = str(package)
    payload["completed_review"]["frozen_package_sha256"] = findings.sha256_file(package)
    payload["workspace"]["parent_dir"] = str(tmp_path / "analysis")
    path = tmp_path / "findings.yaml"
    path.write_text(yaml.safe_dump(payload, sort_keys=False), encoding="utf-8")
    return path


def _materialized_config(tmp_path: Path) -> findings.FindingsConfig:
    return findings.load_findings_config(_write_config(tmp_path), Path.cwd())


def _canonical_rows(count: int = 130) -> pd.DataFrame:
    rows: list[dict[str, str]] = []
    for index in range(count):
        row = {column: "" for column in findings.CANONICAL_COLUMNS}
        row.update(
            {
                "schema_version": "1",
                "review_batch_id": "batch",
                "review_case_id": f"case_{index:03d}",
                "source_04_5k_zip_sha256": "A" * 64,
                "source_case_fingerprint": f"fp_{index:03d}",
                "image_id": f"valid_{index:03d}.jpg",
                "image_filename": f"valid_{index:03d}.jpg",
                "auto_error_category": "false_negative",
                "auto_error_detail_ids": "false_negative",
                "error_case_count": "1",
                "ground_truth_error_count": "1",
                "prediction_error_count": "0",
                "gt_count": "1",
                "prediction_count": "0",
                "max_prediction_confidence": "0.000000",
                "best_iou": "0.000000",
                "threshold_candidate": "0.20",
                "threshold_designation": "BALANCED_VALIDATION_THRESHOLD_CANDIDATE",
                "original_image_relpath": f"assets/original/valid_{index:03d}.jpg",
                "overlay_image_relpath": f"assets/overlay/case_{index:03d}.jpg",
                "review_status": "reviewed",
                "error_disposition": "confirmed_model_error",
                "primary_root_cause": "missed_small_damage",
                "secondary_root_cause": "none",
                "annotation_quality": "correct",
                "annotation_defect_type": "none",
                "recommended_action": "add_positive_sample",
                "retraining_priority": "medium",
                "correction_proposal_required": "no",
                "reviewer": "Vincent",
                "reviewed_at_utc": "2026-07-14T00:00:00+00:00",
                "review_notes": "Reviewed.",
            }
        )
        rows.append(row)
    return pd.DataFrame(rows, columns=findings.CANONICAL_COLUMNS).fillna("").astype(str)


def _scope_frames(count: int = 130) -> tuple[pd.DataFrame, pd.DataFrame]:
    canonical = _canonical_rows(count)
    source = canonical.copy()
    for column in findings.SCOPE_COLUMNS:
        source[column] = "pending" if column == "scope_review_status" else ""
    source = source.loc[:, list(findings.SCOPE_EXPORT_COLUMNS)]
    reviewed = source.copy()
    reviewed["scope_review_status"] = "reviewed"
    reviewed["scope_group"] = "IN_SCOPE_LIGHT_MODERATE"
    reviewed["scope_reason"] = "light_surface_damage"
    reviewed["operability"] = "drivable_or_likely_drivable"
    reviewed["scope_confidence"] = "high"
    reviewed["scope_reviewer"] = "Vincent"
    reviewed["scope_reviewed_at_utc"] = "2026-07-14T00:00:00+00:00"
    return reviewed, source


class _FakeCell:
    def __init__(self, value: str) -> None:
        self.value = value


class _FakeScopeSheet:
    def iter_rows(self, *, min_row: int, max_row: int | None = None):
        if min_row == 1:
            headers = (
                "Original Preview",
                "Overlay Preview",
                *findings.SCOPE_EXPORT_COLUMNS,
            )
            return iter((tuple(_FakeCell(value) for value in headers),))
        return iter(())


class _FakeScopeWorkbook:
    def __init__(self, *, valid_contract: bool = True) -> None:
        self.sheetnames = (
            findings.SCOPE_WORKBOOK_SHEETS
            if valid_contract
            else ("Unexpected",)
        )
        self.closed = False
        self.sheet = _FakeScopeSheet()

    def __getitem__(self, name: str):
        assert name == "Scope_Review"
        return self.sheet

    def close(self) -> None:
        self.closed = True


def test_read_scope_workbook_closes_read_only_handle(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "scope.xlsx"
    path.write_bytes(b"fixture")
    workbook = _FakeScopeWorkbook()
    monkeypatch.setattr(findings, "load_workbook", lambda *_args, **_kwargs: workbook)

    frame = findings.read_scope_workbook(path)

    assert frame.empty
    assert workbook.closed is True


def test_read_scope_workbook_closes_handle_on_contract_error(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "scope.xlsx"
    path.write_bytes(b"fixture")
    workbook = _FakeScopeWorkbook(valid_contract=False)
    monkeypatch.setattr(findings, "load_workbook", lambda *_args, **_kwargs: workbook)

    with pytest.raises(findings.FindingsAnalysisError, match="sheet contract"):
        findings.read_scope_workbook(path)

    assert workbook.closed is True


def test_load_findings_config_has_exact_scope_contract(tmp_path: Path) -> None:
    config = _materialized_config(tmp_path)
    assert config.expected_case_count == 130
    assert config.scope_completed_workbook_filename == (
        "severity_scope_review_completed.xlsx"
    )
    assert config.scope_options["scope_group"] == (
        "IN_SCOPE_LIGHT_MODERATE",
        "BOUNDARY_HEAVY_DAMAGE",
        "OUT_OF_SCOPE_CATASTROPHIC",
    )
    assert config.retraining_status == "NOT_YET_APPROVED"
    assert config.deployment_acceptance == "NOT_YET_APPROVED"


@pytest.mark.parametrize(
    "key",
    [
        "test_split_read",
        "model_inference_executed",
        "annotation_modified",
        "training_started",
    ],
)
def test_load_findings_config_rejects_enabled_prohibited_boundary(
    tmp_path: Path, key: str
) -> None:
    path = _write_config(tmp_path)
    payload = yaml.safe_load(path.read_text(encoding="utf-8"))
    payload["safety"][key] = True
    path.write_text(yaml.safe_dump(payload, sort_keys=False), encoding="utf-8")
    with pytest.raises(findings.FindingsAnalysisError, match=key):
        findings.load_findings_config(path, Path.cwd())


def test_repository_state_requires_equal_heads_and_allowed_status(tmp_path: Path) -> None:
    head = "a" * 40
    outputs = {
        ("branch", "--show-current"): "main\n",
        ("rev-parse", "HEAD"): f"{head}\n",
        ("rev-parse", "origin/main"): f"{head}\n",
        ("ls-remote", "origin", "refs/heads/main"): f"{head}\trefs/heads/main\n",
        ("status", "--porcelain=v1", "--untracked-files=all"): (
            "?? outputs/metadata/external_assets/source.bin\n"
        ),
    }

    def runner(args: Sequence[str], cwd: Path) -> str:
        return outputs[tuple(args)]

    state = findings.inspect_repository_state(tmp_path, head, runner=runner)
    assert state.local_head == head
    assert state.unexpected_status == ()


def test_repository_state_blocks_unexpected_entry(tmp_path: Path) -> None:
    head = "b" * 40

    def runner(args: Sequence[str], cwd: Path) -> str:
        mapping = {
            ("branch", "--show-current"): "main\n",
            ("rev-parse", "HEAD"): f"{head}\n",
            ("rev-parse", "origin/main"): f"{head}\n",
            ("ls-remote", "origin", "refs/heads/main"): f"{head}\trefs/heads/main\n",
            ("status", "--porcelain=v1", "--untracked-files=all"): " M dataset/01_raw/a.jpg\n",
        }
        return mapping[tuple(args)]

    with pytest.raises(findings.FindingsAnalysisError, match="unexpected worktree"):
        findings.inspect_repository_state(tmp_path, head, runner=runner)


def test_authoritative_inputs_verify_size_and_hash(tmp_path: Path) -> None:
    config = _materialized_config(tmp_path)
    results = findings.verify_authoritative_inputs(config)
    assert set(results) == {"completed_workbook", "source_workbook", "frozen_package"}
    assert all(row["match"] == "true" for row in results.values())


def test_hash_mismatch_fails_before_workspace_creation(tmp_path: Path) -> None:
    config = _materialized_config(tmp_path)
    config.completed_workbook.write_bytes(b"changed")
    with pytest.raises(findings.FindingsAnalysisError, match="completed_workbook"):
        findings.verify_authoritative_inputs(config)
    assert not config.workspace_parent.exists()


def test_workspace_is_timestamped_and_no_overwrite(tmp_path: Path) -> None:
    config = _materialized_config(tmp_path)
    paths = findings.create_workspace(config, timestamp="20260714T120000Z")
    assert paths.root.name == "phase04_5l_completed_review_findings_20260714T120000Z"
    with pytest.raises(findings.FindingsAnalysisError, match="already exists"):
        findings.create_workspace(config, timestamp="20260714T120000Z")


def test_safe_extract_rejects_parent_escape(tmp_path: Path) -> None:
    archive_path = tmp_path / "unsafe.zip"
    with zipfile.ZipFile(archive_path, "w") as archive:
        archive.writestr("../escape.txt", b"no")
    with pytest.raises(findings.FindingsAnalysisError, match="unsafe"):
        findings._safe_extract_zip(archive_path, tmp_path / "extract")
    assert not (tmp_path / "escape.txt").exists()


def test_scope_validation_passes_complete_review(tmp_path: Path) -> None:
    config = _materialized_config(tmp_path)
    reviewed, source = _scope_frames()
    result = findings.validate_scope_dataframe(reviewed, source, config)
    assert result.passed
    assert result.counts == {"reviewed": 130, "pending": 0, "needs_adjudication": 0}


@pytest.mark.parametrize(
    ("column", "value"),
    [
        ("scope_review_status", "done"),
        ("scope_group", "HEAVY"),
        ("scope_reason", "unknown_reason"),
        ("operability", "broken"),
        ("scope_confidence", "certain"),
    ],
)
def test_scope_controlled_values_are_enforced(
    tmp_path: Path, column: str, value: str
) -> None:
    config = _materialized_config(tmp_path)
    reviewed, source = _scope_frames()
    reviewed.loc[0, column] = value
    result = findings.validate_scope_dataframe(reviewed, source, config)
    assert not result.passed
    assert "INVALID_SCOPE_CONTROLLED_VALUE" in {
        issue["error_code"] for issue in result.issues
    }


def test_scope_semantics_and_source_immutability(tmp_path: Path) -> None:
    config = _materialized_config(tmp_path)
    reviewed, source = _scope_frames()
    reviewed.loc[0, "scope_confidence"] = "low"
    reviewed.loc[0, "scope_reviewer_notes"] = ""
    reviewed.loc[1, "scope_reason"] = "insufficient_visual_evidence"
    reviewed.loc[1, "scope_confidence"] = "medium"
    reviewed.loc[1, "scope_reviewer_notes"] = ""
    reviewed.loc[2, "image_id"] = "changed.jpg"
    result = findings.validate_scope_dataframe(reviewed, source, config)
    codes = {issue["error_code"] for issue in result.issues}
    assert "LOW_SCOPE_CONFIDENCE_NOTES_REQUIRED" in codes
    assert "INSUFFICIENT_EVIDENCE_CONTRACT_VIOLATION" in codes
    assert "SCOPE_SOURCE_FIELD_CHANGED" in codes


def test_total_variation_distance_detects_shift() -> None:
    assert findings.total_variation_distance(
        {"confirmed_model_error": 70, "annotation_issue": 30},
        {"confirmed_model_error": 90, "annotation_issue": 10},
    ) == pytest.approx(0.20)


def test_recommendation_precedence_is_fail_closed() -> None:
    assert findings.choose_primary_recommendation(
        additional_review=True,
        data_correction=True,
        scope_rebalancing=True,
        retraining_justified=True,
    ) == "ADDITIONAL_REVIEW_REQUIRED"
    assert findings.choose_primary_recommendation(
        additional_review=False,
        data_correction=True,
        scope_rebalancing=True,
        retraining_justified=True,
    ) == "DATA_CORRECTION_REQUIRED_BEFORE_RETRAINING"
    assert findings.choose_primary_recommendation(
        additional_review=False,
        data_correction=False,
        scope_rebalancing=True,
        retraining_justified=True,
    ) == "SCOPE_REBALANCING_REQUIRED_BEFORE_RETRAINING"
    assert findings.choose_primary_recommendation(
        additional_review=False,
        data_correction=False,
        scope_rebalancing=False,
        retraining_justified=True,
    ) == "RETRAINING_PROPOSAL_JUSTIFIED"
    assert findings.choose_primary_recommendation(
        additional_review=False,
        data_correction=False,
        scope_rebalancing=False,
        retraining_justified=False,
    ) == "NO_RETRAINING_RECOMMENDED"


def test_findings_payload_and_recommendation_are_advisory(tmp_path: Path) -> None:
    config = _materialized_config(tmp_path)
    reviewed, _ = _scope_frames()
    reviewed.loc[100:, "scope_group"] = "OUT_OF_SCOPE_CATASTROPHIC"
    reviewed.loc[100:, "scope_reason"] = "catastrophic_collision"
    reviewed.loc[100:, "operability"] = "non_drivable_or_likely_non_drivable"
    payload = findings.build_findings_payload(reviewed, config)
    result = findings.classify_recommendation(reviewed, payload, config)
    assert payload["all_vs_in_scope"]["non_scope_share"] == pytest.approx(30 / 130)
    assert result.primary in findings.PRIMARY_RECOMMENDATIONS
    recommendation = findings._recommendation_payload(result, config)
    assert recommendation["advisory_only"] is True
    assert recommendation["retraining_status"] == "NOT_YET_APPROVED"
    assert recommendation["deployment_acceptance"] == "NOT_YET_APPROVED"



def test_existing_review_workflow_exposes_required_findings_seams() -> None:
    import inspect
    import fleetvision.data.validation_error_human_review as review_module

    assert hasattr(review_module, "write_validation_outputs")
    signature = inspect.signature(review_module.summarize_canonical_review)
    assert "asset_root" in signature.parameters
    assert signature.parameters["asset_root"].kind is inspect.Parameter.KEYWORD_ONLY
    assert signature.parameters["asset_root"].default is None

def test_python_wrappers_and_powershell_contracts() -> None:
    root = Path.cwd()
    wrappers = (
        root / "scripts/phase04_5_run_completed_review_findings_f1.py",
        root / "scripts/phase04_5_run_completed_review_findings_f2.py",
    )
    for wrapper in wrappers:
        text = wrapper.read_text(encoding="utf-8")
        assert "validation_error_review_findings import main_" in text
        assert "ultralytics" not in text.lower()
        assert ".train(" not in text.lower()
        assert ".predict(" not in text.lower()
    for script in (
        root / "scripts/phase04_5_run_completed_review_findings_f1.ps1",
        root / "scripts/phase04_5_run_completed_review_findings_f2.ps1",
    ):
        text = script.read_text(encoding="utf-8")
        assert "#requires -Version 5.1" in text
        assert "Set-StrictMode -Version Latest" in text
        assert '$ErrorActionPreference = "Stop"' in text
        assert "ExpectedHead" in text
        for forbidden in ("git clean", "git reset --hard", "dataset/05_yolo", "YOLO("):
            assert forbidden.lower() not in text.lower()


def test_main_f1_blocks_without_repository_and_prints_safety(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    path = _write_config(tmp_path)
    exit_code = findings.main_f1(
        [
            "--config",
            str(path),
            "--project-root",
            str(tmp_path),
            "--expected-head",
            "a" * 40,
        ]
    )
    output = capsys.readouterr().out
    assert exit_code == 1
    assert "F1_BLOCKED" in output
    assert "TEST_SPLIT_READ: NO" in output
    assert "RETRAINING_STATUS: NOT_YET_APPROVED" in output


def test_scope_workbook_build_and_deterministic_export(tmp_path: Path) -> None:
    from dataclasses import replace

    config = replace(
        _materialized_config(tmp_path),
        expected_case_count=2,
        workspace_parent=tmp_path / "analysis_scope",
    )
    canonical = _canonical_rows(2)
    asset_root = tmp_path / "asset_root"
    for row in canonical.to_dict(orient="records"):
        for column, data in (
            ("original_image_relpath", b"original"),
            ("overlay_image_relpath", b"overlay"),
        ):
            asset = asset_root / row[column]
            asset.parent.mkdir(parents=True, exist_ok=True)
            from PIL import Image

            Image.new("RGB", (80, 60), (100, 110, 120)).save(asset, format="JPEG")
    paths = findings.create_workspace(config, timestamp="20260714T130000Z")
    findings.build_scope_review_package(canonical, asset_root, config, paths)
    assert paths.scope_workbook.is_file()
    assert paths.scope_source_csv.is_file()
    assert paths.scope_asset_manifest.is_file()

    workbook = load_workbook(paths.scope_workbook)
    sheet = workbook["Scope_Review"]
    header_map = {cell.value: cell.column for cell in sheet[1]}
    for row_index in range(2, sheet.max_row + 1):
        values = {
            "scope_review_status": "reviewed",
            "scope_group": "IN_SCOPE_LIGHT_MODERATE",
            "scope_reason": "light_surface_damage",
            "operability": "drivable_or_likely_drivable",
            "scope_confidence": "high",
            "scope_reviewer_notes": "",
            "scope_reviewer": "Vincent",
            "scope_reviewed_at_utc": "2026-07-14T00:00:00+00:00",
        }
        for column, value in values.items():
            sheet.cell(row=row_index, column=header_map[column], value=value)
    workbook.save(paths.scope_workbook)

    result = findings.export_scope_classification(
        config,
        paths.scope_workbook,
        paths.scope_source_csv,
        paths.scope_export_csv,
    )
    assert result.passed
    exported = pd.read_csv(
        paths.scope_export_csv,
        dtype=str,
        keep_default_na=False,
        encoding="utf-8-sig",
    )
    assert exported.columns.tolist() == list(findings.SCOPE_EXPORT_COLUMNS)
    assert exported["scope_review_status"].tolist() == ["reviewed", "reviewed"]


def test_run_f1_and_f2_orchestration_with_controlled_dependencies(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    from dataclasses import replace
    from types import SimpleNamespace
    from PIL import Image

    config = _materialized_config(tmp_path)
    canonical = _canonical_rows(2)
    completed_fingerprint = findings.logical_fingerprint(canonical)
    package = tmp_path / "package.zip"
    package_root = tmp_path / "package_source"
    for row in canonical.to_dict(orient="records"):
        for column in ("original_image_relpath", "overlay_image_relpath"):
            asset = package_root / row[column]
            asset.parent.mkdir(parents=True, exist_ok=True)
            Image.new("RGB", (80, 60), (90, 100, 110)).save(asset, format="JPEG")
    with zipfile.ZipFile(package, "w") as archive:
        for asset in sorted(package_root.rglob("*")):
            if asset.is_file():
                archive.write(asset, asset.relative_to(package_root).as_posix())

    config = replace(
        config,
        expected_case_count=2,
        completed_logical_fingerprint=completed_fingerprint,
        frozen_package=package,
        frozen_package_sha256=findings.sha256_file(package),
        workspace_parent=tmp_path / "analysis_e2e",
    )

    def git_runner(args: Sequence[str], cwd: Path) -> str:
        head = "c" * 40
        mapping = {
            ("branch", "--show-current"): "main\n",
            ("rev-parse", "HEAD"): f"{head}\n",
            ("rev-parse", "origin/main"): f"{head}\n",
            ("ls-remote", "origin", "refs/heads/main"): f"{head}\trefs/heads/main\n",
            ("status", "--porcelain=v1", "--untracked-files=all"): "",
        }
        return mapping[tuple(args)]

    monkeypatch.setattr(findings, "load_existing_review_config", lambda *args: object())
    monkeypatch.setattr(findings, "read_workbook_dataframe", lambda path: canonical.copy())

    def fake_export(existing_config, workbook_path, output_csv):
        canonical.to_csv(output_csv, index=False, encoding="utf-8-sig", lineterminator="\n")
        return SimpleNamespace(logical_fingerprint=completed_fingerprint)

    monkeypatch.setattr(findings, "export_review_workbook", fake_export)
    monkeypatch.setattr(
        findings,
        "validate_canonical_csv",
        lambda *args, **kwargs: SimpleNamespace(
            passed=True,
            issue_count=0,
            counts={"reviewed": 2, "pending": 0, "needs_adjudication": 0},
            logical_fingerprint=completed_fingerprint,
            issues=(),
        ),
    )

    def fake_write_validation(result, report_json, errors_csv):
        report_json.parent.mkdir(parents=True, exist_ok=True)
        report_json.write_text(
            json.dumps({"classification": "VALIDATION_ERROR_HUMAN_REVIEW_VERIFIED"}),
            encoding="utf-8",
        )
        pd.DataFrame(columns=["row_number", "error_code", "message"]).to_csv(
            errors_csv, index=False, encoding="utf-8-sig"
        )

    monkeypatch.setattr(findings, "write_validation_outputs", fake_write_validation)

    def fake_summary(existing_config, canonical_csv, batch_root, *, asset_root=None):
        (batch_root / "canonical").mkdir(parents=True, exist_ok=True)
        (batch_root / "reports").mkdir(parents=True, exist_ok=True)
        pd.DataFrame(
            columns=[
                "proposal_id",
                "review_case_id",
                "image_id",
                "source_annotation_ids",
                "proposal_type",
                "proposal_description",
                "evidence_original_relpath",
                "evidence_overlay_relpath",
                "reviewer",
                "proposal_status",
            ]
        ).to_csv(
            batch_root / "canonical/annotation_correction_proposals.csv",
            index=False,
            encoding="utf-8-sig",
        )
        (batch_root / "reports/review_summary.json").write_text("{}", encoding="utf-8")
        (batch_root / "reports/review_summary.md").write_text("# Summary\n", encoding="utf-8")
        pd.DataFrame().to_csv(
            batch_root / "reports/data_improvement_action_queue.csv", index=False
        )
        pd.DataFrame().to_csv(
            batch_root / "reports/data_improvement_action_summary.csv", index=False
        )
        return SimpleNamespace(action_count=0, correction_proposal_count=0)

    monkeypatch.setattr(findings, "summarize_canonical_review", fake_summary)

    paths = findings.run_f1(
        config,
        "c" * 40,
        timestamp="20260714T140000Z",
        git_runner=git_runner,
    )
    assert (paths.root / "evidence/f1_gate_result.json").is_file()
    assert (paths.root / "evidence/workspace_before.csv").is_file()

    template_hash = findings.sha256_file(paths.scope_workbook)
    paths.scope_completed_workbook.parent.mkdir(parents=True, exist_ok=True)
    import shutil

    shutil.copy2(paths.scope_workbook, paths.scope_completed_workbook)
    workbook = load_workbook(paths.scope_completed_workbook)
    sheet = workbook["Scope_Review"]
    header_map = {cell.value: cell.column for cell in sheet[1]}
    for row_index in range(2, sheet.max_row + 1):
        values = {
            "scope_review_status": "reviewed",
            "scope_group": "IN_SCOPE_LIGHT_MODERATE",
            "scope_reason": "light_surface_damage",
            "operability": "drivable_or_likely_drivable",
            "scope_confidence": "high",
            "scope_reviewer_notes": "",
            "scope_reviewer": "Vincent",
            "scope_reviewed_at_utc": "2026-07-14T00:00:00+00:00",
        }
        for column, value in values.items():
            sheet.cell(row=row_index, column=header_map[column], value=value)
    workbook.save(paths.scope_completed_workbook)
    export_evidence = {
        "outcome": "PASS",
        "classification": "LOCAL_SCOPE_REVIEW_APP_COMPLETED_WORKBOOK_EXPORTED",
        "completed_scope_workbook": str(paths.scope_completed_workbook),
        "completed_scope_workbook_sha256": findings.sha256_file(
            paths.scope_completed_workbook
        ),
        "review_cases": 2,
        "reviewed": 2,
        "pending": 0,
        "needs_adjudication": 0,
        "source_scope_csv_sha256": findings.sha256_file(paths.scope_source_csv),
        "source_scope_template_sha256": findings.sha256_file(paths.scope_workbook),
        "source_scope_asset_manifest_sha256": findings.sha256_file(
            paths.scope_asset_manifest
        ),
        "test_split_read": False,
        "model_inference_executed": False,
        "annotation_modified": False,
        "training_started": False,
        "retraining_status": "NOT_YET_APPROVED",
        "deployment_acceptance": "NOT_YET_APPROVED",
    }
    (paths.scope_completed_workbook.parent / "scope_review_export_result.json").write_text(
        json.dumps(export_evidence), encoding="utf-8"
    )
    assert findings.sha256_file(paths.scope_workbook) == template_hash

    recommendation = findings.run_f2(
        config,
        paths.root,
        "c" * 40,
        git_runner=git_runner,
    )
    assert recommendation.primary == "RETRAINING_PROPOSAL_JUSTIFIED"
    gate = json.loads((paths.root / "evidence/gate_result.json").read_text(encoding="utf-8"))
    assert gate["outcome"] == "PASS"
    assert gate["scope_reviewed"] == 2
    assert gate["retraining_status"] == "NOT_YET_APPROVED"
    assert (paths.root / "evidence/SHA256SUMS.csv").is_file()


def test_f1_checksum_manifest_treats_scope_template_as_immutable(
    tmp_path: Path,
) -> None:
    root = tmp_path / "workspace"
    (root / "evidence").mkdir(parents=True)
    (root / "scope_review").mkdir(parents=True)
    immutable = root / "canonical/review.csv"
    immutable.parent.mkdir(parents=True)
    immutable.write_text("a,b\n1,2\n", encoding="utf-8")
    scope_workbook = root / "scope_review/severity_scope_review.xlsx"
    scope_workbook.write_bytes(b"readonly template")
    pd.DataFrame(
        [
            {
                "relative_path": "canonical/review.csv",
                "size_bytes": str(immutable.stat().st_size),
                "sha256": findings.sha256_file(immutable),
            },
            {
                "relative_path": "scope_review/severity_scope_review.xlsx",
                "size_bytes": str(scope_workbook.stat().st_size),
                "sha256": findings.sha256_file(scope_workbook),
            },
        ]
    ).to_csv(
        root / "evidence/F1_SHA256SUMS.csv",
        index=False,
        encoding="utf-8-sig",
    )
    findings._verify_f1_checksum_manifest(root)
    scope_workbook.write_bytes(b"direct Excel edit is forbidden")
    with pytest.raises(findings.FindingsAnalysisError, match="checksummed output"):
        findings._verify_f1_checksum_manifest(root)


def test_completed_scope_export_verifier_rejects_hash_mismatch(
    tmp_path: Path,
) -> None:
    from types import SimpleNamespace

    root = tmp_path / "workspace"
    scope_dir = root / "scope_review"
    export_dir = root / "scope_review_app/exports"
    scope_dir.mkdir(parents=True)
    export_dir.mkdir(parents=True)
    source = scope_dir / "severity_scope_review_source.csv"
    template = scope_dir / "severity_scope_review.xlsx"
    asset_manifest = scope_dir / "scope_asset_manifest.csv"
    completed = export_dir / "severity_scope_review_completed.xlsx"
    for path, payload in (
        (source, b"source"),
        (template, b"template"),
        (asset_manifest, b"assets"),
        (completed, b"completed"),
    ):
        path.write_bytes(payload)
    paths = SimpleNamespace(
        root=root,
        scope_source_csv=source,
        scope_workbook=template,
        scope_asset_manifest=asset_manifest,
        scope_completed_workbook=completed,
    )
    config = SimpleNamespace(
        expected_case_count=130,
        retraining_status="NOT_YET_APPROVED",
        deployment_acceptance="NOT_YET_APPROVED",
    )
    evidence = {
        "outcome": "PASS",
        "classification": "LOCAL_SCOPE_REVIEW_APP_COMPLETED_WORKBOOK_EXPORTED",
        "completed_scope_workbook": str(completed),
        "completed_scope_workbook_sha256": "0" * 64,
        "review_cases": 130,
        "reviewed": 130,
        "pending": 0,
        "needs_adjudication": 0,
        "source_scope_csv_sha256": findings.sha256_file(source),
        "source_scope_template_sha256": findings.sha256_file(template),
        "source_scope_asset_manifest_sha256": findings.sha256_file(asset_manifest),
        "test_split_read": False,
        "model_inference_executed": False,
        "annotation_modified": False,
        "training_started": False,
        "retraining_status": "NOT_YET_APPROVED",
        "deployment_acceptance": "NOT_YET_APPROVED",
    }
    (export_dir / "scope_review_export_result.json").write_text(
        json.dumps(evidence), encoding="utf-8"
    )
    with pytest.raises(findings.FindingsAnalysisError, match="SHA256 mismatch"):
        findings.verify_completed_scope_review_export(config, paths)


def test_ranked_action_recommendations_are_priority_weighted() -> None:
    frame = pd.DataFrame(
        [
            {"recommended_action": "add_positive_sample", "retraining_priority": "high"},
            {"recommended_action": "add_positive_sample", "retraining_priority": "medium"},
            {"recommended_action": "threshold_analysis_only", "retraining_priority": "low"},
            {"recommended_action": "no_action", "retraining_priority": "not_applicable"},
        ]
    )
    ranked = findings._ranked_action_recommendations(frame)
    assert [row["recommended_action"] for row in ranked] == [
        "add_positive_sample",
        "threshold_analysis_only",
    ]
    assert ranked[0]["priority_score"] == 5
