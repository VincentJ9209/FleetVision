from __future__ import annotations

import csv
import json
import zipfile
from pathlib import Path

import yaml
from openpyxl import Workbook, load_workbook
from PIL import Image

from fleetvision.data.validation_error_human_review import (
    CANONICAL_COLUMNS,
    HUMAN_COLUMNS,
    SOURCE_COLUMNS,
    WORKBOOK_SHEETS,
    sha256_file,
    source_case_fingerprint,
)


def _canonical_config(case_count: int) -> dict[str, object]:
    return {
        "schema_version": "1",
        "expected_source": {
            "zip_filename": "source_04_5k.zip",
            "zip_sha256": "A" * 64,
            "gate_classification": (
                "VALIDATION_ERROR_ANALYSIS_AND_THRESHOLD_CANDIDATES_COMPLETED"
            ),
            "case_count": case_count,
            "validation_image_count": case_count,
            "validation_ground_truth_instances": case_count,
            "raw_prediction_count": case_count,
            "threshold_candidate": 0.20,
            "threshold_designation": (
                "BALANCED_VALIDATION_THRESHOLD_CANDIDATE"
            ),
        },
        "source_files": {
            "gate_result": "04_5K_gate_result.json",
            "worklist": "validation_error_review_worklist.csv",
            "predictions": "validation_predictions.csv",
            "ground_truth": "validation_ground_truth.csv",
            "artifact_manifest": "phase_04_5k_artifact_manifest.csv",
            "checksums": "phase_04_5k_checksums.sha256",
        },
        "output": {"base_dir": "outputs/metadata/phase_04_5l"},
        "workbook": {
            "filename": "validation_error_human_review.xlsx",
            "thumbnail_width": 300,
            "thumbnail_height": 190,
        },
        "options": {
            "review_status": ["pending", "reviewed", "needs_adjudication"],
            "error_disposition": [
                "confirmed_model_error",
                "annotation_issue",
                "ambiguous_case",
                "expected_threshold_tradeoff",
                "invalid_review_case",
            ],
            "primary_root_cause": [
                "missed_small_damage",
                "weak_visual_signal",
                "difficult_lighting_or_reflection",
                "occlusion_or_crop",
                "localization_error",
                "duplicate_prediction",
                "background_false_positive",
                "annotation_missing",
                "annotation_inaccurate_bbox",
                "ambiguous_visual_evidence",
                "invalid_or_low_quality_image",
                "other",
            ],
            "secondary_root_cause": [
                "none",
                "missed_small_damage",
                "weak_visual_signal",
                "difficult_lighting_or_reflection",
                "occlusion_or_crop",
                "localization_error",
                "duplicate_prediction",
                "background_false_positive",
                "annotation_missing",
                "annotation_inaccurate_bbox",
                "ambiguous_visual_evidence",
                "invalid_or_low_quality_image",
                "other",
            ],
            "annotation_quality": [
                "correct",
                "questionable",
                "defect_suspected",
                "not_applicable",
            ],
            "annotation_defect_type": [
                "none",
                "missing_bbox",
                "extra_bbox",
                "inaccurate_bbox",
                "wrong_damage_scope",
                "ambiguous_annotation",
                "invalid_image_assignment",
                "other",
            ],
            "recommended_action": [
                "no_action",
                "add_hard_negative",
                "add_positive_sample",
                "improve_annotation_guideline",
                "create_annotation_correction_proposal",
                "investigate_preprocessing",
                "investigate_image_quality",
                "adjust_model_strategy",
                "threshold_analysis_only",
                "exclude_invalid_image_proposal",
                "other",
            ],
            "retraining_priority": [
                "not_applicable",
                "low",
                "medium",
                "high",
            ],
            "correction_proposal_required": ["no", "yes"],
        },
    }


def _write_workbook(batch_root: Path, case_count: int) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in WORKBOOK_SHEETS:
        workbook.create_sheet(sheet_name)

    sheet = workbook["Review_Cases"]
    sheet.append(["Original Preview", "Overlay Preview", *CANONICAL_COLUMNS])

    for index in range(case_count):
        image_id = f"valid_{index:03d}.jpg"
        review_case_id = f"review_{index:03d}"
        original_relpath = f"assets/original/{image_id}"
        overlay_relpath = f"assets/overlay/{review_case_id}.jpg"

        Image.new("RGB", (32, 24), (100 + index, 110, 120)).save(
            batch_root / original_relpath
        )
        Image.new("RGB", (32, 24), (120, 100 + index, 110)).save(
            batch_root / overlay_relpath
        )

        source = {
            "schema_version": "1",
            "review_batch_id": batch_root.name,
            "review_case_id": review_case_id,
            "source_04_5k_zip_sha256": "A" * 64,
            "source_case_fingerprint": "",
            "image_id": image_id,
            "image_filename": image_id,
            "auto_error_category": "false_negative",
            "auto_error_detail_ids": "false_negative",
            "error_case_count": "1",
            "ground_truth_error_count": "1",
            "prediction_error_count": "0",
            "gt_count": "1",
            "prediction_count": "0",
            "max_prediction_confidence": "0.100000",
            "best_iou": "0.000000",
            "threshold_candidate": "0.20",
            "threshold_designation": (
                "BALANCED_VALIDATION_THRESHOLD_CANDIDATE"
            ),
            "original_image_relpath": original_relpath,
            "overlay_image_relpath": overlay_relpath,
        }
        source["source_case_fingerprint"] = source_case_fingerprint(source)
        human = {
            column: "pending" if column == "review_status" else ""
            for column in HUMAN_COLUMNS
        }
        row = {**source, **human}
        sheet.append(["", "", *[row[column] for column in CANONICAL_COLUMNS]])

    workbook_path = (
        batch_root / "workbook/validation_error_human_review.xlsx"
    )
    workbook.save(workbook_path)
    return workbook_path


def refresh_package_integrity(batch_root: Path, frozen_zip: Path) -> None:
    asset_manifest = batch_root / "manifest/asset_manifest.csv"
    checksums = batch_root / "manifest/checksums.sha256"
    asset_manifest.unlink(missing_ok=True)
    checksums.unlink(missing_ok=True)

    rows = []
    for path in sorted(
        candidate for candidate in batch_root.rglob("*")
        if candidate.is_file()
    ):
        rows.append(
            {
                "relative_path": path.relative_to(batch_root).as_posix(),
                "size_bytes": str(path.stat().st_size),
                "sha256": sha256_file(path),
            }
        )

    with asset_manifest.open(
        "w",
        encoding="utf-8-sig",
        newline="",
    ) as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["relative_path", "size_bytes", "sha256"],
        )
        writer.writeheader()
        writer.writerows(rows)

    checksum_paths = sorted(
        candidate for candidate in batch_root.rglob("*")
        if candidate.is_file()
    )
    checksums.write_text(
        "\n".join(
            f"{sha256_file(path)}  "
            f"{path.relative_to(batch_root).as_posix()}"
            for path in checksum_paths
        )
        + "\n",
        encoding="utf-8",
    )

    frozen_zip.unlink(missing_ok=True)
    with zipfile.ZipFile(
        frozen_zip,
        "w",
        compression=zipfile.ZIP_DEFLATED,
    ) as archive:
        for path in sorted(
            candidate for candidate in batch_root.rglob("*")
            if candidate.is_file()
        ):
            archive.write(
                path,
                arcname=(
                    Path(batch_root.name)
                    / path.relative_to(batch_root)
                ).as_posix(),
            )


def create_review_package(
    tmp_path: Path,
    *,
    case_count: int = 2,
    parent_name: str = "review_source",
) -> tuple[Path, Path, Path]:
    project_root = tmp_path / "project"
    canonical_config_path = (
        project_root
        / "configs/data/validation_error_human_review_config.yaml"
    )
    canonical_config_path.parent.mkdir(parents=True)
    canonical_config_path.write_text(
        yaml.safe_dump(
            _canonical_config(case_count),
            sort_keys=False,
        ),
        encoding="utf-8",
    )

    batch_root = tmp_path / parent_name / "batch_001"
    (batch_root / "assets/original").mkdir(parents=True)
    (batch_root / "assets/overlay").mkdir(parents=True)
    (batch_root / "workbook").mkdir(parents=True)
    (batch_root / "manifest").mkdir(parents=True)

    source_manifest = {
        "gate_id": "04.5L-PREP",
        "classification": (
            "VALIDATION_ERROR_HUMAN_REVIEW_PACKAGE_PREPARED"
        ),
        "created_at_utc": "2026-07-14T00:00:00+00:00",
        "review_batch_id": batch_root.name,
        "schema_version": "1",
        "source_zip_filename": "source_04_5k.zip",
        "source_zip_sha256": "A" * 64,
        "source_gate_classification": (
            "VALIDATION_ERROR_ANALYSIS_AND_THRESHOLD_CANDIDATES_COMPLETED"
        ),
        "case_count": case_count,
        "threshold_candidate": "0.20",
        "threshold_designation": (
            "BALANCED_VALIDATION_THRESHOLD_CANDIDATE"
        ),
        "test_split_read": False,
        "model_inference_executed": False,
        "training_started": False,
        "annotation_modified": False,
        "retraining_status": "NOT_YET_APPROVED",
        "deployment_acceptance": "NOT_YET_APPROVED",
        "source_artifacts": {},
    }
    (batch_root / "manifest/source_manifest.json").write_text(
        json.dumps(source_manifest),
        encoding="utf-8",
    )

    _write_workbook(batch_root, case_count)

    frozen_zip = tmp_path / "batch_001_PACKAGE.zip"
    refresh_package_integrity(batch_root, frozen_zip)
    return project_root, batch_root, frozen_zip


def write_app_config(
    tmp_path: Path,
    project_root: Path,
    batch_root: Path,
    frozen_zip: Path,
    *,
    workbook_sha256: str | None = None,
    frozen_zip_sha256: str | None = None,
    case_count: int = 2,
    workspace_root: Path | None = None,
) -> Path:
    workbook = batch_root / "workbook/validation_error_human_review.xlsx"
    workspace = workspace_root or (tmp_path / "workspace")
    config = {
        "schema_version": "1",
        "source": {
            "batch_root": batch_root.as_posix(),
            "workbook_relative_path": (
                "workbook/validation_error_human_review.xlsx"
            ),
            "workbook_sha256": (
                workbook_sha256 or sha256_file(workbook)
            ),
            "frozen_zip_path": frozen_zip.as_posix(),
            "frozen_zip_sha256": (
                frozen_zip_sha256 or sha256_file(frozen_zip)
            ),
            "expected_case_count": case_count,
            "canonical_config_path": (
                "configs/data/validation_error_human_review_config.yaml"
            ),
        },
        "workspace": {
            "root": workspace.as_posix(),
            "reviewer": "Vincent",
            "timezone": "Asia/Taipei",
            "backup_every_successful_saves": 10,
            "backup_retention": 20,
            "completed_workbook_name": (
                "validation_error_human_review_completed.xlsx"
            ),
        },
    }
    config_path = project_root / "review_app.yaml"
    config_path.write_text(
        yaml.safe_dump(config, sort_keys=False),
        encoding="utf-8",
    )
    return config_path


def set_workbook_value(
    workbook_path: Path,
    review_case_id: str,
    column: str,
    value: str,
) -> None:
    workbook = load_workbook(workbook_path)
    sheet = workbook["Review_Cases"]
    headers = {
        str(cell.value): cell.column
        for cell in sheet[1]
    }
    case_column = headers["review_case_id"]
    for row_index in range(2, sheet.max_row + 1):
        if str(sheet.cell(row=row_index, column=case_column).value) == review_case_id:
            sheet.cell(
                row=row_index,
                column=headers[column],
                value=value,
            )
            workbook.save(workbook_path)
            return
    raise AssertionError(f"review case not found: {review_case_id}")
