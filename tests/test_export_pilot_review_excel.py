from __future__ import annotations

import subprocess
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
import pytest
import yaml
from openpyxl import load_workbook

from fleetvision.data.build_pilot_review_excel import (
    EXCEL_COLUMNS,
    WORKLIST_SHEET,
    PilotReviewExcelConfig,
    build_pilot_review_excel,
    write_workbook,
)
from fleetvision.data.build_pilot_review_worklist import HUMAN_COLUMNS, WORKLIST_COLUMNS
from fleetvision.data.export_pilot_review_excel import (
    PilotReviewExcelExportConfig,
    export_pilot_review_excel,
    merge_excel_human_fields,
    write_export_outputs,
)
from fleetvision.data.validate_pilot_human_review import PilotHumanReviewValidationConfig, validate_pilot_human_review


def make_row(index: int, **overrides: str) -> dict[str, str]:
    row = {
        "review_id": f"review-{index:03d}",
        "image_id": f"image-{index:03d}",
        "source_bucket": "02_claimable_damage",
        "original_path": f"dataset/01_raw/02_claimable_damage/images/{index:03d}.jpg",
        "filename": f"{index:03d}.jpg",
        "suggested_photo_type_review": "exterior",
        "photo_type_confidence": "0.91",
        "suggested_angle_review": "front_left",
        "angle_confidence": "0.82",
        "auto_review_notes": "auto note",
        "seed_photo_type_review": "exterior",
        "seed_angle_review": "front_left",
        "seed_is_exterior_review": "1",
        "seed_has_visible_damage_review": "1",
        "seed_severity_review": "minor",
        "seed_review_status": "pending",
        "seed_reviewer": "Vincent",
        "seed_review_notes": "seed note",
        "human_photo_type_review": "exterior",
        "human_angle_review": "front_left",
        "human_is_exterior_review": "1",
        "human_has_visible_damage_review": "1",
        "human_severity_review": "minor",
        "human_review_status": "pending",
        "human_reviewer": "Vincent",
        "human_reviewed_at": "",
        "human_review_notes": "",
    }
    row.update(overrides)
    return row


def make_worklist(row_count: int = 3) -> pd.DataFrame:
    return pd.DataFrame([make_row(index) for index in range(1, row_count + 1)], columns=WORKLIST_COLUMNS)


def write_csv(path: Path, dataframe: pd.DataFrame) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    dataframe.to_csv(path, index=False, encoding="utf-8-sig")


def config_for(tmp_path: Path, row_count: int = 3) -> PilotReviewExcelExportConfig:
    return PilotReviewExcelExportConfig(
        input_xlsx=tmp_path / "review.xlsx",
        source_worklist_csv=tmp_path / "worklist.csv",
        output_csv=tmp_path / "out" / "results.csv",
        summary_csv=tmp_path / "out" / "summary.csv",
        errors_csv=tmp_path / "out" / "errors.csv",
        expected_rows=row_count,
    )


def build_workbook_from_source(tmp_path: Path, source: pd.DataFrame) -> Path:
    workbook = build_pilot_review_excel(
        source,
        PilotReviewExcelConfig(
            input_csv=tmp_path / "worklist.csv",
            output_xlsx=tmp_path / "review.xlsx",
            project_root=tmp_path,
            expected_rows=len(source),
        ),
    )
    output = tmp_path / "review.xlsx"
    write_workbook(workbook, output)
    return output


def column_index(column_name: str) -> int:
    return EXCEL_COLUMNS.index(column_name) + 1


def set_workbook_cell(path: Path, row: int, column_name: str, value) -> None:
    workbook = load_workbook(path)
    worksheet = workbook[WORKLIST_SHEET]
    worksheet.cell(row=row, column=column_index(column_name), value=value)
    workbook.save(path)


def prepare_files(tmp_path: Path, row_count: int = 3) -> tuple[PilotReviewExcelExportConfig, pd.DataFrame]:
    source = make_worklist(row_count)
    config = config_for(tmp_path, row_count)
    write_csv(config.source_worklist_csv, source)
    build_workbook_from_source(tmp_path, source)
    return config, source


def test_exports_human_edits_and_preserves_canonical_order_and_columns(tmp_path: Path) -> None:
    config, source = prepare_files(tmp_path)
    set_workbook_cell(config.input_xlsx, 2, "human_review_status", "reviewed")
    set_workbook_cell(config.input_xlsx, 2, "human_reviewer", "Joanna")
    set_workbook_cell(config.input_xlsx, 2, "human_reviewed_at", "2026-07-10T09:30:00")
    set_workbook_cell(config.input_xlsx, 2, "human_review_notes", "checked")
    set_workbook_cell(config.input_xlsx, 2, "source_bucket", "tampered")

    result = export_pilot_review_excel(config)
    write_export_outputs(result, config)
    exported = pd.read_csv(config.output_csv, dtype=str, keep_default_na=False, encoding="utf-8-sig")

    assert result.is_valid is True
    assert exported.columns.tolist() == WORKLIST_COLUMNS
    assert exported["review_id"].tolist() == source["review_id"].tolist()
    assert exported.loc[0, "source_bucket"] == source.loc[0, "source_bucket"]
    assert exported.loc[0, "human_review_status"] == "reviewed"
    assert exported.loc[0, "human_reviewer"] == "Joanna"
    assert exported.loc[0, "human_review_notes"] == "checked"
    assert not any(column.endswith(("_x", "_y")) for column in exported.columns)


def test_identity_mismatch_fails_for_image_id_and_original_path(tmp_path: Path) -> None:
    config, _ = prepare_files(tmp_path)
    set_workbook_cell(config.input_xlsx, 2, "image_id", "wrong-image")

    with pytest.raises(ValueError, match="image_id mismatch"):
        export_pilot_review_excel(config)

    config, _ = prepare_files(tmp_path / "second")
    set_workbook_cell(config.input_xlsx, 2, "original_path", "wrong/path.jpg")

    with pytest.raises(ValueError, match="original_path mismatch"):
        export_pilot_review_excel(config)


def test_missing_duplicate_and_unknown_review_id_fail(tmp_path: Path) -> None:
    config, _ = prepare_files(tmp_path)
    set_workbook_cell(config.input_xlsx, 2, "review_id", "")

    with pytest.raises(ValueError, match="blank review_id"):
        export_pilot_review_excel(config)

    config, _ = prepare_files(tmp_path / "duplicate")
    set_workbook_cell(config.input_xlsx, 3, "review_id", "review-001")

    with pytest.raises(ValueError, match="duplicate review_id"):
        export_pilot_review_excel(config)

    config, _ = prepare_files(tmp_path / "unknown")
    set_workbook_cell(config.input_xlsx, 2, "review_id", "review-999")

    with pytest.raises(ValueError, match="missing review_id"):
        export_pilot_review_excel(config)


def test_formula_in_human_field_fails(tmp_path: Path) -> None:
    config, _ = prepare_files(tmp_path)
    set_workbook_cell(config.input_xlsx, 2, "human_review_notes", "=HYPERLINK(\"x\")")

    with pytest.raises(ValueError, match="contains formula"):
        export_pilot_review_excel(config)


def test_excel_datetime_serializes_to_validator_compatible_format(tmp_path: Path) -> None:
    config, _ = prepare_files(tmp_path)
    set_workbook_cell(config.input_xlsx, 2, "human_review_status", "reviewed")
    set_workbook_cell(config.input_xlsx, 2, "human_reviewed_at", datetime(2026, 7, 10, 9, 30, 5))

    result = export_pilot_review_excel(config)

    assert result.is_valid is True
    assert result.exported_dataframe.loc[0, "human_reviewed_at"] == "2026-07-10 09:30:05"


def test_pending_is_not_auto_changed_to_reviewed(tmp_path: Path) -> None:
    config, _ = prepare_files(tmp_path)

    result = export_pilot_review_excel(config)

    assert result.exported_dataframe["human_review_status"].tolist() == ["pending", "pending", "pending"]
    assert result.summary["reviewed_rows"] == 0


def test_failed_export_does_not_overwrite_existing_output(tmp_path: Path) -> None:
    config, _ = prepare_files(tmp_path)
    config.output_csv.parent.mkdir(parents=True, exist_ok=True)
    config.output_csv.write_text("previous success\n", encoding="utf-8")
    set_workbook_cell(config.input_xlsx, 2, "human_review_notes", "=1+1")

    with pytest.raises(ValueError):
        export_pilot_review_excel(config)

    assert config.output_csv.read_text(encoding="utf-8") == "previous success\n"


def test_export_result_passes_existing_validator(tmp_path: Path) -> None:
    config, _ = prepare_files(tmp_path)
    result = export_pilot_review_excel(config)

    validation = validate_pilot_human_review(
        result.exported_dataframe,
        PilotHumanReviewValidationConfig(
            input_csv=config.output_csv,
            summary_csv=config.summary_csv,
            errors_csv=config.errors_csv,
            expected_rows=3,
        ),
    )

    assert validation.is_valid is True


def test_cli_tmp_path_writes_result_summary_and_errors(tmp_path: Path) -> None:
    repo_root = Path(__file__).resolve().parents[1]
    config, _ = prepare_files(tmp_path)
    config_yaml = tmp_path / "export_config.yaml"
    config_yaml.write_text(
        yaml.safe_dump(
            {
                "input_xlsx": str(config.input_xlsx),
                "source_worklist_csv": str(config.source_worklist_csv),
                "output_csv": str(config.output_csv),
                "summary_csv": str(config.summary_csv),
                "errors_csv": str(config.errors_csv),
                "expected_rows": 3,
            },
            sort_keys=False,
        ),
        encoding="utf-8",
    )

    completed = subprocess.run(
        [
            sys.executable,
            str(repo_root / "scripts" / "phase04_export_pilot_review_excel.py"),
            "--config",
            str(config_yaml),
        ],
        text=True,
        capture_output=True,
        check=False,
    )

    assert completed.returncode == 0, completed.stderr
    assert config.output_csv.exists()
    assert config.summary_csv.exists()
    assert config.errors_csv.exists()
    assert "exported_rows: 3" in completed.stdout


def test_source_worklist_file_is_not_modified(tmp_path: Path) -> None:
    config, _ = prepare_files(tmp_path)
    before = config.source_worklist_csv.read_bytes()

    result = export_pilot_review_excel(config)
    write_export_outputs(result, config)

    assert config.source_worklist_csv.read_bytes() == before
