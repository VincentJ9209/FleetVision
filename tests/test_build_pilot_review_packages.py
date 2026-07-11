from __future__ import annotations

import shutil
import zipfile
from pathlib import Path
from zipfile import ZipFile

import pandas as pd
import pytest
from openpyxl import load_workbook

from fleetvision.data.build_pilot_review_excel import DROPDOWN_COLUMNS, EXCEL_COLUMNS, OPTIONS_SHEET, WORKLIST_SHEET
from fleetvision.data.build_pilot_review_packages import (
    MANIFEST_FILENAME,
    PACKAGE_GUIDE_FILENAME,
    README_FILENAME,
    WORKBOOK_FILENAME,
    PilotReviewCollaborationConfig,
    ReviewerConfig,
    build_assignments,
    build_pilot_review_packages,
)
from fleetvision.data.build_pilot_review_worklist import WORKLIST_COLUMNS


REVIEWERS = (ReviewerConfig("vincent", "Vincent"), ReviewerConfig("sister", "Reviewer_Sister"))


def make_row(index: int, project_root: Path) -> dict[str, str]:
    rel_path = f"dataset/01_raw/02_claimable_damage/images/{index:03d}.jpg"
    image_path = project_root / rel_path
    image_path.parent.mkdir(parents=True, exist_ok=True)
    image_path.write_bytes(f"image-{index}".encode("utf-8"))
    return {
        "review_id": f"review-{index:03d}",
        "image_id": f"image-{index:03d}",
        "source_bucket": "02_claimable_damage",
        "original_path": rel_path,
        "filename": f"{index:03d}.jpg",
        "suggested_photo_type_review": "exterior",
        "photo_type_confidence": "0.91",
        "suggested_angle_review": "front_left",
        "angle_confidence": "0.82",
        "auto_review_notes": "auto note",
        "seed_photo_type_review": "exterior",
        "seed_angle_review": "front_left",
        "seed_is_exterior_review": "1",
        "seed_has_visible_damage_review": "1",
        "seed_severity_review": "minor",
        "seed_review_status": "pending",
        "seed_reviewer": "",
        "seed_review_notes": "seed note",
        "human_photo_type_review": "exterior",
        "human_angle_review": "front_left",
        "human_is_exterior_review": "1",
        "human_has_visible_damage_review": "1",
        "human_severity_review": "minor",
        "human_review_status": "pending",
        "human_reviewer": "",
        "human_reviewed_at": "",
        "human_review_notes": "",
    }


def make_worklist(project_root: Path, row_count: int = 4) -> pd.DataFrame:
    return pd.DataFrame([make_row(index, project_root) for index in range(1, row_count + 1)], columns=WORKLIST_COLUMNS)


def write_worklist(path: Path, dataframe: pd.DataFrame) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    dataframe.to_csv(path, index=False, encoding="utf-8-sig")


def image_formula_path(formula: str) -> str:
    marker = '&"'
    start = formula.index(marker) + len(marker)
    end = formula.index('"', start)
    return formula[start:end]


def review_sheet_xml_contains_data_validations(workbook_path: Path) -> bool:
    with ZipFile(workbook_path) as workbook_zip:
        for name in workbook_zip.namelist():
            if name.startswith("xl/worksheets/") and name.endswith(".xml"):
                if b"<dataValidations" in workbook_zip.read(name):
                    return True
    return False


def data_validation_by_range(worksheet) -> dict[str, object]:
    return {str(validation.sqref): validation for validation in worksheet.data_validations.dataValidation}


def make_config(tmp_path: Path, row_count: int = 4) -> tuple[PilotReviewCollaborationConfig, pd.DataFrame]:
    worklist = make_worklist(tmp_path, row_count)
    guide_pdf = tmp_path / "docs" / "01_phase_guides" / "FleetVision_人工審核填寫指南_Excel協作版.pdf"
    guide_pdf.parent.mkdir(parents=True, exist_ok=True)
    guide_pdf.write_bytes(b"pdf")
    source_csv = tmp_path / "dataset" / "00_catalog" / "worklist.csv"
    write_worklist(source_csv, worklist)
    return (
        PilotReviewCollaborationConfig(
            source_worklist_csv=source_csv,
            guide_pdf=guide_pdf,
            output_root=tmp_path / "outputs" / "manual_review" / "collaboration",
            project_root=tmp_path,
            reviewers=REVIEWERS,
            expected_rows=row_count,
        ),
        worklist,
    )


def test_round_robin_assignments_are_deterministic_balanced_and_complete(tmp_path: Path) -> None:
    _, worklist = make_config(tmp_path)
    assignments = build_assignments(worklist, REVIEWERS, expected_rows=4)

    assert assignments["reviewer_id"].tolist() == ["vincent", "sister", "vincent", "sister"]
    assert assignments["review_id"].tolist() == worklist["review_id"].tolist()
    assert assignments["review_id"].is_unique
    assert set(assignments["review_id"]) == set(worklist["review_id"])
    assert set(assignments.loc[assignments["reviewer_id"] == "vincent", "review_id"]).isdisjoint(
        set(assignments.loc[assignments["reviewer_id"] == "sister", "review_id"])
    )


def test_build_packages_contain_reviewer_workbooks_images_manifest_guide_readme_and_zip(tmp_path: Path) -> None:
    config, _ = make_config(tmp_path)

    assignments = build_pilot_review_packages(config)

    assert len(assignments) == 4
    for reviewer_id, reviewer_name, expected_ids in [
        ("vincent", "Vincent", ["review-001", "review-003"]),
        ("sister", "Reviewer_Sister", ["review-002", "review-004"]),
    ]:
        package_dir = config.output_root / "packages" / reviewer_id
        assert (package_dir / WORKBOOK_FILENAME).exists()
        assert (package_dir / MANIFEST_FILENAME).exists()
        assert (package_dir / PACKAGE_GUIDE_FILENAME).exists()
        assert (package_dir / README_FILENAME).exists()
        assert (config.output_root / "packages" / f"{reviewer_id}.zip").exists()
        manifest = pd.read_csv(package_dir / MANIFEST_FILENAME, dtype=str, keep_default_na=False, encoding="utf-8-sig")
        assert manifest["review_id"].tolist() == expected_ids
        workbook = load_workbook(package_dir / WORKBOOK_FILENAME)
        worksheet = workbook[WORKLIST_SHEET]
        workbook_review_ids = [worksheet.cell(row=row, column=EXCEL_COLUMNS.index("review_id") + 1).value for row in range(2, worksheet.max_row + 1)]
        assert workbook_review_ids == expected_ids
        assert {worksheet.cell(row=row, column=EXCEL_COLUMNS.index("human_reviewer") + 1).value for row in range(2, worksheet.max_row + 1)} == {reviewer_name}
        assert {worksheet.cell(row=row, column=EXCEL_COLUMNS.index("human_review_status") + 1).value for row in range(2, worksheet.max_row + 1)} == {"pending"}
        assert {worksheet.cell(row=row, column=EXCEL_COLUMNS.index("human_reviewed_at") + 1).value for row in range(2, worksheet.max_row + 1)} == {None}
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=1)
            formula = cell.value
            target = image_formula_path(formula)
            assert formula.startswith("=HYPERLINK(")
            assert 'CELL("filename",A1)' in formula
            assert cell.hyperlink is None
            assert target.startswith("images\\")
            assert not target.startswith("file:///")
            assert "G:\\" not in formula
            assert "file:///" not in formula
            assert "http://" not in formula
            assert "https://" not in formula
            assert ":" not in target
            assert (package_dir / target).exists()
        assert workbook.calculation.calcMode == "auto"
        assert workbook.calculation.fullCalcOnLoad is True
        assert workbook.calculation.forceFullCalc is True
        assert workbook.calculation.calcOnSave is True


def test_package_can_be_moved_or_unzipped_with_valid_relative_links(tmp_path: Path) -> None:
    config, _ = make_config(tmp_path)
    build_pilot_review_packages(config)
    extract_root = tmp_path / "unzipped"
    with zipfile.ZipFile(config.output_root / "packages" / "vincent.zip") as zip_file:
        zip_file.extractall(extract_root)
    workbook = load_workbook(extract_root / "vincent" / WORKBOOK_FILENAME)
    worksheet = workbook[WORKLIST_SHEET]
    for row in range(2, worksheet.max_row + 1):
        target = image_formula_path(worksheet.cell(row=row, column=1).value)
        assert (extract_root / "vincent" / target).exists()


def test_vincent_and_sister_open_image_formulas_are_excel_compatible(tmp_path: Path) -> None:
    config, _ = make_config(tmp_path)
    build_pilot_review_packages(config)

    for reviewer_id in ["vincent", "sister"]:
        workbook = load_workbook(config.output_root / "packages" / reviewer_id / WORKBOOK_FILENAME)
        worksheet = workbook[WORKLIST_SHEET]
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=1)
            formula = cell.value
            assert cell.data_type == "f"
            assert cell.hyperlink is None
            assert formula.startswith("=HYPERLINK(")
            assert 'LEFT(CELL("filename",A1),FIND("[",CELL("filename",A1))-1)' in formula
            assert "images\\" in formula
            assert "G:\\" not in formula
            assert "file:///" not in formula
            assert "http://" not in formula
            assert "https://" not in formula


def test_vincent_and_sister_workbooks_keep_human_dropdown_validations(tmp_path: Path) -> None:
    config, _ = make_config(tmp_path, row_count=500)
    build_pilot_review_packages(config)

    expected_ranges = {
        "human_photo_type_review": "T2:T251",
        "human_angle_review": "U2:U251",
        "human_is_exterior_review": "V2:V251",
        "human_has_visible_damage_review": "W2:W251",
        "human_severity_review": "X2:X251",
        "human_review_status": "Y2:Y251",
    }
    expected_defined_names = {f"package_{column_name}_options" for column_name in DROPDOWN_COLUMNS}

    for reviewer_id, reviewer_name in [("vincent", "Vincent"), ("sister", "Reviewer_Sister")]:
        workbook_path = config.output_root / "packages" / reviewer_id / WORKBOOK_FILENAME
        workbook = load_workbook(workbook_path)
        worksheet = workbook[WORKLIST_SHEET]
        validations = data_validation_by_range(worksheet)

        assert worksheet.max_row == 251
        assert {worksheet.cell(row=row, column=EXCEL_COLUMNS.index("human_reviewer") + 1).value for row in range(2, 252)} == {reviewer_name}
        assert {worksheet.cell(row=row, column=EXCEL_COLUMNS.index("human_review_status") + 1).value for row in range(2, 252)} == {"pending"}
        assert {worksheet.cell(row=row, column=EXCEL_COLUMNS.index("human_reviewed_at") + 1).value for row in range(2, 252)} == {None}
        assert workbook[OPTIONS_SHEET].sheet_state == "hidden"
        assert set(workbook.defined_names.keys()).issuperset(expected_defined_names)
        assert review_sheet_xml_contains_data_validations(workbook_path)
        assert len(worksheet.data_validations.dataValidation) == 6
        for column_name, target_range in expected_ranges.items():
            validation = validations[target_range]
            assert validation.type == "list"
            assert validation.allow_blank is True
            assert validation.formula1 == f"=package_{column_name}_options"


def test_missing_image_stops_without_partial_package(tmp_path: Path) -> None:
    config, worklist = make_config(tmp_path)
    missing_path = tmp_path / worklist.loc[0, "original_path"]
    missing_path.unlink()

    with pytest.raises(FileNotFoundError, match="image not found"):
        build_pilot_review_packages(config)

    assert not config.output_root.exists()
