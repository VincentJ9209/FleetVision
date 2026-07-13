from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

import pytest
from openpyxl import load_workbook

from fleetvision.data.validation_error_human_review import (
    SOURCE_COLUMNS,
    WORKBOOK_SHEETS,
    read_workbook_dataframe,
    sha256_file,
)
from fleetvision.review.validation_error_review_export import (
    CompletedWorkbookExportError,
    export_completed_workbook,
)
from fleetvision.review.validation_error_review_mapping import (
    ReviewSelection,
    derive_canonical_fields,
)
from fleetvision.review.validation_error_review_package import (
    load_review_app_config,
    load_verified_source_package,
)
from fleetvision.review.validation_error_review_state import (
    ReviewStateStore,
)
from review_app_fixtures import (
    create_review_package,
    write_app_config,
)


NOW = datetime(2026, 7, 14, 4, 0, tzinfo=timezone.utc)


def _prepared(tmp_path: Path):
    project_root, batch_root, frozen_zip = create_review_package(
        tmp_path
    )
    config_path = write_app_config(
        tmp_path,
        project_root,
        batch_root,
        frozen_zip,
    )
    config = load_review_app_config(config_path, project_root)
    package = load_verified_source_package(config)
    store = ReviewStateStore(
        config.workspace_root,
        backup_retention=20,
    )
    store.initialize(package)
    return package, store


def _save_model_miss(
    store: ReviewStateStore,
    review_case_id: str,
) -> None:
    selection = ReviewSelection(
        outcome="model_miss",
        reason="missed_small_damage",
        annotation_quality="correct",
        recommended_action="add_positive_sample",
        retraining_priority="medium",
    )
    canonical = derive_canonical_fields(
        selection,
        reviewer="Vincent",
        reviewed_at=NOW,
    )
    store.save_review(
        review_case_id,
        selection,
        canonical,
    )


def _save_ambiguous(
    store: ReviewStateStore,
    review_case_id: str,
) -> None:
    selection = ReviewSelection(
        outcome="ambiguous",
        reason="ambiguous_visual_evidence",
        annotation_quality="questionable",
        recommended_action="investigate_image_quality",
        retraining_priority="not_applicable",
        review_notes="影像證據不足。",
    )
    canonical = derive_canonical_fields(
        selection,
        reviewer="Vincent",
        reviewed_at=NOW,
    )
    store.save_review(
        review_case_id,
        selection,
        canonical,
    )


def test_export_is_blocked_while_pending_cases_exist(
    tmp_path: Path,
) -> None:
    package, store = _prepared(tmp_path)
    _save_model_miss(
        store,
        package.cases[0].review_case_id,
    )

    with pytest.raises(
        CompletedWorkbookExportError,
        match="every case",
    ):
        export_completed_workbook(package, store)


def test_export_is_blocked_while_adjudication_exists(
    tmp_path: Path,
) -> None:
    package, store = _prepared(tmp_path)
    _save_model_miss(
        store,
        package.cases[0].review_case_id,
    )
    _save_ambiguous(
        store,
        package.cases[1].review_case_id,
    )

    with pytest.raises(
        CompletedWorkbookExportError,
        match="needs_adjudication=1",
    ):
        export_completed_workbook(package, store)


def test_completed_workbook_preserves_source_contract(
    tmp_path: Path,
) -> None:
    package, store = _prepared(tmp_path)
    for case in package.cases:
        _save_model_miss(store, case.review_case_id)

    source_hash_before = sha256_file(package.workbook_path)
    source_frame = read_workbook_dataframe(
        package.workbook_path
    )
    result = export_completed_workbook(package, store)
    output_frame = read_workbook_dataframe(result.output_path)

    assert result.output_path.is_file()
    assert result.row_count == 2
    assert result.sha256 == sha256_file(result.output_path)
    assert result.backup_path.is_file()
    assert set(output_frame["review_status"]) == {"reviewed"}
    assert (
        source_frame.loc[:, list(SOURCE_COLUMNS)]
        .equals(output_frame.loc[:, list(SOURCE_COLUMNS)])
    )
    assert sha256_file(package.workbook_path) == source_hash_before

    workbook = load_workbook(result.output_path)
    try:
        assert tuple(workbook.sheetnames) == WORKBOOK_SHEETS
    finally:
        workbook.close()


def test_completed_workbook_refuses_overwrite(
    tmp_path: Path,
) -> None:
    package, store = _prepared(tmp_path)
    for case in package.cases:
        _save_model_miss(store, case.review_case_id)

    first = export_completed_workbook(package, store)
    with pytest.raises(
        CompletedWorkbookExportError,
        match="overwrite",
    ):
        export_completed_workbook(package, store)

    assert first.output_path.is_file()
