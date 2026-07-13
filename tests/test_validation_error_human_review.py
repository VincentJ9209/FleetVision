from __future__ import annotations

import json
import zipfile
from pathlib import Path

import pandas as pd
import pytest
import yaml
from openpyxl import load_workbook
from PIL import Image

import fleetvision.data.validation_error_human_review as review_module

from fleetvision.data.validation_error_human_review import (
    CANONICAL_COLUMNS,
    HumanReviewError,
    export_review_workbook,
    load_config,
    prepare_review_package,
    sha256_file,
    summarize_canonical_review,
    validate_canonical_csv,
)


def _write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(rows).to_csv(path, index=False, encoding="utf-8-sig")


def _make_source(tmp_path: Path, *, case_count: int = 2) -> tuple[Path, Path, Path, list[str]]:
    source_root = tmp_path / "source"
    images_dir = tmp_path / "valid"
    source_root.mkdir()
    images_dir.mkdir()
    image_ids: list[str] = []
    worklist_rows: list[dict[str, object]] = []
    prediction_rows: list[dict[str, object]] = []
    ground_truth_rows: list[dict[str, object]] = []

    for index in range(case_count):
        image_id = f"valid_{index:03d}.jpg"
        image_ids.append(image_id)
        Image.new("RGB", (320, 240), (100 + index, 110, 120)).save(images_dir / image_id)
        worklist_rows.append(
            {
                "image_id": image_id,
                "primary_error_type": "false_negative" if index == 0 else "background_false_positive",
                "error_types": "false_negative" if index == 0 else "background_false_positive|duplicate_prediction",
                "error_case_count": 2 if index == 0 else 1,
                "ground_truth_error_count": 1 if index == 0 else 0,
                "prediction_error_count": 1,
                "review_status": "pending",
                "human_error_category": "",
                "review_notes": "",
            }
        )
        prediction_rows.append(
            {
                "split": "valid",
                "image_id": image_id,
                "prediction_id": f"p_{index}",
                "class_id": 0,
                "confidence": 0.35 + index * 0.1,
                "x1": 30,
                "y1": 35,
                "x2": 130,
                "y2": 140,
                "bbox_area_ratio": 0.1,
                "matched_iou": 0.25,
            }
        )
        ground_truth_rows.append(
            {
                "split": "valid",
                "image_id": image_id,
                "gt_id": f"g_{index}",
                "class_id": 0,
                "x1": 45,
                "y1": 50,
                "x2": 145,
                "y2": 150,
                "bbox_area_ratio": 0.1,
            }
        )

    gate = {
        "gate_id": "04.5K",
        "outcome": "PASS",
        "classification": "VALIDATION_ERROR_ANALYSIS_AND_THRESHOLD_CANDIDATES_COMPLETED",
        "allowed_split": "valid",
        "validation_image_count": case_count,
        "validation_ground_truth_instances": case_count,
        "raw_prediction_count": case_count,
        "balanced_threshold": 0.20,
        "test_set_used_for_tuning": False,
        "test_set_read": False,
        "training_started": False,
        "annotation_modified": False,
        "deployment_acceptance": "NOT_YET_APPROVED",
    }
    gate_path = source_root / "04_5K_gate_result.json"
    worklist_path = source_root / "validation_error_review_worklist.csv"
    predictions_path = source_root / "validation_predictions.csv"
    ground_truth_path = source_root / "validation_ground_truth.csv"
    gate_path.write_text(json.dumps(gate), encoding="utf-8")
    _write_csv(worklist_path, worklist_rows)
    _write_csv(predictions_path, prediction_rows)
    _write_csv(ground_truth_path, ground_truth_rows)

    artifact_paths = [gate_path, worklist_path, predictions_path, ground_truth_path]
    manifest_path = source_root / "phase_04_5k_artifact_manifest.csv"
    _write_csv(
        manifest_path,
        [
            {
                "relative_path": path.name,
                "size_bytes": path.stat().st_size,
                "sha256": sha256_file(path),
            }
            for path in artifact_paths
        ],
    )
    checksum_path = source_root / "phase_04_5k_checksums.sha256"
    checksum_paths = artifact_paths + [manifest_path]
    checksum_path.write_text(
        "\n".join(f"{sha256_file(path)}  {path.name}" for path in checksum_paths) + "\n",
        encoding="ascii",
    )

    source_zip = tmp_path / "source.zip"
    with zipfile.ZipFile(source_zip, "w") as archive:
        for path in sorted(source_root.iterdir()):
            archive.write(path, arcname=path.name)
    return source_root, source_zip, images_dir, image_ids


def _refresh_source_integrity(source_root: Path, source_zip: Path) -> None:
    artifact_paths = [
        source_root / "04_5K_gate_result.json",
        source_root / "validation_error_review_worklist.csv",
        source_root / "validation_predictions.csv",
        source_root / "validation_ground_truth.csv",
    ]
    manifest_path = source_root / "phase_04_5k_artifact_manifest.csv"
    _write_csv(
        manifest_path,
        [
            {
                "relative_path": path.name,
                "size_bytes": path.stat().st_size,
                "sha256": sha256_file(path),
            }
            for path in artifact_paths
        ],
    )
    checksum_path = source_root / "phase_04_5k_checksums.sha256"
    checksum_paths = artifact_paths + [manifest_path]
    checksum_path.write_text(
        "\n".join(f"{sha256_file(path)}  {path.name}" for path in checksum_paths) + "\n",
        encoding="ascii",
    )
    with zipfile.ZipFile(source_zip, "w") as archive:
        for path in sorted(source_root.iterdir()):
            archive.write(path, arcname=path.name)


def _make_config(tmp_path: Path, source_zip: Path, *, case_count: int = 2) -> Path:
    config = {
        "schema_version": "1",
        "expected_source": {
            "zip_filename": source_zip.name,
            "zip_sha256": sha256_file(source_zip),
            "gate_classification": "VALIDATION_ERROR_ANALYSIS_AND_THRESHOLD_CANDIDATES_COMPLETED",
            "case_count": case_count,
            "validation_image_count": case_count,
            "validation_ground_truth_instances": case_count,
            "raw_prediction_count": case_count,
            "threshold_candidate": 0.20,
            "threshold_designation": "BALANCED_VALIDATION_THRESHOLD_CANDIDATE",
        },
        "source_files": {
            "gate_result": "04_5K_gate_result.json",
            "worklist": "validation_error_review_worklist.csv",
            "predictions": "validation_predictions.csv",
            "ground_truth": "validation_ground_truth.csv",
            "artifact_manifest": "phase_04_5k_artifact_manifest.csv",
            "checksums": "phase_04_5k_checksums.sha256",
        },
        "output": {"base_dir": "outputs/metadata/phase_04_5l"},
        "workbook": {
            "filename": "validation_error_human_review.xlsx",
            "thumbnail_width": 160,
            "thumbnail_height": 100,
        },
        "options": {
            "review_status": ["pending", "reviewed", "needs_adjudication"],
            "error_disposition": [
                "confirmed_model_error",
                "annotation_issue",
                "ambiguous_case",
                "expected_threshold_tradeoff",
                "invalid_review_case",
            ],
            "primary_root_cause": ["missed_small_damage", "annotation_missing", "other"],
            "secondary_root_cause": ["none", "other"],
            "annotation_quality": ["correct", "questionable", "defect_suspected", "not_applicable"],
            "annotation_defect_type": ["none", "missing_bbox", "other"],
            "recommended_action": [
                "no_action",
                "add_positive_sample",
                "create_annotation_correction_proposal",
                "other",
            ],
            "retraining_priority": ["not_applicable", "low", "medium", "high"],
            "correction_proposal_required": ["no", "yes"],
        },
    }
    config_path = tmp_path / "config.yaml"
    config_path.write_text(yaml.safe_dump(config, sort_keys=False), encoding="utf-8")
    return config_path


def _prepare(tmp_path: Path):
    source_root, source_zip, images_dir, image_ids = _make_source(tmp_path)
    config_path = _make_config(tmp_path, source_zip)
    config = load_config(config_path, tmp_path)
    result = prepare_review_package(
        config,
        source_root,
        source_zip,
        images_dir,
        "batch_001",
    )
    return config, result, image_ids


def _complete_workbook(workbook_path: Path, *, defect_second: bool = True) -> None:
    workbook = load_workbook(workbook_path)
    sheet = workbook["Review_Cases"]
    header_map = {cell.value: cell.column for cell in sheet[1]}
    for row_index in range(2, sheet.max_row + 1):
        values = {
            "review_status": "reviewed",
            "error_disposition": "confirmed_model_error",
            "primary_root_cause": "missed_small_damage",
            "secondary_root_cause": "none",
            "annotation_quality": "correct",
            "annotation_defect_type": "none",
            "recommended_action": "add_positive_sample",
            "retraining_priority": "medium",
            "correction_proposal_required": "no",
            "reviewer": "Vincent",
            "reviewed_at_utc": "2026-07-13T10:00:00+00:00",
            "review_notes": "Verified validation-only case.",
        }
        if defect_second and row_index == 3:
            values.update(
                {
                    "error_disposition": "annotation_issue",
                    "primary_root_cause": "annotation_missing",
                    "annotation_quality": "defect_suspected",
                    "annotation_defect_type": "missing_bbox",
                    "recommended_action": "create_annotation_correction_proposal",
                    "retraining_priority": "high",
                    "correction_proposal_required": "yes",
                    "review_notes": "Suspected missing GT; proposal only, not applied.",
                }
            )
        for column, value in values.items():
            sheet.cell(row=row_index, column=header_map[column], value=value)
    workbook.save(workbook_path)


def test_prepare_review_package_creates_contract_and_assets(tmp_path: Path) -> None:
    config, result, image_ids = _prepare(tmp_path)

    assert result.row_count == 2
    assert result.asset_count == 4
    assert result.workbook_path.is_file()
    assert sorted(path.name for path in (result.batch_root / "assets/original").iterdir()) == image_ids
    assert len(list((result.batch_root / "assets/overlay").glob("*.jpg"))) == 2

    workbook = load_workbook(result.workbook_path)
    assert workbook.sheetnames == [
        "Instructions",
        "Review_Cases",
        "Option_Lists",
        "Manifest",
        "Progress_Summary",
    ]
    assert len(workbook["Review_Cases"]._images) == 4
    assert workbook["Option_Lists"].sheet_state == "hidden"
    assert config.threshold_designation == "BALANCED_VALIDATION_THRESHOLD_CANDIDATE"


def test_prepare_is_fail_closed_on_zip_hash_mismatch(tmp_path: Path) -> None:
    source_root, source_zip, images_dir, _ = _make_source(tmp_path)
    config_path = _make_config(tmp_path, source_zip)
    raw = yaml.safe_load(config_path.read_text(encoding="utf-8"))
    raw["expected_source"]["zip_sha256"] = "0" * 64
    config_path.write_text(yaml.safe_dump(raw, sort_keys=False), encoding="utf-8")
    config = load_config(config_path, tmp_path)

    with pytest.raises(HumanReviewError, match="SHA256 mismatch"):
        prepare_review_package(config, source_root, source_zip, images_dir, "batch_001")
    assert not config.output_base_dir.exists()


def test_prepare_refuses_overwrite(tmp_path: Path) -> None:
    config, result, _ = _prepare(tmp_path)
    source_root = tmp_path / "source"
    source_zip = tmp_path / "source.zip"
    images_dir = tmp_path / "valid"

    with pytest.raises(HumanReviewError, match="overwrite is forbidden"):
        prepare_review_package(config, source_root, source_zip, images_dir, "batch_001")
    assert result.workbook_path.is_file()


def test_export_and_validate_completed_workbook(tmp_path: Path) -> None:
    config, result, _ = _prepare(tmp_path)
    _complete_workbook(result.workbook_path)
    output_csv = result.batch_root / "canonical/validation_error_human_review.csv"

    exported = export_review_workbook(config, result.workbook_path, output_csv)
    validation = validate_canonical_csv(
        config,
        output_csv,
        workbook_path=result.workbook_path,
        batch_root=result.batch_root,
    )

    assert exported.row_count == 2
    assert output_csv.is_file()
    assert validation.passed
    assert validation.issue_count == 0
    assert validation.counts["reviewed"] == 2
    assert pd.read_csv(output_csv, encoding="utf-8-sig").columns.tolist() == list(CANONICAL_COLUMNS)


def test_export_blocks_pending_rows(tmp_path: Path) -> None:
    config, result, _ = _prepare(tmp_path)
    output_csv = result.batch_root / "canonical/validation_error_human_review.csv"

    with pytest.raises(HumanReviewError, match="cannot be exported"):
        export_review_workbook(config, result.workbook_path, output_csv)
    assert not output_csv.exists()


def test_validator_detects_source_mutation_and_forbidden_test_path(tmp_path: Path) -> None:
    config, result, _ = _prepare(tmp_path)
    _complete_workbook(result.workbook_path)
    output_csv = result.batch_root / "canonical/validation_error_human_review.csv"
    export_review_workbook(config, result.workbook_path, output_csv)
    frame = pd.read_csv(output_csv, dtype=str, keep_default_na=False, encoding="utf-8-sig")
    frame.loc[0, "original_image_relpath"] = "assets/test/forbidden.jpg"
    frame.to_csv(output_csv, index=False, encoding="utf-8-sig")

    validation = validate_canonical_csv(config, output_csv, batch_root=result.batch_root)
    codes = {issue["error_code"] for issue in validation.issues}
    assert not validation.passed
    assert "SOURCE_FINGERPRINT_MISMATCH" in codes
    assert "FORBIDDEN_PATH" in codes


def test_summary_creates_non_applied_proposals_and_action_queue(tmp_path: Path) -> None:
    config, result, _ = _prepare(tmp_path)
    _complete_workbook(result.workbook_path)
    output_csv = result.batch_root / "canonical/validation_error_human_review.csv"
    export_review_workbook(config, result.workbook_path, output_csv)

    summary = summarize_canonical_review(config, output_csv, result.batch_root)

    assert summary.action_count == 2
    assert summary.correction_proposal_count == 1
    proposal = pd.read_csv(
        result.batch_root / "canonical/annotation_correction_proposals.csv",
        dtype=str,
        keep_default_na=False,
        encoding="utf-8-sig",
    )
    assert proposal.loc[0, "proposal_status"] == "PROPOSED_NOT_APPLIED"
    assert "x1" not in proposal.columns
    summary_json = json.loads(
        (result.batch_root / "reports/review_summary.json").read_text(encoding="utf-8")
    )
    assert summary_json["annotation_modified"] is False
    assert summary_json["test_set_read"] is False
    assert summary_json["training_started"] is False
    assert summary_json["deployment_acceptance"] == "NOT_YET_APPROVED"


def test_summary_refuses_overwrite(tmp_path: Path) -> None:
    config, result, _ = _prepare(tmp_path)
    _complete_workbook(result.workbook_path, defect_second=False)
    output_csv = result.batch_root / "canonical/validation_error_human_review.csv"
    export_review_workbook(config, result.workbook_path, output_csv)
    summarize_canonical_review(config, output_csv, result.batch_root)

    with pytest.raises(HumanReviewError, match="overwrite is forbidden"):
        summarize_canonical_review(config, output_csv, result.batch_root)


def test_prepare_supports_long_controlled_option_lists(tmp_path: Path) -> None:
    source_root, source_zip, images_dir, _ = _make_source(tmp_path)
    config_path = _make_config(tmp_path, source_zip)
    raw = yaml.safe_load(config_path.read_text(encoding="utf-8"))
    raw["options"]["primary_root_cause"] = [
        f"long_root_cause_option_{index:02d}_for_excel_validation"
        for index in range(12)
    ]
    config_path.write_text(yaml.safe_dump(raw, sort_keys=False), encoding="utf-8")
    config = load_config(config_path, tmp_path)

    result = prepare_review_package(
        config,
        source_root,
        source_zip,
        images_dir,
        "batch_long_options",
    )

    workbook = load_workbook(result.workbook_path)
    validations = list(workbook["Review_Cases"].data_validations.dataValidation)
    assert any(
        validation.formula1 == "=phase04_5l_option_primary_root_cause"
        for validation in validations
    )
    assert "phase04_5l_option_primary_root_cause" in workbook.defined_names


def test_prepare_blocks_source_artifact_that_does_not_match_verified_zip(tmp_path: Path) -> None:
    source_root, source_zip, images_dir, _ = _make_source(tmp_path)
    config_path = _make_config(tmp_path, source_zip)
    config = load_config(config_path, tmp_path)
    worklist_path = source_root / "validation_error_review_worklist.csv"
    worklist_path.write_text(worklist_path.read_text(encoding="utf-8-sig") + "\n", encoding="utf-8-sig")

    with pytest.raises(HumanReviewError, match="does not match verified ZIP"):
        prepare_review_package(
            config,
            source_root,
            source_zip,
            images_dir,
            "batch_source_mismatch",
        )


def test_prepare_blocks_validation_directory_nested_under_test_path(tmp_path: Path) -> None:
    source_root, source_zip, images_dir, _ = _make_source(tmp_path)
    test_parent = tmp_path / "test"
    test_parent.mkdir()
    forbidden_images_dir = test_parent / "valid"
    images_dir.rename(forbidden_images_dir)
    config_path = _make_config(tmp_path, source_zip)
    config = load_config(config_path, tmp_path)

    with pytest.raises(HumanReviewError, match="test-split directory is forbidden"):
        prepare_review_package(
            config,
            source_root,
            source_zip,
            forbidden_images_dir,
            "batch_test_path",
        )


def test_prepare_computes_best_iou_from_validation_predictions_and_gt(tmp_path: Path) -> None:
    _, result, _ = _prepare(tmp_path)
    workbook = load_workbook(result.workbook_path, data_only=True)
    sheet = workbook["Review_Cases"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    best_iou = float(sheet.cell(row=2, column=headers["best_iou"]).value)
    assert 0.5 < best_iou < 0.7


def test_prepare_blocks_non_validation_split_in_source_records(tmp_path: Path) -> None:
    source_root, source_zip, images_dir, _ = _make_source(tmp_path)
    predictions_path = source_root / "validation_predictions.csv"
    predictions = pd.read_csv(predictions_path, dtype=str, keep_default_na=False, encoding="utf-8-sig")
    predictions.loc[0, "split"] = "test"
    predictions.to_csv(predictions_path, index=False, encoding="utf-8-sig")
    _refresh_source_integrity(source_root, source_zip)
    config_path = _make_config(tmp_path, source_zip)
    config = load_config(config_path, tmp_path)

    with pytest.raises(HumanReviewError, match="split must contain only valid"):
        prepare_review_package(
            config,
            source_root,
            source_zip,
            images_dir,
            "batch_forbidden_split",
        )


def test_prepare_blocks_unsafe_04_5k_gate_boundary(tmp_path: Path) -> None:
    source_root, source_zip, images_dir, _ = _make_source(tmp_path)
    gate_path = source_root / "04_5K_gate_result.json"
    gate = json.loads(gate_path.read_text(encoding="utf-8"))
    gate["test_set_read"] = True
    gate_path.write_text(json.dumps(gate), encoding="utf-8")
    _refresh_source_integrity(source_root, source_zip)
    config_path = _make_config(tmp_path, source_zip)
    config = load_config(config_path, tmp_path)

    with pytest.raises(HumanReviewError, match="04.5K gate boundary mismatch"):
        prepare_review_package(
            config,
            source_root,
            source_zip,
            images_dir,
            "batch_unsafe_gate",
        )


def test_prepare_blocks_invalid_04_5k_artifact_manifest(tmp_path: Path) -> None:
    source_root, source_zip, images_dir, _ = _make_source(tmp_path)
    manifest_path = source_root / "phase_04_5k_artifact_manifest.csv"
    manifest = pd.read_csv(manifest_path, dtype=str, keep_default_na=False, encoding="utf-8-sig")
    manifest.loc[manifest["relative_path"] == "validation_predictions.csv", "sha256"] = "0" * 64
    manifest.to_csv(manifest_path, index=False, encoding="utf-8-sig")
    checksum_path = source_root / "phase_04_5k_checksums.sha256"
    checksum_members = [
        source_root / "04_5K_gate_result.json",
        source_root / "validation_error_review_worklist.csv",
        source_root / "validation_predictions.csv",
        source_root / "validation_ground_truth.csv",
        manifest_path,
    ]
    checksum_path.write_text(
        "\n".join(f"{sha256_file(path)}  {path.name}" for path in checksum_members) + "\n",
        encoding="ascii",
    )
    with zipfile.ZipFile(source_zip, "w") as archive:
        for path in sorted(source_root.iterdir()):
            archive.write(path, arcname=path.name)
    config_path = _make_config(tmp_path, source_zip)
    config = load_config(config_path, tmp_path)

    with pytest.raises(HumanReviewError, match="04.5K artifact manifest mismatch"):
        prepare_review_package(
            config,
            source_root,
            source_zip,
            images_dir,
            "batch_bad_manifest",
        )


def test_validator_protects_review_case_id_with_source_fingerprint(tmp_path: Path) -> None:
    config, result, _ = _prepare(tmp_path)
    _complete_workbook(result.workbook_path)
    output_csv = result.batch_root / "canonical/validation_error_human_review.csv"
    export_review_workbook(config, result.workbook_path, output_csv)
    frame = pd.read_csv(output_csv, dtype=str, keep_default_na=False, encoding="utf-8-sig")
    frame.loc[0, "review_case_id"] = "l_manually_changed"
    frame.to_csv(output_csv, index=False, encoding="utf-8-sig")

    validation = validate_canonical_csv(config, output_csv, batch_root=result.batch_root)
    codes = {issue["error_code"] for issue in validation.issues}
    assert "SOURCE_FINGERPRINT_MISMATCH" in codes


def test_validation_report_pair_rolls_back_on_second_write_failure(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    config, result, _ = _prepare(tmp_path)
    _complete_workbook(result.workbook_path)
    output_csv = result.batch_root / "canonical/validation_error_human_review.csv"
    export_review_workbook(config, result.workbook_path, output_csv)
    validation = validate_canonical_csv(config, output_csv, batch_root=result.batch_root)
    report_json = tmp_path / "reports/validation_report.json"
    errors_csv = tmp_path / "reports/validation_errors.csv"

    def fail_csv(*args, **kwargs):
        raise OSError("simulated CSV write failure")

    monkeypatch.setattr(review_module, "_write_csv", fail_csv)
    with pytest.raises(OSError, match="simulated CSV write failure"):
        review_module._write_validation_outputs(validation, report_json, errors_csv)

    assert not report_json.exists()
    assert not errors_csv.exists()


def test_summary_promotion_rolls_back_all_outputs_on_failure(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    config, result, _ = _prepare(tmp_path)
    _complete_workbook(result.workbook_path)
    output_csv = result.batch_root / "canonical/validation_error_human_review.csv"
    export_review_workbook(config, result.workbook_path, output_csv)

    original_replace = Path.replace
    promotion_calls = 0

    def fail_second_promotion(self: Path, target: Path):
        nonlocal promotion_calls
        if ".summary.staging-" in str(self):
            promotion_calls += 1
            if promotion_calls == 2:
                raise OSError("simulated promotion failure")
        return original_replace(self, target)

    monkeypatch.setattr(Path, "replace", fail_second_promotion)
    with pytest.raises(OSError, match="simulated promotion failure"):
        summarize_canonical_review(config, output_csv, result.batch_root)

    expected = [
        result.batch_root / "canonical/annotation_correction_proposals.csv",
        result.batch_root / "reports/review_summary.json",
        result.batch_root / "reports/review_summary.md",
        result.batch_root / "reports/data_improvement_action_queue.csv",
        result.batch_root / "reports/data_improvement_action_summary.csv",
    ]
    assert all(not path.exists() for path in expected)



def test_validator_rejects_inconsistent_annotation_proposal_semantics(tmp_path: Path) -> None:
    config, result, _ = _prepare(tmp_path)
    _complete_workbook(result.workbook_path, defect_second=False)
    output_csv = result.batch_root / "canonical/validation_error_human_review.csv"
    export_review_workbook(config, result.workbook_path, output_csv)
    frame = pd.read_csv(output_csv, dtype=str, keep_default_na=False, encoding="utf-8-sig")
    frame.loc[0, "annotation_quality"] = "questionable"
    frame.loc[0, "annotation_defect_type"] = "missing_bbox"
    frame.loc[0, "recommended_action"] = "create_annotation_correction_proposal"
    frame.loc[0, "correction_proposal_required"] = "yes"
    frame.to_csv(output_csv, index=False, encoding="utf-8-sig")

    validation = validate_canonical_csv(config, output_csv, batch_root=result.batch_root)
    codes = {issue["error_code"] for issue in validation.issues}
    assert "CONTRADICTORY_CORRECTION" in codes
    assert "CONTRADICTORY_CORRECTION_ACTION" in codes
    assert "CONTRADICTORY_DEFECT_TYPE" in codes


def test_validator_requires_specific_defect_type_for_suspected_defect(tmp_path: Path) -> None:
    config, result, _ = _prepare(tmp_path)
    _complete_workbook(result.workbook_path, defect_second=False)
    output_csv = result.batch_root / "canonical/validation_error_human_review.csv"
    export_review_workbook(config, result.workbook_path, output_csv)
    frame = pd.read_csv(output_csv, dtype=str, keep_default_na=False, encoding="utf-8-sig")
    frame.loc[0, "annotation_quality"] = "defect_suspected"
    frame.loc[0, "annotation_defect_type"] = "none"
    frame.loc[0, "recommended_action"] = "create_annotation_correction_proposal"
    frame.loc[0, "correction_proposal_required"] = "yes"
    frame.loc[0, "review_notes"] = "Suspected defect but type was not selected."
    frame.to_csv(output_csv, index=False, encoding="utf-8-sig")

    validation = validate_canonical_csv(config, output_csv, batch_root=result.batch_root)
    codes = {issue["error_code"] for issue in validation.issues}
    assert "CORRECTION_EVIDENCE_REQUIRED" in codes


def test_prepare_supports_production_nested_04_5k_layout(tmp_path: Path) -> None:
    source_root, source_zip, images_dir, _ = _make_source(tmp_path)
    nested_paths = {
        "worklist": Path("review/validation_error_review_worklist.csv"),
        "predictions": Path("records/validation_predictions.csv"),
        "ground_truth": Path("records/validation_ground_truth.csv"),
        "artifact_manifest": Path("manifest/phase_04_5k_artifact_manifest.csv"),
        "checksums": Path("manifest/phase_04_5k_checksums.sha256"),
    }
    original_names = {
        "worklist": "validation_error_review_worklist.csv",
        "predictions": "validation_predictions.csv",
        "ground_truth": "validation_ground_truth.csv",
        "artifact_manifest": "phase_04_5k_artifact_manifest.csv",
        "checksums": "phase_04_5k_checksums.sha256",
    }
    for key, relative in nested_paths.items():
        target = source_root / relative
        target.parent.mkdir(parents=True, exist_ok=True)
        (source_root / original_names[key]).replace(target)

    gate_path = source_root / "04_5K_gate_result.json"
    worklist_path = source_root / nested_paths["worklist"]
    predictions_path = source_root / nested_paths["predictions"]
    ground_truth_path = source_root / nested_paths["ground_truth"]
    manifest_path = source_root / nested_paths["artifact_manifest"]
    checksum_path = source_root / nested_paths["checksums"]
    core = [gate_path, worklist_path, predictions_path, ground_truth_path]
    _write_csv(
        manifest_path,
        [
            {
                "relative_path": path.relative_to(source_root).as_posix(),
                "size_bytes": path.stat().st_size,
                "sha256": sha256_file(path),
            }
            for path in core
        ],
    )
    checksum_path.write_text(
        "\n".join(
            f"{sha256_file(path)}  {path.relative_to(source_root).as_posix()}"
            for path in core + [manifest_path]
        )
        + "\n",
        encoding="ascii",
    )
    with zipfile.ZipFile(source_zip, "w") as archive:
        for path in sorted(item for item in source_root.rglob("*") if item.is_file()):
            archive.write(path, arcname=path.relative_to(source_root).as_posix())

    config_path = _make_config(tmp_path, source_zip)
    raw = yaml.safe_load(config_path.read_text(encoding="utf-8"))
    raw["source_files"] = {
        "gate_result": "04_5K_gate_result.json",
        **{key: relative.as_posix() for key, relative in nested_paths.items()},
    }
    raw["expected_source"]["zip_sha256"] = sha256_file(source_zip)
    config_path.write_text(yaml.safe_dump(raw, sort_keys=False), encoding="utf-8")
    config = load_config(config_path, tmp_path)

    result = prepare_review_package(
        config,
        source_root,
        source_zip,
        images_dir,
        "batch_nested_layout",
    )
    assert result.row_count == 2
    assert result.asset_count == 4



def test_dataframe_validator_reports_column_contract_before_reindex(tmp_path: Path) -> None:
    config, result, _ = _prepare(tmp_path)
    _complete_workbook(result.workbook_path, defect_second=False)
    output_csv = result.batch_root / "canonical/validation_error_human_review.csv"
    export_review_workbook(config, result.workbook_path, output_csv)
    frame = pd.read_csv(output_csv, dtype=str, keep_default_na=False, encoding="utf-8-sig")
    reordered = frame[[frame.columns[1], frame.columns[0], *frame.columns[2:]]]

    validation = review_module.validate_canonical_dataframe(reordered, config)
    codes = {issue["error_code"] for issue in validation.issues}
    assert "CANONICAL_COLUMN_CONTRACT_MISMATCH" in codes
