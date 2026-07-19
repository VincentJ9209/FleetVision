from __future__ import annotations

import csv
import importlib
import json
import os
import subprocess
import sys
from dataclasses import replace
from datetime import datetime, timezone
from pathlib import Path

import pytest
from openpyxl import load_workbook

from team_pairing_audit_fixtures import create_rgb_image, create_test_project
from fleetvision.data.team_pairing_audit import (
    build_source_snapshot,
    sha256_file,
    verify_source_snapshots,
)
from fleetvision.review.team_pairing_review_app import create_runtime
from fleetvision.review.team_pairing_review_mapping import (
    AngleReviewSelection,
    BatchReviewSelection,
    PairReviewSelection,
    derive_canonical_angle_fields,
    derive_canonical_batch_fields,
    derive_canonical_pair_fields,
    load_team_pairing_audit_config,
)
from fleetvision.review.team_pairing_review_state import (
    BatchMemberSeed,
    CandidateBatchSeed,
    PairCandidateSeed,
    SourceImageSeed,
    TeamPairingCandidatePackage,
    TeamPairingReviewStateStore,
    TeamPairingWorkspaceIdentity,
)


def exporter():
    return importlib.import_module("fleetvision.review.team_pairing_review_export")


def _write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _runtime(tmp_path: Path):
    project_root, source_root, config_path = create_test_project(tmp_path)
    config = load_team_pairing_audit_config(config_path, project_root)
    workspace = config.output_root / "workspaces" / "team_pairing_audit_test"
    candidates = workspace / "candidates"
    source_dir = workspace / "source"
    candidates.mkdir(parents=True)
    source_dir.mkdir(parents=True)

    images = []
    inventory_rows: list[dict[str, object]] = []
    batch_rows: list[dict[str, object]] = []
    member_rows: list[dict[str, object]] = []
    batches = []
    members = []

    for index in range(1, 7):
        image_id = f"team_{index:03d}"
        batch_id = f"batch_{index:03d}"
        relative_path = f"image_{index:03d}.jpg"
        create_rgb_image(source_root / relative_path, color=(20 + index, 70, 110))
        images.append(SourceImageSeed(image_id, index, relative_path, True))
        batches.append(
            CandidateBatchSeed(
                batch_id,
                index,
                f"2026-07-19T{index:02d}:00:00+00:00",
                f"2026-07-19T{index:02d}:05:00+00:00",
            )
        )
        members.append(BatchMemberSeed(batch_id, image_id, 1))
        inventory_rows.append(
            {
                "inventory_sequence": index,
                "image_id": image_id,
                "relative_path": relative_path,
                "original_path": f"dataset/01_raw/04_team/{relative_path}",
                "filename": relative_path,
                "is_readable": "True",
                "sha256": sha256_file(source_root / relative_path),
            }
        )
        batch_rows.append(
            {
                "batch_sequence": index,
                "batch_id": batch_id,
                "start_time": f"2026-07-19T{index:02d}:00:00+00:00",
                "end_time": f"2026-07-19T{index:02d}:05:00+00:00",
                "image_count": 1,
            }
        )
        member_rows.append(
            {
                "batch_id": batch_id,
                "image_id": image_id,
                "member_sequence": 1,
                "membership_role": "candidate_representative",
            }
        )

    pairs = (
        PairCandidateSeed("pair_001", 1, "batch_001", "batch_002", "TEAMCAR-001", 3600, '["front_left"]', 1, 1),
        PairCandidateSeed("pair_002", 2, "batch_003", "batch_004", "TEAMCAR-002", 3600, '["front_left"]', 1, 1),
        PairCandidateSeed("pair_003", 3, "batch_005", "batch_006", "TEAMCAR-003", 3600, '["front_left"]', 1, 1),
    )

    inventory_path = candidates / "team_image_inventory.csv"
    batch_path = candidates / "team_capture_batch_candidates.csv"
    member_path = candidates / "team_capture_batch_members.csv"
    pair_path = candidates / "team_before_after_pair_candidates.csv"
    _write_csv(inventory_path, list(inventory_rows[0]), inventory_rows)
    _write_csv(batch_path, list(batch_rows[0]), batch_rows)
    _write_csv(member_path, list(member_rows[0]), member_rows)
    _write_csv(
        pair_path,
        [
            "pair_candidate_id",
            "pair_sequence",
            "before_batch_id",
            "after_batch_id",
            "manual_vehicle_id",
            "elapsed_seconds",
            "overlap_angles_json",
            "overlap_count",
            "four_angle_overlap_count",
        ],
        [
            {
                "pair_candidate_id": item.pair_candidate_id,
                "pair_sequence": item.pair_sequence,
                "before_batch_id": item.before_batch_id,
                "after_batch_id": item.after_batch_id,
                "manual_vehicle_id": item.manual_vehicle_id,
                "elapsed_seconds": item.elapsed_seconds,
                "overlap_angles_json": item.overlap_angles_json,
                "overlap_count": item.overlap_count,
                "four_angle_overlap_count": item.four_angle_overlap_count,
            }
            for item in pairs
        ],
    )

    before = build_source_snapshot(source_root)
    after = build_source_snapshot(source_root)
    verification = verify_source_snapshots(before, after)
    (source_dir / "source_snapshot_before.json").write_text(
        json.dumps(before, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8"
    )
    (source_dir / "source_snapshot_after.json").write_text(
        json.dumps(after, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8"
    )
    (source_dir / "source_snapshot_verification.json").write_text(
        json.dumps(verification, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8"
    )

    manifest = {
        "inventory_sha256": sha256_file(inventory_path),
        "batch_candidates_sha256": sha256_file(batch_path),
        "batch_members_sha256": sha256_file(member_path),
        "config_sha256": sha256_file(config_path),
        "expected_image_count": 6,
        "expected_batch_count": 6,
        "expected_pair_count": 3,
    }
    manifest_path = candidates / "candidate_manifest.json"
    manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8"
    )

    identity = TeamPairingWorkspaceIdentity(
        schema_version="1",
        project_root=str(project_root.resolve()),
        source_root=str(source_root.resolve()),
        candidate_manifest_sha256=sha256_file(manifest_path),
        inventory_sha256=manifest["inventory_sha256"],
        batch_candidates_sha256=manifest["batch_candidates_sha256"],
        batch_members_sha256=manifest["batch_members_sha256"],
        config_sha256=manifest["config_sha256"],
        reviewer="Vincent",
        timezone="Asia/Taipei",
        expected_image_count=6,
        expected_batch_count=6,
        expected_pair_count=3,
    )
    package = TeamPairingCandidatePackage(
        workspace,
        identity,
        tuple(images),
        tuple(batches),
        tuple(members),
        pairs,
    )
    store = TeamPairingReviewStateStore(
        workspace,
        identity=identity,
        backup_every_successful_saves=10,
        backup_retention=20,
    )
    roles = {(item.batch_id, item.image_id): "candidate_representative" for item in members}
    runtime = create_runtime(config, package, store=store, member_roles=roles)
    return runtime


def _complete(runtime, *, primary_pair: str | None = "pair_001", rejected_pair: str | None = None) -> None:
    reviewed_at = datetime(2026, 7, 19, 12, 0, tzinfo=timezone.utc)
    for index, batch in enumerate(runtime.package.batches, start=1):
        stage = "before" if index % 2 else "after"
        vehicle_id = f"TEAMCAR-{(index + 1) // 2:03d}"
        selection = BatchReviewSelection("confirmed", vehicle_id, stage, "")
        runtime.store.save_batch_review(
            batch.batch_id,
            selection,
            derive_canonical_batch_fields(selection, reviewer="Vincent", reviewed_at=reviewed_at),
        )
    for image in runtime.package.images:
        selection = AngleReviewSelection("reviewed", "front_left", "")
        runtime.store.save_image_review(
            image.image_id,
            selection,
            derive_canonical_angle_fields(selection, reviewer="Vincent", reviewed_at=reviewed_at),
        )
    for pair in runtime.package.pairs:
        status = "rejected" if pair.pair_candidate_id == rejected_pair else "confirmed"
        role = "none"
        if status == "confirmed" and pair.pair_candidate_id == primary_pair:
            role = "primary"
        elif status == "confirmed" and pair.pair_candidate_id == "pair_003" and primary_pair is not None:
            role = "backup"
        notes = "人工拒絕配對" if status == "rejected" else ""
        selection = PairReviewSelection(status, "no", "none", role, notes)
        runtime.store.save_pair_review(
            pair.pair_candidate_id,
            selection,
            derive_canonical_pair_fields(selection, reviewer="Vincent", reviewed_at=reviewed_at),
        )


def test_incomplete_export_is_blocked(tmp_path: Path) -> None:
    module = exporter()
    runtime = _runtime(tmp_path)
    with pytest.raises(module.TeamPairingReviewExportError, match="pending|未完成"):
        module.export_completed_team_pairing_review(runtime, timestamp="20260719_120000")


def test_fewer_than_three_confirmed_pairs_is_blocked(tmp_path: Path) -> None:
    module = exporter()
    runtime = _runtime(tmp_path)
    _complete(runtime, rejected_pair="pair_003")
    with pytest.raises(module.TeamPairingReviewExportError, match="3"):
        module.export_completed_team_pairing_review(runtime, timestamp="20260719_120000")


def test_primary_demo_pair_must_be_exactly_one_and_reliable() -> None:
    module = exporter()
    rows = [
        {"pair_candidate_id": "pair_1", "manual_demo_role": "none", "derived_case_classification": "NO_NEW_DAMAGE"},
        {"pair_candidate_id": "pair_2", "manual_demo_role": "none", "derived_case_classification": "NO_NEW_DAMAGE"},
    ]
    with pytest.raises(module.TeamPairingReviewExportError, match="primary"):
        module.validate_primary_demo_rows(rows)
    duplicate = [dict(rows[0], manual_demo_role="primary"), dict(rows[1], manual_demo_role="primary")]
    with pytest.raises(module.TeamPairingReviewExportError, match="primary"):
        module.validate_primary_demo_rows(duplicate)
    unreliable = [dict(rows[0], manual_demo_role="primary", derived_case_classification="NEW_DAMAGE_CANDIDATE")]
    with pytest.raises(module.TeamPairingReviewExportError, match="classification"):
        module.validate_primary_demo_rows(unreliable)


def test_source_snapshot_mismatch_is_blocked(tmp_path: Path) -> None:
    module = exporter()
    runtime = _runtime(tmp_path)
    _complete(runtime)
    (runtime.config.source_root / "image_001.jpg").write_bytes(b"changed")
    with pytest.raises(module.TeamPairingReviewExportError, match="snapshot|source"):
        module.export_completed_team_pairing_review(runtime, timestamp="20260719_120000")


def test_completed_export_creates_exact_artifacts_and_is_no_overwrite(tmp_path: Path) -> None:
    module = exporter()
    runtime = _runtime(tmp_path)
    _complete(runtime)
    result = module.export_completed_team_pairing_review(
        runtime,
        timestamp="20260719_120000",
        repository_commit="583ec82c7fe9bdb600bf4d85899588ba7a6ec399",
    )
    expected = {
        "team_image_inventory.csv",
        "team_capture_batch_candidates.csv",
        "team_image_reviews_completed.csv",
        "team_before_after_pair_candidates.csv",
        "team_pair_review_completed.xlsx",
        "team_pairing_summary.json",
        "SHA256SUMS.csv",
    }
    assert {path.name for path in result.export_root.iterdir()} == expected
    with pytest.raises(module.TeamPairingReviewExportError, match="overwrite"):
        module.export_completed_team_pairing_review(runtime, timestamp="20260719_120000")


def test_export_failure_cleans_staging_and_partial_output(tmp_path: Path) -> None:
    module = exporter()
    runtime = _runtime(tmp_path)
    _complete(runtime)
    with pytest.raises(module.TeamPairingReviewExportError, match="simulated"):
        module.export_completed_team_pairing_review(
            runtime,
            timestamp="20260719_120000",
            simulate_failure_after_first_artifact=True,
        )
    exports = runtime.package.workspace_root / "exports"
    assert not list(exports.glob("completed_*"))
    assert not list(exports.glob(".*staging*"))


def test_csv_order_checksum_manifest_and_source_fields_are_deterministic(tmp_path: Path) -> None:
    module = exporter()
    runtime = _runtime(tmp_path)
    source_inventory = (runtime.package.workspace_root / "candidates" / "team_image_inventory.csv").read_bytes()
    source_batches = (runtime.package.workspace_root / "candidates" / "team_capture_batch_candidates.csv").read_bytes()
    _complete(runtime)
    result = module.export_completed_team_pairing_review(runtime, timestamp="20260719_120000")
    assert result.inventory_csv.read_bytes() == source_inventory
    assert result.batch_candidates_csv.read_bytes() == source_batches
    with result.image_reviews_csv.open(newline="", encoding="utf-8") as handle:
        image_rows = list(csv.DictReader(handle))
    with result.pair_candidates_csv.open(newline="", encoding="utf-8") as handle:
        pair_rows = list(csv.DictReader(handle))
    assert [int(row["inventory_sequence"]) for row in image_rows] == list(range(1, 7))
    assert [int(row["pair_sequence"]) for row in pair_rows] == [1, 2, 3]
    with result.checksum_manifest.open(newline="", encoding="utf-8-sig") as handle:
        checksum_rows = list(csv.DictReader(handle))
    assert [row["relative_path"] for row in checksum_rows] == sorted(row["relative_path"] for row in checksum_rows)
    for row in checksum_rows:
        path = result.export_root / row["relative_path"]
        assert int(row["size_bytes"]) == path.stat().st_size
        assert row["sha256"] == sha256_file(path)


def test_workbook_sheet_contract_summary_counts_and_cli_help(tmp_path: Path) -> None:
    module = exporter()
    runtime = _runtime(tmp_path)
    _complete(runtime)
    result = module.export_completed_team_pairing_review(runtime, timestamp="20260719_120000")
    workbook = load_workbook(result.workbook, read_only=True, data_only=True)
    try:
        assert tuple(workbook.sheetnames) == module.WORKBOOK_SHEETS
        assert workbook["圖片清單"].max_row == 7
        assert workbook["候選批次"].max_row == 7
        assert workbook["批次成員"].max_row == 7
        assert workbook["圖片角度"].max_row == 7
        assert workbook["配對候選"].max_row == 4
        assert workbook["確認案例"].max_row == 4
    finally:
        workbook.close()
    summary = json.loads(result.summary_json.read_text(encoding="utf-8"))
    assert summary["confirmed_pair_count"] == 3
    assert summary["primary_demo_pair_id"] == "pair_001"
    assert summary["source_snapshot_verification"]["byte_identical"] is True
    script = Path(__file__).resolve().parents[1] / "scripts" / "phase05s_export_team_pairing_review.py"
    clean_env = dict(os.environ)
    clean_env.pop("PYTHONPATH", None)
    completed = subprocess.run(
        [sys.executable, str(script), "--help"],
        cwd=Path(__file__).resolve().parents[1],
        env=clean_env,
        capture_output=True,
        text=True,
        check=False,
    )
    assert completed.returncode == 0, completed.stderr
    assert "--workspace-root" in completed.stdout
