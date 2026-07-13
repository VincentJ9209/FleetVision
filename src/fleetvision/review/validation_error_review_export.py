from __future__ import annotations

import shutil
import uuid
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from fleetvision.data.validation_error_human_review import (
    HUMAN_COLUMNS,
    SOURCE_COLUMNS,
    WORKBOOK_SHEETS,
    load_config,
    read_workbook_dataframe,
    sha256_file,
    validate_canonical_dataframe,
)
from fleetvision.review.validation_error_review_package import (
    VerifiedSourcePackage,
)
from fleetvision.review.validation_error_review_state import (
    ReviewStateStore,
)


class CompletedWorkbookExportError(RuntimeError):
    """Raised when completed Workbook export is incomplete or unsafe."""


@dataclass(frozen=True)
class CompletedWorkbookExport:
    output_path: Path
    sha256: str
    row_count: int
    logical_fingerprint: str
    backup_path: Path
    exported_at_utc: str


def _embedded_image_count(workbook_path: Path) -> int:
    workbook = load_workbook(
        workbook_path,
        read_only=False,
        data_only=False,
    )
    try:
        return sum(
            len(getattr(sheet, "_images", ()))
            for sheet in workbook.worksheets
        )
    finally:
        workbook.close()


def _require_complete_state(
    package: VerifiedSourcePackage,
    store: ReviewStateStore,
) -> None:
    progress = store.progress()
    if progress.total != len(package.cases):
        raise CompletedWorkbookExportError(
            "state/package case count mismatch: "
            f"state={progress.total} package={len(package.cases)}"
        )
    if (
        progress.reviewed != progress.total
        or progress.pending != 0
        or progress.needs_adjudication != 0
    ):
        raise CompletedWorkbookExportError(
            "completed Workbook requires every case to be reviewed; "
            f"reviewed={progress.reviewed}/{progress.total} "
            f"pending={progress.pending} "
            f"needs_adjudication={progress.needs_adjudication}"
        )


def export_completed_workbook(
    package: VerifiedSourcePackage,
    store: ReviewStateStore,
) -> CompletedWorkbookExport:
    """Export a no-overwrite completed Workbook from trusted local state."""

    _require_complete_state(package, store)

    source_workbook_hash = sha256_file(package.workbook_path)
    if source_workbook_hash != package.config.workbook_sha256:
        raise CompletedWorkbookExportError(
            "source Workbook SHA256 changed before export"
        )
    if (
        not package.config.frozen_zip_path.is_file()
        or sha256_file(package.config.frozen_zip_path)
        != package.config.frozen_zip_sha256
    ):
        raise CompletedWorkbookExportError(
            "frozen package ZIP SHA256 changed before export"
        )

    store.export_dir.mkdir(parents=True, exist_ok=True)
    output_path = (
        store.export_dir
        / package.config.completed_workbook_name
    )
    if output_path.exists():
        raise CompletedWorkbookExportError(
            f"completed Workbook overwrite is forbidden: {output_path}"
        )

    token = uuid.uuid4().hex[:12]
    staging_path = (
        store.export_dir
        / f".{output_path.name}.staging-{token}.xlsx"
    )
    if staging_path.exists():
        raise CompletedWorkbookExportError(
            f"unexpected staging collision: {staging_path}"
        )

    source_frame = read_workbook_dataframe(package.workbook_path)
    source_image_count = _embedded_image_count(package.workbook_path)
    backup_path = store.create_backup()

    shutil.copy2(package.workbook_path, staging_path)
    try:
        workbook = load_workbook(
            staging_path,
            read_only=False,
            data_only=False,
        )
        try:
            if tuple(workbook.sheetnames) != WORKBOOK_SHEETS:
                raise CompletedWorkbookExportError(
                    "source Workbook sheet contract changed"
                )

            sheet = workbook["Review_Cases"]
            headers = {
                str(cell.value): cell.column
                for cell in sheet[1]
            }
            missing_headers = [
                column
                for column in HUMAN_COLUMNS
                if column not in headers
            ]
            if missing_headers:
                raise CompletedWorkbookExportError(
                    "Workbook missing human-review columns: "
                    f"{missing_headers}"
                )
            if "review_case_id" not in headers:
                raise CompletedWorkbookExportError(
                    "Workbook missing review_case_id"
                )

            row_by_case_id: dict[str, int] = {}
            for row_index in range(2, sheet.max_row + 1):
                review_case_id = str(
                    sheet.cell(
                        row=row_index,
                        column=headers["review_case_id"],
                    ).value
                    or ""
                ).strip()
                if not review_case_id:
                    continue
                if review_case_id in row_by_case_id:
                    raise CompletedWorkbookExportError(
                        "duplicate review_case_id in Workbook: "
                        f"{review_case_id}"
                    )
                row_by_case_id[review_case_id] = row_index

            expected_case_ids = {
                case.review_case_id
                for case in package.cases
            }
            if set(row_by_case_id) != expected_case_ids:
                raise CompletedWorkbookExportError(
                    "Workbook/state review_case_id set mismatch"
                )

            for case in package.cases:
                stored = store.get_review(case.review_case_id)
                if stored is None:
                    raise CompletedWorkbookExportError(
                        "stored review unexpectedly missing: "
                        f"{case.review_case_id}"
                    )
                canonical = dict(stored.canonical_fields)
                if set(canonical) != set(HUMAN_COLUMNS):
                    raise CompletedWorkbookExportError(
                        "stored canonical human-field contract mismatch: "
                        f"{case.review_case_id}"
                    )
                if canonical["review_status"] != "reviewed":
                    raise CompletedWorkbookExportError(
                        "stored review is not final reviewed status: "
                        f"{case.review_case_id}"
                    )

                row_index = row_by_case_id[case.review_case_id]
                for column in HUMAN_COLUMNS:
                    sheet.cell(
                        row=row_index,
                        column=headers[column],
                        value=canonical[column],
                    )

            workbook.save(staging_path)
        finally:
            workbook.close()

        output_frame = read_workbook_dataframe(staging_path)
        canonical_config = load_config(
            package.config.canonical_config_path,
            package.config.project_root,
        )
        validation = validate_canonical_dataframe(
            output_frame,
            canonical_config,
            require_complete=True,
            batch_root=package.batch_root,
        )
        if not validation.passed:
            preview = "; ".join(
                issue["error_code"]
                for issue in validation.issues[:10]
            )
            raise CompletedWorkbookExportError(
                "completed Workbook semantic validation failed: "
                f"issues={validation.issue_count} {preview}"
            )

        if not source_frame.loc[:, list(SOURCE_COLUMNS)].equals(
            output_frame.loc[:, list(SOURCE_COLUMNS)]
        ):
            raise CompletedWorkbookExportError(
                "immutable source columns changed during export"
            )

        output_image_count = _embedded_image_count(staging_path)
        if output_image_count != source_image_count:
            raise CompletedWorkbookExportError(
                "embedded image count changed during export: "
                f"source={source_image_count} output={output_image_count}"
            )

        if sha256_file(package.workbook_path) != source_workbook_hash:
            raise CompletedWorkbookExportError(
                "source Workbook changed during export"
            )
        if (
            sha256_file(package.config.frozen_zip_path)
            != package.config.frozen_zip_sha256
        ):
            raise CompletedWorkbookExportError(
                "frozen package ZIP changed during export"
            )

        staging_path.replace(output_path)
    except Exception:
        staging_path.unlink(missing_ok=True)
        output_path.unlink(missing_ok=True)
        raise

    exported_at = datetime.now(timezone.utc).isoformat(
        timespec="microseconds"
    )
    return CompletedWorkbookExport(
        output_path=output_path,
        sha256=sha256_file(output_path),
        row_count=len(package.cases),
        logical_fingerprint=validation.logical_fingerprint,
        backup_path=backup_path,
        exported_at_utc=exported_at,
    )
