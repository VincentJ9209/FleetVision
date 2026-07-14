from __future__ import annotations

import csv
import json
import shutil
import uuid
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Mapping

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Protection
from PIL import Image, ImageDraw

from fleetvision.review.annotation_correction_review_mapping import (
    CanonicalCorrectionFields,
    parse_replacement_bbox,
    parse_target_bbox_ids,
    proposal_fingerprint,
)
from fleetvision.review.annotation_correction_review_package import (
    CorrectionSourceCase,
    VerifiedCorrectionReviewPackage,
    sha256_file,
)
from fleetvision.review.annotation_correction_review_state import CorrectionReviewStateStore


class CorrectionReviewExportError(RuntimeError):
    """Raised when completed correction proposal export is incomplete or unsafe."""


EXPORT_CLASSIFICATION = "PHASE_04_5M_ANNOTATION_CORRECTION_PROPOSALS_REVIEWED"
EXPORT_COLUMNS = (
    "schema_version", "correction_review_batch_id", "correction_case_id", "review_case_id", "image_id",
    "source_split", "source_case_fingerprint", "original_annotation_defect_type", "original_review_notes",
    "source_gt_bbox_records_json", "source_prediction_bbox_records_json", "correction_review_status",
    "correction_decision", "correction_operation", "target_gt_bbox_ids_json",
    "replacement_bbox_coordinates_json", "correction_reason", "correction_reviewer",
    "correction_reviewed_at_utc", "proposal_fingerprint",
)


@dataclass(frozen=True)
class CorrectionReviewExport:
    export_root: Path
    reviewed_csv_path: Path
    reviewed_json_path: Path
    completed_workbook_path: Path
    result_json_path: Path
    checksum_manifest_path: Path
    proposed_overlay_paths: tuple[Path, ...]
    exported_at_utc: str


def _canonical_fields(value: Mapping[str, str]) -> CanonicalCorrectionFields:
    return CanonicalCorrectionFields(**{key: str(value[key]) for key in CanonicalCorrectionFields.__dataclass_fields__})


def proposed_gt_boxes(source_boxes: tuple[Mapping[str, Any], ...], canonical: CanonicalCorrectionFields) -> tuple[dict[str, Any], ...]:
    boxes = [dict(row) for row in source_boxes]
    targets = set(parse_target_bbox_ids(canonical.target_gt_bbox_ids_json))
    replacement = parse_replacement_bbox(canonical.replacement_bbox_coordinates_json)
    operation = canonical.correction_operation
    if canonical.correction_decision == "REJECT_CORRECTION_KEEP_CURRENT_GT":
        return tuple(boxes)
    if operation in {"REMOVE_DUPLICATE_BBOX", "REMOVE_INVALID_BBOX"}:
        return tuple(row for row in boxes if row["bbox_id"] not in targets)
    if operation == "RESIZE_OR_REDRAW_BBOX":
        assert replacement is not None
        for row in boxes:
            if row["bbox_id"] in targets:
                row.update(replacement.as_dict())
        return tuple(boxes)
    if operation == "ADD_MISSING_BBOX":
        assert replacement is not None
        next_id = f"proposed_{1 + sum(str(row['bbox_id']).startswith('proposed_') for row in boxes):03d}"
        boxes.append({"bbox_id": next_id, **replacement.as_dict()})
        return tuple(boxes)
    if operation == "OTHER" and replacement is not None:
        boxes.append({"bbox_id": "proposed_001", **replacement.as_dict()})
    return tuple(boxes)


def _draw_proposed_overlay(case: CorrectionSourceCase, boxes: tuple[Mapping[str, Any], ...], target: Path) -> None:
    with Image.open(case.original_path) as image:
        canvas = image.convert("RGB")
    draw = ImageDraw.Draw(canvas)
    for row in boxes:
        draw.rectangle((row["x1"], row["y1"], row["x2"], row["y2"]), outline=(50, 205, 50), width=4)
        draw.text((row["x1"] + 3, row["y1"] + 3), str(row["bbox_id"]), fill=(255,255,255), stroke_width=2, stroke_fill=(50,205,50))
    target.parent.mkdir(parents=True, exist_ok=True)
    canvas.save(target, format="JPEG", quality=95, subsampling=0, optimize=False)


def _write_csv(path: Path, rows: list[Mapping[str, Any]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(EXPORT_COLUMNS))
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row[key] for key in EXPORT_COLUMNS})


def _write_workbook(path: Path, rows: list[Mapping[str, Any]], package: VerifiedCorrectionReviewPackage, overlays: Mapping[str, Path]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    instructions = workbook.create_sheet("Instructions")
    instructions["A1"] = "FleetVision Phase 04.5M — Annotation Correction Proposals"
    instructions["A1"].font = Font(size=16, bold=True)
    instructions["A3"] = "Archive only"
    instructions["B3"] = "This Workbook is an export/archive artifact. SQLite remains the live source of truth."
    instructions["A4"] = "Safety"
    instructions["B4"] = "No canonical annotation, COCO, dataset, Registry, fixed split, inference, or training mutation occurred."
    sheet = workbook.create_sheet("Correction_Proposals")
    headers = [*EXPORT_COLUMNS, "original_image", "current_gt_overlay", "combined_overlay", "proposed_overlay"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.protection = Protection(locked=True)
    case_by_id = package.case_by_id
    for row_index, row in enumerate(rows, start=2):
        case = case_by_id[row["correction_case_id"]]
        for col, header in enumerate(EXPORT_COLUMNS, start=1):
            sheet.cell(row=row_index, column=col, value=row[header])
        links = [case.original_path, case.gt_overlay_path, case.combined_overlay_path, overlays[case.correction_case_id]]
        for offset, link in enumerate(links, start=len(EXPORT_COLUMNS)+1):
            cell = sheet.cell(row=row_index, column=offset, value=str(link))
            cell.hyperlink = str(link)
            cell.style = "Hyperlink"
    hashes = workbook.create_sheet("Source_Hashes")
    hashes.append(["artifact", "sha256"])
    hashes.append(["source_csv", package.source_csv_sha256])
    hashes.append(["source_manifest", package.source_manifest_sha256])
    hashes.append(["source_contract", package.source_contract_sha256])
    hashes.append(["package_gate", package.package_gate_sha256])
    manifest = workbook.create_sheet("Manifest")
    manifest.append(["review_cases", len(rows)])
    manifest.append(["classification", EXPORT_CLASSIFICATION])
    workbook.save(path)


def _require_complete(package: VerifiedCorrectionReviewPackage, store: CorrectionReviewStateStore) -> None:
    progress = store.progress()
    if progress.total != 2 or progress.reviewed != 2 or progress.pending != 0 or progress.needs_adjudication != 0:
        raise CorrectionReviewExportError(
            f"completed export requires 2/2 reviewed；reviewed={progress.reviewed}/{progress.total} pending={progress.pending} needs_adjudication={progress.needs_adjudication}"
        )
    for path, expected, label in (
        (package.source_csv_path, package.source_csv_sha256, "source CSV"),
        (package.source_manifest_path, package.source_manifest_sha256, "source manifest"),
        (package.source_contract_path, package.source_contract_sha256, "source contract"),
        (package.package_gate_path, package.package_gate_sha256, "package gate"),
    ):
        if sha256_file(path) != expected:
            raise CorrectionReviewExportError(f"{label} 已改變")


def export_completed_correction_review(
    package: VerifiedCorrectionReviewPackage,
    store: CorrectionReviewStateStore,
) -> CorrectionReviewExport:
    _require_complete(package, store)
    export_root = package.export_root
    names = package.config
    final_paths = [
        export_root / names.reviewed_csv_name,
        export_root / names.reviewed_json_name,
        export_root / names.completed_workbook_name,
        export_root / names.result_json_name,
        export_root / "SHA256SUMS.csv",
    ]
    if export_root.exists() or any(path.exists() for path in final_paths):
        raise CorrectionReviewExportError("completed correction export already exists；overwrite is forbidden")
    staging = package.workspace_root / f".exports.staging-{uuid.uuid4().hex[:12]}"
    if staging.exists():
        raise CorrectionReviewExportError("unexpected export staging collision")
    staging.mkdir(parents=True)
    try:
        rows: list[dict[str, Any]] = []
        overlay_by_case: dict[str, Path] = {}
        for case in package.cases:
            stored = store.get_review(case.correction_case_id)
            if stored is None:
                raise CorrectionReviewExportError(f"stored review missing：{case.correction_case_id}")
            canonical = _canonical_fields(stored.canonical_fields)
            if canonical.correction_review_status != "reviewed":
                raise CorrectionReviewExportError(f"review not final：{case.correction_case_id}")
            fingerprint = proposal_fingerprint(case.source_case_fingerprint, canonical)
            row = {
                "schema_version": package.config.schema_version,
                "correction_review_batch_id": package.workspace_root.name,
                "correction_case_id": case.correction_case_id,
                "review_case_id": case.review_case_id,
                "image_id": case.image_id,
                "source_split": case.source_split,
                "source_case_fingerprint": case.source_case_fingerprint,
                "original_annotation_defect_type": case.original_annotation_defect_type,
                "original_review_notes": case.original_review_notes,
                "source_gt_bbox_records_json": case.gt_bbox_records_json,
                "source_prediction_bbox_records_json": case.prediction_bbox_records_json,
                **canonical.as_dict(),
                "proposal_fingerprint": fingerprint,
            }
            rows.append(row)
            source_boxes = tuple(json.loads(case.gt_bbox_records_json))
            proposed = proposed_gt_boxes(source_boxes, canonical)
            target = staging / "proposed_overlay" / f"{case.correction_case_id}.jpg"
            _draw_proposed_overlay(case, proposed, target)
            overlay_by_case[case.correction_case_id] = (
                export_root / "proposed_overlay" / f"{case.correction_case_id}.jpg"
            )

        csv_path = staging / names.reviewed_csv_name
        json_path = staging / names.reviewed_json_name
        workbook_path = staging / names.completed_workbook_name
        result_path = staging / names.result_json_name
        checksum_path = staging / "SHA256SUMS.csv"
        _write_csv(csv_path, rows)
        roundtrip = list(csv.DictReader(csv_path.open("r", encoding="utf-8-sig")))
        if roundtrip != [{key: str(value) for key, value in row.items()} for row in rows]:
            raise CorrectionReviewExportError("CSV round-trip validation failed")
        json_path.write_text(json.dumps({"schema_version":"1","proposal_count":2,"proposals":rows}, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
        _write_workbook(workbook_path, rows, package, overlay_by_case)
        exported_at = datetime.now(timezone.utc).isoformat(timespec="microseconds")
        result_payload = {
            "gate_id": "PHASE_04_5M_ANNOTATION_CORRECTION_REVIEW_EXPORT",
            "outcome": "PASS",
            "classification": EXPORT_CLASSIFICATION,
            "review_cases": 2,
            "reviewed": 2,
            "pending": 0,
            "needs_adjudication": 0,
            "canonical_annotation_modified": False,
            "canonical_coco_modified": False,
            "dataset_modified": False,
            "registry_modified": False,
            "fixed_splits_modified": False,
            "test_split_read": False,
            "model_inference_executed": False,
            "training_started": False,
            "retraining_status": "NOT_YET_APPROVED",
            "deployment_acceptance": "NOT_YET_APPROVED",
            "exported_at_utc": exported_at,
        }
        result_path.write_text(json.dumps(result_payload, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
        artifacts = sorted(path for path in staging.rglob("*") if path.is_file())
        with checksum_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=["relative_path","size_bytes","sha256"])
            writer.writeheader()
            for path in artifacts:
                writer.writerow({"relative_path": path.relative_to(staging).as_posix(), "size_bytes": path.stat().st_size, "sha256": sha256_file(path)})
        staging.replace(export_root)
    except Exception:
        shutil.rmtree(staging, ignore_errors=True)
        raise
    reviewed_csv = export_root / names.reviewed_csv_name
    store.record_export(reviewed_csv, sha256_file(reviewed_csv))
    return CorrectionReviewExport(
        export_root=export_root,
        reviewed_csv_path=reviewed_csv,
        reviewed_json_path=export_root / names.reviewed_json_name,
        completed_workbook_path=export_root / names.completed_workbook_name,
        result_json_path=export_root / names.result_json_name,
        checksum_manifest_path=export_root / "SHA256SUMS.csv",
        proposed_overlay_paths=tuple(export_root / "proposed_overlay" / f"{case.correction_case_id}.jpg" for case in package.cases),
        exported_at_utc=exported_at,
    )
