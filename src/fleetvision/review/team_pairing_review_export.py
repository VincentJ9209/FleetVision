from __future__ import annotations

import csv
import json
import shutil
import sqlite3
import uuid
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from fleetvision.data.team_pairing_audit import (
    SourceMutationError,
    build_source_snapshot,
    sha256_file,
    verify_source_snapshots,
)
from fleetvision.review.team_pairing_review_app import (
    TeamPairingReviewRuntime,
    required_angle_image_ids,
)


class TeamPairingReviewExportError(RuntimeError):
    """Raised when a completed Team Pairing export is incomplete or unsafe."""


WORKBOOK_SHEETS = (
    "執行摘要",
    "圖片清單",
    "候選批次",
    "批次成員",
    "圖片角度",
    "配對候選",
    "確認案例",
    "稽核資訊",
)

_ALLOWED_PRIMARY_CLASSIFICATIONS = {
    "NO_NEW_DAMAGE",
    "EXISTING_DAMAGE_UNCHANGED",
}


@dataclass(frozen=True)
class TeamPairingCompletedExport:
    export_root: Path
    inventory_csv: Path
    batch_candidates_csv: Path
    image_reviews_csv: Path
    pair_candidates_csv: Path
    workbook: Path
    summary_json: Path
    checksum_manifest: Path
    final_backup: Path
    exported_at_utc: str


def _read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    if not path.is_file():
        raise TeamPairingReviewExportError(f"required candidate artifact missing: {path}")
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            raise TeamPairingReviewExportError(f"CSV header missing: {path}")
        return list(reader.fieldnames), [dict(row) for row in reader]


def _write_csv(path: Path, fieldnames: Sequence[str], rows: Sequence[Mapping[str, Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(fieldnames), extrasaction="raise")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})


def _load_json_object(path: Path) -> dict[str, Any]:
    if not path.is_file():
        raise TeamPairingReviewExportError(f"required JSON artifact missing: {path}")
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise TeamPairingReviewExportError(f"invalid JSON artifact: {path}") from exc
    if not isinstance(value, dict):
        raise TeamPairingReviewExportError(f"JSON artifact must be an object: {path}")
    return value


def validate_primary_demo_rows(rows: Sequence[Mapping[str, Any]]) -> str:
    primary = [row for row in rows if str(row.get("manual_demo_role", "")) == "primary"]
    if len(primary) != 1:
        raise TeamPairingReviewExportError(
            f"completed export requires exactly one primary demo pair; found={len(primary)}"
        )
    classification = str(primary[0].get("derived_case_classification", ""))
    if classification not in _ALLOWED_PRIMARY_CLASSIFICATIONS:
        raise TeamPairingReviewExportError(
            f"primary demo pair classification is not reliable: {classification}"
        )
    return str(primary[0].get("pair_candidate_id", ""))


def _batch_rows(runtime: TeamPairingReviewRuntime) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for batch in sorted(runtime.package.batches, key=lambda item: item.batch_sequence):
        stored = runtime.store.get_batch_review(batch.batch_id)
        canonical = dict(stored.canonical_fields)
        rows.append(
            {
                "batch_id": batch.batch_id,
                "batch_sequence": batch.batch_sequence,
                "start_time_utc": batch.start_time_utc,
                "end_time_utc": batch.end_time_utc,
                "manual_batch_status": canonical.get("manual_batch_status", "pending"),
                "manual_vehicle_id": canonical.get("manual_vehicle_id", ""),
                "manual_stage": canonical.get("manual_stage", "unknown"),
                "manual_notes": canonical.get("manual_notes", ""),
                "review_reviewer": canonical.get("review_reviewer", ""),
                "reviewed_at_utc": canonical.get("reviewed_at_utc", ""),
                "review_revision": stored.revision,
            }
        )
    return rows


def _image_review_rows(
    runtime: TeamPairingReviewRuntime,
    inventory_rows: Sequence[Mapping[str, str]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for source in sorted(
        inventory_rows,
        key=lambda row: (int(row["inventory_sequence"]), str(row["image_id"])),
    ):
        stored = runtime.store.get_image_review(str(source["image_id"]))
        canonical = dict(stored.canonical_fields)
        rows.append(
            {
                **dict(source),
                "angle_review_status": canonical.get("review_status", "pending"),
                "manual_angle": canonical.get("manual_angle", "unknown"),
                "angle_review_notes": canonical.get("manual_notes", ""),
                "angle_review_reviewer": canonical.get("review_reviewer", ""),
                "angle_reviewed_at_utc": canonical.get("reviewed_at_utc", ""),
                "angle_review_revision": stored.revision,
            }
        )
    return rows


def _pair_review_rows(runtime: TeamPairingReviewRuntime) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for candidate in runtime.store.pair_candidate_rows():
        pair_id = str(candidate["pair_candidate_id"])
        stored = runtime.store.get_pair_review(pair_id)
        canonical = dict(stored.canonical_fields)
        rows.append(
            {
                "pair_candidate_id": pair_id,
                "pair_sequence": int(candidate["pair_sequence"]),
                "before_batch_id": candidate["before_batch_id"],
                "after_batch_id": candidate["after_batch_id"],
                "manual_vehicle_id": candidate["manual_vehicle_id"],
                "elapsed_seconds": int(candidate["elapsed_seconds"]),
                "overlap_angles_json": json.dumps(
                    list(candidate["overlap_angles"]),
                    ensure_ascii=False,
                    separators=(",", ":"),
                ),
                "overlap_count": int(candidate["overlap_count"]),
                "four_angle_overlap_count": int(candidate["four_angle_overlap_count"]),
                "manual_pair_status": canonical.get("manual_pair_status", "pending"),
                "manual_existing_damage_visible": canonical.get(
                    "manual_existing_damage_visible", "uncertain"
                ),
                "manual_new_damage_status": canonical.get(
                    "manual_new_damage_status", "uncertain"
                ),
                "manual_demo_role": canonical.get("manual_demo_role", "none"),
                "manual_notes": canonical.get("manual_notes", ""),
                "derived_case_classification": canonical.get(
                    "derived_case_classification", "MANUAL_REVIEW_REQUIRED"
                ),
                "review_reviewer": canonical.get("review_reviewer", ""),
                "reviewed_at_utc": canonical.get("reviewed_at_utc", ""),
                "review_revision": stored.revision,
            }
        )
    return rows


def _validate_readiness(
    runtime: TeamPairingReviewRuntime,
    batch_rows: Sequence[Mapping[str, Any]],
    image_rows: Sequence[Mapping[str, Any]],
    pair_rows: Sequence[Mapping[str, Any]],
) -> tuple[str, tuple[str, ...]]:
    if runtime.store.integrity_check() != "ok":
        raise TeamPairingReviewExportError("SQLite integrity check failed")
    runtime.store.verify_event_log_continuity()

    pending_batches = [row for row in batch_rows if row["manual_batch_status"] == "pending"]
    if pending_batches:
        raise TeamPairingReviewExportError(
            f"candidate batches are still pending: {len(pending_batches)}"
        )
    for row in batch_rows:
        if row["manual_batch_status"] == "confirmed":
            if not str(row["manual_vehicle_id"]).strip() or row["manual_stage"] not in {
                "before",
                "after",
            }:
                raise TeamPairingReviewExportError(
                    f"confirmed batch is missing vehicle/stage: {row['batch_id']}"
                )

    image_by_id = {str(row["image_id"]): row for row in image_rows}
    required_images: set[str] = set()
    for row in batch_rows:
        if row["manual_batch_status"] == "confirmed":
            required_images.update(required_angle_image_ids(runtime, str(row["batch_id"])))
    for image_id in sorted(required_images):
        row = image_by_id.get(image_id)
        if row is None or row["angle_review_status"] != "reviewed" or row["manual_angle"] == "unknown":
            raise TeamPairingReviewExportError(
                f"required image angle is incomplete: {image_id}"
            )

    pending_pairs = [row for row in pair_rows if row["manual_pair_status"] == "pending"]
    if pending_pairs:
        raise TeamPairingReviewExportError(
            f"pair candidates are still pending: {len(pending_pairs)}"
        )
    confirmed = [row for row in pair_rows if row["manual_pair_status"] == "confirmed"]
    if len(confirmed) < 3:
        raise TeamPairingReviewExportError(
            f"completed export requires at least 3 confirmed pairs; found={len(confirmed)}"
        )
    primary_id = validate_primary_demo_rows(confirmed)
    backup_ids = tuple(
        str(row["pair_candidate_id"])
        for row in confirmed
        if row["manual_demo_role"] == "backup"
    )
    return primary_id, backup_ids


def _verify_source(runtime: TeamPairingReviewRuntime) -> dict[str, Any]:
    source_dir = runtime.package.workspace_root / "source"
    before = _load_json_object(source_dir / "source_snapshot_before.json")
    stored_after = _load_json_object(source_dir / "source_snapshot_after.json")
    stored_verification = _load_json_object(
        source_dir / "source_snapshot_verification.json"
    )
    try:
        verify_source_snapshots(before, stored_after)
        current = build_source_snapshot(runtime.config.source_root)
        current_verification = verify_source_snapshots(before, current)
    except SourceMutationError as exc:
        raise TeamPairingReviewExportError(
            "source snapshot verification failed"
        ) from exc
    if stored_verification.get("byte_identical") is not True:
        raise TeamPairingReviewExportError("stored source snapshot verification is not PASS")
    return current_verification


def _style_table_sheet(sheet: Any, headers: Sequence[str], rows: Sequence[Mapping[str, Any]]) -> None:
    sheet.append(list(headers))
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column_cells in sheet.columns:
        width = min(40, max(10, max(len(str(cell.value or "")) for cell in column_cells) + 2))
        sheet.column_dimensions[column_cells[0].column_letter].width = width


def _write_workbook(
    path: Path,
    *,
    summary_rows: Sequence[Mapping[str, Any]],
    inventory_rows: Sequence[Mapping[str, Any]],
    batch_rows: Sequence[Mapping[str, Any]],
    member_rows: Sequence[Mapping[str, Any]],
    image_rows: Sequence[Mapping[str, Any]],
    pair_rows: Sequence[Mapping[str, Any]],
    audit_rows: Sequence[Mapping[str, Any]],
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    summary = workbook.create_sheet(WORKBOOK_SHEETS[0])
    _style_table_sheet(summary, ("項目", "內容"), summary_rows)
    _style_table_sheet(
        workbook.create_sheet(WORKBOOK_SHEETS[1]),
        tuple(inventory_rows[0]) if inventory_rows else (),
        inventory_rows,
    )
    _style_table_sheet(
        workbook.create_sheet(WORKBOOK_SHEETS[2]),
        tuple(batch_rows[0]) if batch_rows else (),
        batch_rows,
    )
    _style_table_sheet(
        workbook.create_sheet(WORKBOOK_SHEETS[3]),
        tuple(member_rows[0]) if member_rows else (),
        member_rows,
    )
    _style_table_sheet(
        workbook.create_sheet(WORKBOOK_SHEETS[4]),
        tuple(image_rows[0]) if image_rows else (),
        image_rows,
    )
    _style_table_sheet(
        workbook.create_sheet(WORKBOOK_SHEETS[5]),
        tuple(pair_rows[0]) if pair_rows else (),
        pair_rows,
    )
    confirmed = [row for row in pair_rows if row["manual_pair_status"] == "confirmed"]
    _style_table_sheet(
        workbook.create_sheet(WORKBOOK_SHEETS[6]),
        tuple(pair_rows[0]) if pair_rows else (),
        confirmed,
    )
    _style_table_sheet(
        workbook.create_sheet(WORKBOOK_SHEETS[7]),
        ("項目", "內容"),
        audit_rows,
    )
    workbook.save(path)


def _validate_workbook(path: Path, expected_counts: Mapping[str, int]) -> None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if tuple(workbook.sheetnames) != WORKBOOK_SHEETS:
            raise TeamPairingReviewExportError("completed workbook sheet contract mismatch")
        for sheet_name, row_count in expected_counts.items():
            if workbook[sheet_name].max_row != row_count + 1:
                raise TeamPairingReviewExportError(
                    f"completed workbook row count mismatch: {sheet_name}"
                )
    finally:
        workbook.close()


def _record_export(database_path: Path, output_path: Path, digest: str) -> None:
    connection = sqlite3.connect(database_path)
    try:
        connection.execute(
            "INSERT INTO export_history(output_path,sha256,exported_at_utc) VALUES(?,?,?)",
            (
                str(output_path.resolve()),
                digest,
                datetime.now(timezone.utc).isoformat(timespec="microseconds"),
            ),
        )
        connection.commit()
    except sqlite3.IntegrityError as exc:
        raise TeamPairingReviewExportError("export history already exists") from exc
    finally:
        connection.close()


def export_completed_team_pairing_review(
    runtime: TeamPairingReviewRuntime,
    *,
    timestamp: str | None = None,
    repository_commit: str = "UNKNOWN",
    simulate_failure_after_first_artifact: bool = False,
) -> TeamPairingCompletedExport:
    workspace = runtime.package.workspace_root.resolve()
    exports_root = workspace / "exports"
    token = timestamp or datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    final_root = exports_root / f"completed_{token}"
    if final_root.exists():
        raise TeamPairingReviewExportError("completed export overwrite is forbidden")
    exports_root.mkdir(parents=True, exist_ok=True)
    staging = exports_root / f".completed_{token}.staging-{uuid.uuid4().hex[:12]}"
    if staging.exists():
        raise TeamPairingReviewExportError("unexpected export staging collision")

    candidates = workspace / "candidates"
    inventory_headers, inventory_rows = _read_csv(candidates / "team_image_inventory.csv")
    batch_candidate_headers, batch_candidate_rows = _read_csv(
        candidates / "team_capture_batch_candidates.csv"
    )
    member_headers, member_rows = _read_csv(
        candidates / "team_capture_batch_members.csv"
    )
    batch_rows = _batch_rows(runtime)
    image_rows = _image_review_rows(runtime, inventory_rows)
    pair_rows = _pair_review_rows(runtime)
    primary_id, backup_ids = _validate_readiness(runtime, batch_rows, image_rows, pair_rows)
    source_verification = _verify_source(runtime)
    final_backup = runtime.store.create_backup()
    exported_at = datetime.now(timezone.utc).isoformat(timespec="microseconds")

    inventory_name = "team_image_inventory.csv"
    batches_name = "team_capture_batch_candidates.csv"
    image_reviews_name = "team_image_reviews_completed.csv"
    pair_candidates_name = "team_before_after_pair_candidates.csv"
    workbook_name = "team_pair_review_completed.xlsx"
    summary_name = "team_pairing_summary.json"
    checksum_name = "SHA256SUMS.csv"

    try:
        staging.mkdir(parents=True)
        shutil.copy2(candidates / inventory_name, staging / inventory_name)
        if simulate_failure_after_first_artifact:
            raise TeamPairingReviewExportError("simulated export failure")
        shutil.copy2(candidates / batches_name, staging / batches_name)

        image_headers = [
            *inventory_headers,
            "angle_review_status",
            "manual_angle",
            "angle_review_notes",
            "angle_review_reviewer",
            "angle_reviewed_at_utc",
            "angle_review_revision",
        ]
        pair_headers = list(pair_rows[0]) if pair_rows else []
        _write_csv(staging / image_reviews_name, image_headers, image_rows)
        _write_csv(staging / pair_candidates_name, pair_headers, pair_rows)

        batch_status_counts = Counter(str(row["manual_batch_status"]) for row in batch_rows)
        pair_status_counts = Counter(str(row["manual_pair_status"]) for row in pair_rows)
        classification_counts = Counter(
            str(row["derived_case_classification"]) for row in pair_rows
        )
        summary_rows = [
            {"項目": "匯出時間 UTC", "內容": exported_at},
            {"項目": "圖片數", "內容": len(inventory_rows)},
            {"項目": "候選批次數", "內容": len(batch_rows)},
            {"項目": "確認配對數", "內容": pair_status_counts["confirmed"]},
            {"項目": "主要展示案例", "內容": primary_id},
        ]
        audit_rows = [
            {"項目": "repository_commit", "內容": repository_commit},
            {"項目": "config_sha256", "內容": sha256_file(runtime.config.config_path)},
            {"項目": "candidate_manifest_sha256", "內容": runtime.package.identity.candidate_manifest_sha256},
            {"項目": "source_snapshot_sha256", "內容": source_verification["before_snapshot_sha256"]},
            {"項目": "SQLite_integrity", "內容": "ok"},
            {"項目": "Frozen_Test_access", "內容": "false"},
            {"項目": "training_inference", "內容": "false"},
        ]
        _write_workbook(
            staging / workbook_name,
            summary_rows=summary_rows,
            inventory_rows=inventory_rows,
            batch_rows=batch_rows,
            member_rows=member_rows,
            image_rows=image_rows,
            pair_rows=pair_rows,
            audit_rows=audit_rows,
        )
        _validate_workbook(
            staging / workbook_name,
            {
                "圖片清單": len(inventory_rows),
                "候選批次": len(batch_rows),
                "批次成員": len(member_rows),
                "圖片角度": len(image_rows),
                "配對候選": len(pair_rows),
                "確認案例": pair_status_counts["confirmed"],
            },
        )

        artifact_metadata = []
        for name in sorted(
            (inventory_name, batches_name, image_reviews_name, pair_candidates_name, workbook_name)
        ):
            path = staging / name
            artifact_metadata.append(
                {
                    "relative_path": name,
                    "size_bytes": path.stat().st_size,
                    "sha256": sha256_file(path),
                }
            )
        summary_payload = {
            "schema_version": runtime.config.schema_version,
            "run_id": workspace.name,
            "exported_at_utc": exported_at,
            "repository_commit": repository_commit,
            "config_path": str(runtime.config.config_path),
            "config_sha256": sha256_file(runtime.config.config_path),
            "source_root": str(runtime.config.source_root),
            "output_root": str(runtime.config.output_root),
            "source_count": len(inventory_rows),
            "readable_count": sum(str(row.get("is_readable", "")).lower() == "true" for row in inventory_rows),
            "unreadable_count": sum(str(row.get("is_readable", "")).lower() != "true" for row in inventory_rows),
            "batch_status_distribution": dict(sorted(batch_status_counts.items())),
            "pair_status_distribution": dict(sorted(pair_status_counts.items())),
            "derived_classification_distribution": dict(sorted(classification_counts.items())),
            "confirmed_pair_count": pair_status_counts["confirmed"],
            "primary_demo_pair_id": primary_id,
            "backup_demo_pair_ids": list(backup_ids),
            "source_snapshot_verification": source_verification,
            "warnings": [],
            "frozen_test_access": False,
            "model_inference_executed": False,
            "training_started": False,
            "artifacts": artifact_metadata,
        }
        (staging / summary_name).write_text(
            json.dumps(summary_payload, ensure_ascii=False, indent=2, sort_keys=True),
            encoding="utf-8",
        )

        checksum_targets = sorted(
            path for path in staging.iterdir() if path.is_file() and path.name != checksum_name
        )
        with (staging / checksum_name).open("w", newline="", encoding="utf-8-sig") as handle:
            writer = csv.DictWriter(
                handle,
                fieldnames=("relative_path", "size_bytes", "sha256"),
            )
            writer.writeheader()
            for path in checksum_targets:
                writer.writerow(
                    {
                        "relative_path": path.name,
                        "size_bytes": path.stat().st_size,
                        "sha256": sha256_file(path),
                    }
                )

        staging.replace(final_root)
        _record_export(
            runtime.store.database_path,
            final_root,
            sha256_file(final_root / checksum_name),
        )
    except Exception:
        shutil.rmtree(staging, ignore_errors=True)
        shutil.rmtree(final_root, ignore_errors=True)
        raise

    return TeamPairingCompletedExport(
        export_root=final_root,
        inventory_csv=final_root / inventory_name,
        batch_candidates_csv=final_root / batches_name,
        image_reviews_csv=final_root / image_reviews_name,
        pair_candidates_csv=final_root / pair_candidates_name,
        workbook=final_root / workbook_name,
        summary_json=final_root / summary_name,
        checksum_manifest=final_root / checksum_name,
        final_backup=final_backup,
        exported_at_utc=exported_at,
    )
