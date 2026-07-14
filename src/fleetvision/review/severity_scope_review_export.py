from __future__ import annotations

import json
import shutil
import uuid
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from fleetvision.data.validation_error_human_review import sha256_file
from fleetvision.data.validation_error_review_findings import (
    SCOPE_COLUMNS,
    SCOPE_WORKBOOK_SHEETS,
    load_findings_config,
    read_scope_workbook,
    validate_scope_dataframe,
)
from fleetvision.review.severity_scope_review_package import VerifiedScopePackage
from fleetvision.review.severity_scope_review_state import ScopeReviewStateStore


class ScopeCompletedWorkbookExportError(RuntimeError):
    """Raised when completed scope Workbook export is incomplete or unsafe."""


@dataclass(frozen=True)
class ScopeCompletedWorkbookExport:
    output_path: Path
    result_path: Path
    sha256: str
    row_count: int
    backup_path: Path
    exported_at_utc: str


def _embedded_image_count(path: Path) -> int:
    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        return sum(
            len(getattr(sheet, "_images", ()))
            for sheet in workbook.worksheets
        )
    finally:
        workbook.close()


def _require_complete_state(
    package: VerifiedScopePackage,
    store: ScopeReviewStateStore,
) -> None:
    progress = store.progress()
    if progress.total != len(package.cases):
        raise ScopeCompletedWorkbookExportError(
            "state/package case count mismatch："
            f"state={progress.total} package={len(package.cases)}"
        )
    if (
        progress.reviewed != progress.total
        or progress.pending != 0
        or progress.needs_adjudication != 0
    ):
        raise ScopeCompletedWorkbookExportError(
            "completed scope Workbook requires 130/130 reviewed；"
            f"reviewed={progress.reviewed}/{progress.total} "
            f"pending={progress.pending} "
            f"needs_adjudication={progress.needs_adjudication}"
        )


def export_completed_scope_workbook(
    package: VerifiedScopePackage,
    store: ScopeReviewStateStore,
) -> ScopeCompletedWorkbookExport:
    """Export a no-overwrite completed scope Workbook from trusted SQLite state."""

    _require_complete_state(package, store)
    if sha256_file(package.source_csv_path) != package.source_csv_sha256:
        raise ScopeCompletedWorkbookExportError("F1 scope source CSV 已改變")
    if sha256_file(package.template_workbook_path) != package.template_workbook_sha256:
        raise ScopeCompletedWorkbookExportError("F1 scope template Workbook 已改變")
    if sha256_file(package.asset_manifest_path) != package.asset_manifest_sha256:
        raise ScopeCompletedWorkbookExportError("F1 scope asset manifest 已改變")

    store.export_dir.mkdir(parents=True, exist_ok=True)
    output_path = package.completed_workbook_path
    result_path = store.export_dir / "scope_review_export_result.json"
    if output_path.exists() or result_path.exists():
        raise ScopeCompletedWorkbookExportError(
            "completed scope export already exists；overwrite is forbidden"
        )

    token = uuid.uuid4().hex[:12]
    staging = store.export_dir / f".{output_path.name}.staging-{token}.xlsx"
    if staging.exists():
        raise ScopeCompletedWorkbookExportError("unexpected staging collision")

    source_frame = pd.read_csv(
        package.source_csv_path,
        dtype=str,
        keep_default_na=False,
        encoding="utf-8-sig",
    ).fillna("").astype(str)
    source_image_count = _embedded_image_count(package.template_workbook_path)
    backup_path = store.create_backup()

    shutil.copy2(package.template_workbook_path, staging)
    try:
        workbook = load_workbook(staging, read_only=False, data_only=False)
        try:
            if tuple(workbook.sheetnames) != SCOPE_WORKBOOK_SHEETS:
                raise ScopeCompletedWorkbookExportError(
                    "scope template Workbook sheet contract changed"
                )
            sheet = workbook["Scope_Review"]
            headers = {str(cell.value): cell.column for cell in sheet[1]}
            required_headers = ["review_case_id", *SCOPE_COLUMNS]
            missing = [header for header in required_headers if header not in headers]
            if missing:
                raise ScopeCompletedWorkbookExportError(
                    f"scope template missing headers：{missing}"
                )

            row_by_case_id: dict[str, int] = {}
            for row_index in range(2, sheet.max_row + 1):
                review_case_id = str(
                    sheet.cell(
                        row=row_index,
                        column=headers["review_case_id"],
                    ).value
                    or ""
                ).strip()
                if not review_case_id:
                    continue
                if review_case_id in row_by_case_id:
                    raise ScopeCompletedWorkbookExportError(
                        f"duplicate review_case_id：{review_case_id}"
                    )
                row_by_case_id[review_case_id] = row_index

            expected_ids = {case.review_case_id for case in package.cases}
            if set(row_by_case_id) != expected_ids:
                raise ScopeCompletedWorkbookExportError(
                    "scope template/state review_case_id set mismatch"
                )

            for case in package.cases:
                stored = store.get_review(case.review_case_id)
                if stored is None:
                    raise ScopeCompletedWorkbookExportError(
                        f"stored scope review missing：{case.review_case_id}"
                    )
                canonical = dict(stored.canonical_fields)
                if set(canonical) != set(SCOPE_COLUMNS):
                    raise ScopeCompletedWorkbookExportError(
                        f"stored scope field contract mismatch：{case.review_case_id}"
                    )
                if canonical["scope_review_status"] != "reviewed":
                    raise ScopeCompletedWorkbookExportError(
                        f"stored scope review is not final：{case.review_case_id}"
                    )
                row_index = row_by_case_id[case.review_case_id]
                for column in SCOPE_COLUMNS:
                    sheet.cell(
                        row=row_index,
                        column=headers[column],
                        value=canonical[column],
                    )
            workbook.save(staging)
        finally:
            workbook.close()

        output_frame = read_scope_workbook(staging)
        findings_config = load_findings_config(
            package.config.findings_config_path,
            package.config.project_root,
        )
        validation = validate_scope_dataframe(
            output_frame,
            source_frame,
            findings_config,
        )
        if not validation.passed:
            preview = "; ".join(
                issue["error_code"] for issue in validation.issues[:10]
            )
            raise ScopeCompletedWorkbookExportError(
                "completed scope Workbook semantic validation failed："
                f"issues={validation.issue_count} {preview}"
            )

        if _embedded_image_count(staging) != source_image_count:
            raise ScopeCompletedWorkbookExportError(
                "embedded image count changed during scope export"
            )
        if sha256_file(package.source_csv_path) != package.source_csv_sha256:
            raise ScopeCompletedWorkbookExportError("F1 scope source changed during export")
        if sha256_file(package.template_workbook_path) != package.template_workbook_sha256:
            raise ScopeCompletedWorkbookExportError(
                "F1 scope template changed during export"
            )
        staging.replace(output_path)
    except Exception:
        staging.unlink(missing_ok=True)
        output_path.unlink(missing_ok=True)
        raise

    exported_at = datetime.now(timezone.utc).isoformat(timespec="microseconds")
    output_hash = sha256_file(output_path)
    payload = {
        "gate_id": "PHASE_04_5L_SCOPE_REVIEW_APP_EXPORT",
        "outcome": "PASS",
        "classification": "LOCAL_SCOPE_REVIEW_APP_COMPLETED_WORKBOOK_EXPORTED",
        "completed_scope_workbook": str(output_path),
        "completed_scope_workbook_sha256": output_hash,
        "review_cases": len(package.cases),
        "reviewed": len(package.cases),
        "pending": 0,
        "needs_adjudication": 0,
        "source_scope_csv_sha256": package.source_csv_sha256,
        "source_scope_template_sha256": package.template_workbook_sha256,
        "source_scope_asset_manifest_sha256": package.asset_manifest_sha256,
        "pre_export_backup": str(backup_path),
        "exported_at_utc": exported_at,
        "test_split_read": False,
        "model_inference_executed": False,
        "annotation_modified": False,
        "training_started": False,
        "retraining_status": "NOT_YET_APPROVED",
        "deployment_acceptance": "NOT_YET_APPROVED",
    }
    result_staging = store.export_dir / (
        f".{result_path.name}.staging-{uuid.uuid4().hex[:12]}.json"
    )
    try:
        result_staging.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True),
            encoding="utf-8",
        )
        result_staging.replace(result_path)
        store.record_export(output_path, output_hash)
    except Exception:
        result_staging.unlink(missing_ok=True)
        result_path.unlink(missing_ok=True)
        output_path.unlink(missing_ok=True)
        raise
    return ScopeCompletedWorkbookExport(
        output_path=output_path,
        result_path=result_path,
        sha256=output_hash,
        row_count=len(package.cases),
        backup_path=backup_path,
        exported_at_utc=exported_at,
    )
