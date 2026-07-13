"""Controlled Phase 04.5L validation-error human-review workflow.

This module prepares a validation-only Excel review package, exports completed
human review to canonical CSV, validates the canonical result, and creates
non-destructive data-improvement proposals. It never reads the test split,
modifies annotations, starts training, or grants deployment acceptance.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import os
import shutil
import tempfile
import uuid
import zipfile
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path, PurePosixPath
from typing import Any, Iterable, Mapping, Sequence

import pandas as pd
import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from PIL import Image, ImageDraw


DEFAULT_CONFIG_PATH = Path("configs/data/validation_error_human_review_config.yaml")

WORKLIST_REQUIRED_COLUMNS = (
    "image_id",
    "primary_error_type",
    "error_types",
    "error_case_count",
    "ground_truth_error_count",
    "prediction_error_count",
)
PREDICTION_REQUIRED_COLUMNS = ("split", "image_id", "confidence", "x1", "y1", "x2", "y2")
GROUND_TRUTH_REQUIRED_COLUMNS = ("split", "image_id", "x1", "y1", "x2", "y2")

SOURCE_COLUMNS = (
    "schema_version",
    "review_batch_id",
    "review_case_id",
    "source_04_5k_zip_sha256",
    "source_case_fingerprint",
    "image_id",
    "image_filename",
    "auto_error_category",
    "auto_error_detail_ids",
    "error_case_count",
    "ground_truth_error_count",
    "prediction_error_count",
    "gt_count",
    "prediction_count",
    "max_prediction_confidence",
    "best_iou",
    "threshold_candidate",
    "threshold_designation",
    "original_image_relpath",
    "overlay_image_relpath",
)

HUMAN_COLUMNS = (
    "review_status",
    "error_disposition",
    "primary_root_cause",
    "secondary_root_cause",
    "annotation_quality",
    "annotation_defect_type",
    "recommended_action",
    "retraining_priority",
    "correction_proposal_required",
    "reviewer",
    "reviewed_at_utc",
    "review_notes",
)

CANONICAL_COLUMNS = SOURCE_COLUMNS + HUMAN_COLUMNS

SOURCE_FINGERPRINT_COLUMNS = (
    "schema_version",
    "review_batch_id",
    "review_case_id",
    "source_04_5k_zip_sha256",
    "image_id",
    "image_filename",
    "auto_error_category",
    "auto_error_detail_ids",
    "error_case_count",
    "ground_truth_error_count",
    "prediction_error_count",
    "gt_count",
    "prediction_count",
    "max_prediction_confidence",
    "best_iou",
    "threshold_candidate",
    "threshold_designation",
    "original_image_relpath",
    "overlay_image_relpath",
)

WORKBOOK_SHEETS = (
    "Instructions",
    "Review_Cases",
    "Option_Lists",
    "Manifest",
    "Progress_Summary",
)

PREVIEW_HEADERS = ("Original Preview", "Overlay Preview")


class HumanReviewError(RuntimeError):
    """Raised when Phase 04.5L safety or data-contract checks fail."""


@dataclass(frozen=True)
class ReviewConfig:
    """Resolved Phase 04.5L configuration."""

    project_root: Path
    schema_version: str
    expected_source_zip_name: str
    expected_source_zip_sha256: str
    expected_gate_classification: str
    expected_case_count: int
    expected_validation_image_count: int
    expected_validation_ground_truth_instances: int
    expected_raw_prediction_count: int
    threshold_candidate: float
    threshold_designation: str
    output_base_dir: Path
    source_files: Mapping[str, str]
    workbook_name: str
    thumbnail_width: int
    thumbnail_height: int
    options: Mapping[str, tuple[str, ...]]


@dataclass(frozen=True)
class PrepareResult:
    """Result of preparing a controlled review package."""

    batch_root: Path
    workbook_path: Path
    row_count: int
    asset_count: int
    source_zip_sha256: str


@dataclass(frozen=True)
class ExportResult:
    """Result of exporting a completed workbook to canonical CSV."""

    output_csv: Path
    row_count: int
    logical_fingerprint: str


@dataclass(frozen=True)
class ValidationResult:
    """Canonical validation outcome."""

    passed: bool
    row_count: int
    issue_count: int
    issues: tuple[dict[str, str], ...]
    logical_fingerprint: str
    counts: Mapping[str, int]


@dataclass(frozen=True)
class SummaryResult:
    """Result of producing Phase 04.5L downstream proposal artifacts."""

    output_paths: tuple[Path, ...]
    action_count: int
    correction_proposal_count: int


def _normalize_scalar(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.tzinfo is None:
            value = value.replace(tzinfo=timezone.utc)
        return value.astimezone(timezone.utc).isoformat()
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> str:
    """Return uppercase SHA256 for a file."""

    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(chunk_size), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def _sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest().upper()


def _sha256_stream(handle: Any, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    for chunk in iter(lambda: handle.read(chunk_size), b""):
        digest.update(chunk)
    return digest.hexdigest().upper()


def _verify_source_artifacts_against_zip(
    source_zip: Path,
    source_paths: Mapping[str, Path],
    configured_relpaths: Mapping[str, str],
) -> None:
    """Verify source artifacts are byte-identical to unique safe ZIP members."""

    with zipfile.ZipFile(source_zip, mode="r") as archive:
        bad_member = archive.testzip()
        if bad_member is not None:
            raise HumanReviewError(f"verified source ZIP CRC failure: {bad_member}")
        safe_members: list[zipfile.ZipInfo] = []
        for member in archive.infolist():
            normalized = member.filename.replace("\\", "/")
            pure = PurePosixPath(normalized)
            unix_mode = member.external_attr >> 16
            is_symlink = (unix_mode & 0o170000) == 0o120000
            if pure.is_absolute() or ".." in pure.parts or is_symlink:
                raise HumanReviewError(f"unsafe member in verified source ZIP: {member.filename}")
            if not member.is_dir():
                safe_members.append(member)

        for key, source_path in source_paths.items():
            relative = configured_relpaths[key].replace("\\", "/").lstrip("./")
            candidates = [
                member
                for member in safe_members
                if member.filename.replace("\\", "/") == relative
                or member.filename.replace("\\", "/").endswith("/" + relative)
            ]
            if len(candidates) != 1:
                raise HumanReviewError(
                    f"verified source ZIP must contain exactly one {key} artifact; "
                    f"relative={relative!r} matched={len(candidates)}"
                )
            with archive.open(candidates[0], mode="r") as handle:
                zip_hash = _sha256_stream(handle)
            source_hash = sha256_file(source_path)
            if source_hash != zip_hash:
                raise HumanReviewError(
                    f"source artifact does not match verified ZIP ({key}): "
                    f"source={source_hash} zip={zip_hash}"
                )


def _json_dump(path: Path, payload: Any) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )


def _write_csv(path: Path, rows: Iterable[Mapping[str, Any]], columns: Sequence[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(columns), extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({column: _normalize_scalar(row.get(column, "")) for column in columns})


def _load_csv(path: Path, required_columns: Sequence[str]) -> pd.DataFrame:
    if not path.is_file():
        raise HumanReviewError(f"required CSV not found: {path}")
    frame = pd.read_csv(path, dtype=str, keep_default_na=False, encoding="utf-8-sig")
    missing = [column for column in required_columns if column not in frame.columns]
    if missing:
        raise HumanReviewError(f"CSV missing required columns {missing}: {path}")
    return frame.fillna("").astype(str)


def _verify_04_5k_manifest_contract(
    source_paths: Mapping[str, Path],
    configured_relpaths: Mapping[str, str],
) -> None:
    """Verify selected 04.5K artifacts against its manifest and checksum ledger."""

    manifest = _load_csv(
        source_paths["artifact_manifest"],
        ("relative_path", "size_bytes", "sha256"),
    )
    manifest["relative_path"] = manifest["relative_path"].str.replace("\\", "/", regex=False)
    if manifest["relative_path"].eq("").any() or manifest["relative_path"].duplicated().any():
        raise HumanReviewError("04.5K artifact manifest contains blank or duplicate paths")
    manifest_by_path = {
        row["relative_path"]: row for row in manifest.to_dict(orient="records")
    }

    core_keys = ("gate_result", "worklist", "predictions", "ground_truth")
    for key in core_keys:
        relative = configured_relpaths[key].replace("\\", "/").lstrip("./")
        row = manifest_by_path.get(relative)
        actual_hash = sha256_file(source_paths[key])
        actual_size = str(source_paths[key].stat().st_size)
        if row is None or row["sha256"].strip().upper() != actual_hash or row["size_bytes"].strip() != actual_size:
            raise HumanReviewError(
                f"04.5K artifact manifest mismatch ({key}): relative={relative!r}"
            )

    checksum_entries: dict[str, str] = {}
    for line_number, raw_line in enumerate(
        source_paths["checksums"].read_text(encoding="ascii").splitlines(),
        start=1,
    ):
        line = raw_line.strip()
        if not line:
            continue
        parts = line.split(maxsplit=1)
        if len(parts) != 2:
            raise HumanReviewError(f"04.5K checksum ledger malformed at line {line_number}")
        digest, relative = parts[0].upper(), parts[1].strip().replace("\\", "/")
        if len(digest) != 64 or any(character not in "0123456789ABCDEF" for character in digest):
            raise HumanReviewError(f"04.5K checksum digest malformed at line {line_number}")
        if relative in checksum_entries:
            raise HumanReviewError(f"04.5K checksum ledger duplicate path: {relative}")
        checksum_entries[relative] = digest

    checksum_keys = core_keys + ("artifact_manifest",)
    for key in checksum_keys:
        relative = configured_relpaths[key].replace("\\", "/").lstrip("./")
        if checksum_entries.get(relative) != sha256_file(source_paths[key]):
            raise HumanReviewError(
                f"04.5K checksum ledger mismatch ({key}): relative={relative!r}"
            )


def _resolve_path(project_root: Path, value: str) -> Path:
    path = Path(value)
    return path if path.is_absolute() else project_root / path


def load_config(config_path: Path, project_root: Path) -> ReviewConfig:
    """Load YAML configuration and resolve project-relative output paths."""

    config_path = _resolve_path(project_root, str(config_path))
    if not config_path.is_file():
        raise FileNotFoundError(f"Phase 04.5L config not found: {config_path}")
    raw = yaml.safe_load(config_path.read_text(encoding="utf-8")) or {}
    expected = raw.get("expected_source", {})
    workbook = raw.get("workbook", {})
    output = raw.get("output", {})
    source_files = raw.get("source_files", {})
    options_raw = raw.get("options", {})

    required_source_keys = {
        "gate_result",
        "worklist",
        "predictions",
        "ground_truth",
        "artifact_manifest",
        "checksums",
    }
    missing_source_keys = sorted(required_source_keys - set(source_files))
    if missing_source_keys:
        raise HumanReviewError(f"config source_files missing keys: {missing_source_keys}")

    options: dict[str, tuple[str, ...]] = {}
    for key, values in options_raw.items():
        if not isinstance(values, list) or not values:
            raise HumanReviewError(f"config options.{key} must be a non-empty list")
        normalized = tuple(_normalize_scalar(value) for value in values)
        if any(not value for value in normalized) or len(set(normalized)) != len(normalized):
            raise HumanReviewError(f"config options.{key} contains blank or duplicate values")
        options[str(key)] = normalized

    required_option_keys = {
        "review_status",
        "error_disposition",
        "primary_root_cause",
        "secondary_root_cause",
        "annotation_quality",
        "annotation_defect_type",
        "recommended_action",
        "retraining_priority",
        "correction_proposal_required",
    }
    missing_options = sorted(required_option_keys - set(options))
    if missing_options:
        raise HumanReviewError(f"config options missing keys: {missing_options}")

    config = ReviewConfig(
        project_root=project_root.resolve(),
        schema_version=_normalize_scalar(raw.get("schema_version", "1")),
        expected_source_zip_name=_normalize_scalar(expected.get("zip_filename")),
        expected_source_zip_sha256=_normalize_scalar(expected.get("zip_sha256")).upper(),
        expected_gate_classification=_normalize_scalar(expected.get("gate_classification")),
        expected_case_count=int(expected.get("case_count", 0)),
        expected_validation_image_count=int(expected.get("validation_image_count", 0)),
        expected_validation_ground_truth_instances=int(
            expected.get("validation_ground_truth_instances", 0)
        ),
        expected_raw_prediction_count=int(expected.get("raw_prediction_count", 0)),
        threshold_candidate=float(expected.get("threshold_candidate", 0.0)),
        threshold_designation=_normalize_scalar(expected.get("threshold_designation")),
        output_base_dir=_resolve_path(project_root, _normalize_scalar(output.get("base_dir"))),
        source_files={str(key): _normalize_scalar(value) for key, value in source_files.items()},
        workbook_name=_normalize_scalar(workbook.get("filename", "validation_error_human_review.xlsx")),
        thumbnail_width=int(workbook.get("thumbnail_width", 300)),
        thumbnail_height=int(workbook.get("thumbnail_height", 190)),
        options=options,
    )

    digest = config.expected_source_zip_sha256
    if len(digest) != 64 or any(character not in "0123456789ABCDEF" for character in digest):
        raise HumanReviewError("config expected_source.zip_sha256 must be 64 uppercase hex characters")
    if not config.expected_source_zip_name.lower().endswith(".zip"):
        raise HumanReviewError("config expected_source.zip_filename must be a ZIP filename")
    if not config.expected_gate_classification:
        raise HumanReviewError("config expected_source.gate_classification is required")
    for field_name, value in (
        ("case_count", config.expected_case_count),
        ("validation_image_count", config.expected_validation_image_count),
        (
            "validation_ground_truth_instances",
            config.expected_validation_ground_truth_instances,
        ),
        ("raw_prediction_count", config.expected_raw_prediction_count),
    ):
        if value <= 0:
            raise HumanReviewError(f"config expected_source.{field_name} must be positive")
    if not 0.0 <= config.threshold_candidate <= 1.0:
        raise HumanReviewError("config threshold_candidate must be between 0 and 1")
    if config.threshold_designation != "BALANCED_VALIDATION_THRESHOLD_CANDIDATE":
        raise HumanReviewError(
            "config threshold_designation must remain BALANCED_VALIDATION_THRESHOLD_CANDIDATE"
        )
    if not config.workbook_name.lower().endswith(".xlsx"):
        raise HumanReviewError("config workbook.filename must end with .xlsx")
    if config.thumbnail_width <= 0 or config.thumbnail_height <= 0:
        raise HumanReviewError("config workbook thumbnail dimensions must be positive")
    for key, relative in config.source_files.items():
        try:
            _assert_validation_only_relpath(relative, f"source_files.{key}")
        except HumanReviewError as exc:
            raise HumanReviewError(f"invalid config source file path: {exc}") from exc
    required_values = {
        "review_status": {"pending", "reviewed", "needs_adjudication"},
        "annotation_quality": {"correct", "questionable", "defect_suspected", "not_applicable"},
        "recommended_action": {"no_action", "create_annotation_correction_proposal"},
        "retraining_priority": {"not_applicable", "low", "medium", "high"},
        "correction_proposal_required": {"no", "yes"},
    }
    for option_key, required_values_for_key in required_values.items():
        missing_values = sorted(required_values_for_key - set(config.options[option_key]))
        if missing_values:
            raise HumanReviewError(
                f"config options.{option_key} missing required values: {missing_values}"
            )
    return config


def _canonical_row_payload(row: Mapping[str, Any], columns: Sequence[str]) -> str:
    return "\x1f".join(_normalize_scalar(row.get(column, "")) for column in columns)


def source_case_fingerprint(row: Mapping[str, Any]) -> str:
    """Create a deterministic fingerprint for immutable source fields."""

    return _sha256_text(_canonical_row_payload(row, SOURCE_FINGERPRINT_COLUMNS))


def logical_fingerprint(frame: pd.DataFrame) -> str:
    """Create a deterministic canonical CSV logical fingerprint."""

    normalized = frame.loc[:, list(CANONICAL_COLUMNS)].fillna("").astype(str)
    payload = "\n".join(
        _canonical_row_payload(row, CANONICAL_COLUMNS)
        for row in normalized.to_dict(orient="records")
    )
    return _sha256_text(payload)


def _format_float(value: Any, digits: int = 6) -> str:
    try:
        number = float(value)
    except (TypeError, ValueError):
        number = 0.0
    if not math.isfinite(number):
        number = 0.0
    return f"{number:.{digits}f}"


def _assert_validation_only_relpath(value: str, label: str) -> None:
    pure = PurePosixPath(value.replace("\\", "/"))
    if pure.is_absolute() or ".." in pure.parts:
        raise HumanReviewError(f"{label} must be a safe relative path: {value}")
    if any(part.lower() == "test" for part in pure.parts):
        raise HumanReviewError(f"test-split path is forbidden in {label}: {value}")


def _assert_image_id(image_id: str) -> None:
    pure = PurePosixPath(image_id.replace("\\", "/"))
    if pure.name != image_id or pure.is_absolute() or ".." in pure.parts:
        raise HumanReviewError(f"image_id must be a filename only: {image_id}")


def _numeric_series(frame: pd.DataFrame, column: str) -> pd.Series:
    if column not in frame.columns:
        return pd.Series([0.0] * len(frame), index=frame.index, dtype=float)
    return pd.to_numeric(frame[column], errors="coerce").fillna(0.0)


def _box_iou(left: Mapping[str, Any], right: Mapping[str, Any]) -> float:
    left_x1, left_y1, left_x2, left_y2 = (
        float(left[key]) for key in ("x1", "y1", "x2", "y2")
    )
    right_x1, right_y1, right_x2, right_y2 = (
        float(right[key]) for key in ("x1", "y1", "x2", "y2")
    )
    intersection_width = max(0.0, min(left_x2, right_x2) - max(left_x1, right_x1))
    intersection_height = max(0.0, min(left_y2, right_y2) - max(left_y1, right_y1))
    intersection = intersection_width * intersection_height
    left_area = max(0.0, left_x2 - left_x1) * max(0.0, left_y2 - left_y1)
    right_area = max(0.0, right_x2 - right_x1) * max(0.0, right_y2 - right_y1)
    union = left_area + right_area - intersection
    return intersection / union if union > 0.0 else 0.0


def _max_pairwise_iou(predictions: pd.DataFrame, ground_truth: pd.DataFrame) -> float:
    if predictions.empty or ground_truth.empty:
        return 0.0
    return max(
        _box_iou(prediction, target)
        for prediction in predictions.to_dict(orient="records")
        for target in ground_truth.to_dict(orient="records")
    )


def _draw_overlay(
    source_image: Path,
    output_image: Path,
    ground_truth: pd.DataFrame,
    predictions: pd.DataFrame,
    threshold: float,
) -> None:
    with Image.open(source_image) as opened:
        image = opened.convert("RGB")
    draw = ImageDraw.Draw(image)

    for row in ground_truth.to_dict(orient="records"):
        box = tuple(float(row[key]) for key in ("x1", "y1", "x2", "y2"))
        draw.rectangle(box, outline=(0, 220, 80), width=3)
        draw.text((box[0], max(0.0, box[1] - 14.0)), "GT", fill=(0, 220, 80))

    active = predictions[_numeric_series(predictions, "confidence") >= threshold]
    for row in active.to_dict(orient="records"):
        box = tuple(float(row[key]) for key in ("x1", "y1", "x2", "y2"))
        confidence = float(row["confidence"])
        draw.rectangle(box, outline=(235, 45, 55), width=3)
        draw.text(
            (box[0], max(0.0, box[1] - 14.0)),
            f"P {confidence:.2f}",
            fill=(235, 45, 55),
        )

    draw.rectangle((0, 0, min(image.width, 520), 25), fill=(20, 20, 20))
    draw.text(
        (6, 6),
        f"Validation review overlay | candidate threshold={threshold:.2f}",
        fill=(255, 255, 255),
    )
    output_image.parent.mkdir(parents=True, exist_ok=True)
    image.save(output_image, format="JPEG", quality=92, optimize=True)


def _fit_xl_image(path: Path, max_width: int, max_height: int) -> XLImage:
    image = XLImage(str(path))
    scale = min(max_width / image.width, max_height / image.height, 1.0)
    image.width = int(image.width * scale)
    image.height = int(image.height * scale)
    return image


def _add_named_range_validation(
    sheet: Any,
    column_letter: str,
    defined_name: str,
    row_end: int,
) -> None:
    """Attach controlled Excel validation through a workbook named range."""

    validation = DataValidation(
        type="list",
        formula1=f"={defined_name}",
        allow_blank=True,
    )
    validation.error = "Please select an approved controlled value."
    validation.errorTitle = "Invalid FleetVision review value"
    validation.prompt = "Select one approved value from the list."
    validation.promptTitle = "FleetVision Phase 04.5L"
    sheet.add_data_validation(validation)
    validation.add(f"{column_letter}2:{column_letter}{row_end}")


def _build_workbook(
    path: Path,
    rows: list[dict[str, str]],
    batch_root: Path,
    config: ReviewConfig,
    source_manifest: Mapping[str, Any],
) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    instructions = workbook.create_sheet("Instructions")
    instructions["A1"] = "FleetVision Phase 04.5L — Validation Error Human Review"
    instructions["A1"].font = Font(size=16, bold=True)
    instructions["A3"] = "Purpose"
    instructions["B3"] = "Human-review validation-only baseline error cases."
    instructions["A4"] = "Critical boundary"
    instructions["B4"] = "Do not use test data, modify GT, start training, or approve deployment."
    instructions["A5"] = "Threshold"
    instructions["B5"] = (
        f"{config.threshold_candidate:.2f} is only "
        f"{config.threshold_designation}; it is not a deployment threshold."
    )
    instructions["A7"] = "Required completion"
    instructions["B7"] = "Every row must be reviewed; pending or needs_adjudication blocks final validation."
    instructions["A8"] = "Annotation defect"
    instructions["B8"] = "Create a correction proposal only. Do not edit GT or canonical COCO."
    instructions.column_dimensions["A"].width = 24
    instructions.column_dimensions["B"].width = 105
    for cell in instructions["A"]:
        cell.font = Font(bold=True)
    instructions.protection.sheet = True

    sheet = workbook.create_sheet("Review_Cases")
    headers = PREVIEW_HEADERS + CANONICAL_COLUMNS
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.protection = Protection(locked=True)
    sheet.freeze_panes = "C2"
    sheet.auto_filter.ref = f"A1:{sheet.cell(row=1, column=len(headers)).coordinate}"
    sheet.row_dimensions[1].height = 34
    sheet.column_dimensions["A"].width = 42
    sheet.column_dimensions["B"].width = 42

    source_fill = PatternFill("solid", fgColor="D9EAF7")
    human_fill = PatternFill("solid", fgColor="FFF2CC")
    human_start_column = 3 + len(SOURCE_COLUMNS)
    header_to_column: dict[str, int] = {
        header: index for index, header in enumerate(headers, start=1)
    }

    for row_index, row in enumerate(rows, start=2):
        for column_index, column in enumerate(CANONICAL_COLUMNS, start=3):
            value = row[column]
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if column in SOURCE_COLUMNS:
                cell.fill = source_fill
                cell.protection = Protection(locked=True)
            else:
                cell.fill = human_fill
                cell.protection = Protection(locked=False)

        original = batch_root / row["original_image_relpath"]
        overlay = batch_root / row["overlay_image_relpath"]
        original_image = _fit_xl_image(original, config.thumbnail_width, config.thumbnail_height)
        overlay_image = _fit_xl_image(overlay, config.thumbnail_width, config.thumbnail_height)
        sheet.add_image(original_image, f"A{row_index}")
        sheet.add_image(overlay_image, f"B{row_index}")
        sheet.row_dimensions[row_index].height = max(145, config.thumbnail_height * 0.78)

        for column in ("original_image_relpath", "overlay_image_relpath"):
            cell = sheet.cell(row=row_index, column=header_to_column[column])
            cell.hyperlink = "../" + row[column].replace("\\", "/")
            cell.style = "Hyperlink"
            cell.protection = Protection(locked=True)

    for column_index in range(3, len(headers) + 1):
        header = sheet.cell(row=1, column=column_index).value
        width = 18
        if header in {"image_id", "image_filename", "auto_error_detail_ids"}:
            width = 30
        elif header in {"original_image_relpath", "overlay_image_relpath", "review_notes"}:
            width = 42
        elif header in {"source_case_fingerprint", "source_04_5k_zip_sha256"}:
            width = 26
        sheet.column_dimensions[sheet.cell(row=1, column=column_index).column_letter].width = width

    option_sheet = workbook.create_sheet("Option_Lists")
    option_defined_names: dict[str, str] = {}
    for column_index, (option_key, values) in enumerate(config.options.items(), start=1):
        option_sheet.cell(row=1, column=column_index, value=option_key)
        for option_row_index, value in enumerate(values, start=2):
            option_sheet.cell(row=option_row_index, column=column_index, value=value)
        column_letter = get_column_letter(column_index)
        defined_name = f"phase04_5l_option_{option_key}"
        reference = (
            f"'{option_sheet.title}'!${column_letter}$2:"
            f"${column_letter}${len(values) + 1}"
        )
        workbook.defined_names.add(DefinedName(defined_name, attr_text=reference))
        option_defined_names[option_key] = defined_name
    option_sheet.sheet_state = "hidden"
    option_sheet.protection.sheet = True

    row_end = len(rows) + 1
    validation_columns = {
        "review_status": "review_status",
        "error_disposition": "error_disposition",
        "primary_root_cause": "primary_root_cause",
        "secondary_root_cause": "secondary_root_cause",
        "annotation_quality": "annotation_quality",
        "annotation_defect_type": "annotation_defect_type",
        "recommended_action": "recommended_action",
        "retraining_priority": "retraining_priority",
        "correction_proposal_required": "correction_proposal_required",
    }
    for column, option_key in validation_columns.items():
        column_letter = sheet.cell(row=1, column=header_to_column[column]).column_letter
        _add_named_range_validation(
            sheet,
            column_letter,
            option_defined_names[option_key],
            row_end,
        )

    status_letter = sheet.cell(row=1, column=header_to_column["review_status"]).column_letter
    sheet.conditional_formatting.add(
        f"{status_letter}2:{status_letter}{row_end}",
        FormulaRule(
            formula=[f'${status_letter}2="pending"'],
            fill=PatternFill("solid", fgColor="F4CCCC"),
        ),
    )
    sheet.conditional_formatting.add(
        f"{status_letter}2:{status_letter}{row_end}",
        FormulaRule(
            formula=[f'${status_letter}2="reviewed"'],
            fill=PatternFill("solid", fgColor="D9EAD3"),
        ),
    )
    sheet.protection.sheet = True
    sheet.protection.password = "FleetVision"
    sheet.protection.autoFilter = False
    sheet.protection.sort = False

    manifest_sheet = workbook.create_sheet("Manifest")
    manifest_sheet.append(["key", "value"])
    for key, value in sorted(source_manifest.items()):
        if isinstance(value, (dict, list)):
            value = json.dumps(value, ensure_ascii=False, sort_keys=True)
        manifest_sheet.append([key, _normalize_scalar(value)])
    manifest_sheet.column_dimensions["A"].width = 38
    manifest_sheet.column_dimensions["B"].width = 110
    manifest_sheet.protection.sheet = True

    progress = workbook.create_sheet("Progress_Summary")
    progress.append(["metric", "value"])
    progress.append(["total_cases", len(rows)])
    progress.append(["reviewed", f'=COUNTIF(Review_Cases!${status_letter}$2:${status_letter}${row_end},"reviewed")'])
    progress.append(["pending", f'=COUNTIF(Review_Cases!${status_letter}$2:${status_letter}${row_end},"pending")'])
    progress.append(
        [
            "needs_adjudication",
            f'=COUNTIF(Review_Cases!${status_letter}$2:${status_letter}${row_end},"needs_adjudication")',
        ]
    )
    progress.append(["completion_rate", "=IF(B2=0,0,B3/B2)"])
    progress["B6"].number_format = "0.0%"
    progress.column_dimensions["A"].width = 28
    progress.column_dimensions["B"].width = 18
    progress.protection.sheet = True

    if tuple(workbook.sheetnames) != WORKBOOK_SHEETS:
        raise HumanReviewError(f"unexpected workbook sheet order: {workbook.sheetnames}")
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _artifact_manifest_rows(batch_root: Path) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for path in sorted(candidate for candidate in batch_root.rglob("*") if candidate.is_file()):
        relative = path.relative_to(batch_root).as_posix()
        rows.append(
            {
                "relative_path": relative,
                "size_bytes": str(path.stat().st_size),
                "sha256": sha256_file(path),
            }
        )
    return rows


def prepare_review_package(
    config: ReviewConfig,
    source_root: Path,
    source_zip: Path,
    validation_images_dir: Path,
    batch_id: str,
) -> PrepareResult:
    """Prepare a no-overwrite 130-case review workbook and review assets."""

    source_root = source_root.resolve()
    source_zip = source_zip.resolve()
    validation_images_dir = validation_images_dir.resolve()
    batch_id = _normalize_scalar(batch_id)
    if not batch_id or any(character not in "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789_-" for character in batch_id):
        raise HumanReviewError("batch_id may contain only letters, digits, underscore, and hyphen")
    if not source_zip.is_file():
        raise HumanReviewError(f"source ZIP not found: {source_zip}")
    if source_zip.name != config.expected_source_zip_name:
        raise HumanReviewError(
            f"source ZIP filename mismatch: expected={config.expected_source_zip_name} actual={source_zip.name}"
        )
    actual_zip_hash = sha256_file(source_zip)
    if actual_zip_hash != config.expected_source_zip_sha256:
        raise HumanReviewError(
            f"source ZIP SHA256 mismatch: expected={config.expected_source_zip_sha256} actual={actual_zip_hash}"
        )
    if not validation_images_dir.is_dir():
        raise HumanReviewError(f"validation image directory not found: {validation_images_dir}")
    if any(part.lower() == "test" for part in validation_images_dir.parts):
        raise HumanReviewError(
            f"test-split directory is forbidden for validation review: {validation_images_dir}"
        )

    source_paths = {
        key: source_root / relative for key, relative in config.source_files.items()
    }
    for key, path in source_paths.items():
        if not path.is_file():
            raise HumanReviewError(f"required 04.5K artifact missing ({key}): {path}")
    _verify_source_artifacts_against_zip(source_zip, source_paths, config.source_files)
    _verify_04_5k_manifest_contract(source_paths, config.source_files)

    gate = json.loads(source_paths["gate_result"].read_text(encoding="utf-8"))
    expected_gate_values = {
        "gate_id": "04.5K",
        "outcome": "PASS",
        "classification": config.expected_gate_classification,
        "allowed_split": "valid",
        "validation_image_count": config.expected_validation_image_count,
        "validation_ground_truth_instances": (
            config.expected_validation_ground_truth_instances
        ),
        "raw_prediction_count": config.expected_raw_prediction_count,
        "test_set_used_for_tuning": False,
        "test_set_read": False,
        "training_started": False,
        "annotation_modified": False,
        "deployment_acceptance": "NOT_YET_APPROVED",
    }
    for key, expected_value in expected_gate_values.items():
        if gate.get(key) != expected_value:
            raise HumanReviewError(
                f"04.5K gate boundary mismatch: {key}="
                f"{gate.get(key)!r}, expected={expected_value!r}"
            )
    try:
        balanced_threshold = float(gate.get("balanced_threshold"))
    except (TypeError, ValueError) as exc:
        raise HumanReviewError("04.5K gate balanced_threshold is invalid") from exc
    if not math.isclose(
        balanced_threshold,
        config.threshold_candidate,
        rel_tol=0.0,
        abs_tol=1e-12,
    ):
        raise HumanReviewError(
            "04.5K gate boundary mismatch: balanced_threshold="
            f"{balanced_threshold!r}, expected={config.threshold_candidate!r}"
        )

    worklist = _load_csv(source_paths["worklist"], WORKLIST_REQUIRED_COLUMNS)
    predictions = _load_csv(source_paths["predictions"], PREDICTION_REQUIRED_COLUMNS)
    ground_truth = _load_csv(source_paths["ground_truth"], GROUND_TRUTH_REQUIRED_COLUMNS)
    for label, frame in (("predictions", predictions), ("ground_truth", ground_truth)):
        split_values = set(frame["split"].astype(str).str.strip())
        if split_values != {"valid"}:
            raise HumanReviewError(
                f"{label} split must contain only valid; got={sorted(split_values)}"
            )
    if len(predictions) != config.expected_raw_prediction_count:
        raise HumanReviewError(
            "raw prediction count mismatch: "
            f"expected={config.expected_raw_prediction_count} actual={len(predictions)}"
        )
    if len(ground_truth) != config.expected_validation_ground_truth_instances:
        raise HumanReviewError(
            "validation GT count mismatch: "
            f"expected={config.expected_validation_ground_truth_instances} "
            f"actual={len(ground_truth)}"
        )
    validation_image_ids = set(predictions["image_id"]) | set(ground_truth["image_id"])
    if len(validation_image_ids) != config.expected_validation_image_count:
        raise HumanReviewError(
            "validation image coverage mismatch: "
            f"expected={config.expected_validation_image_count} "
            f"actual={len(validation_image_ids)}"
        )
    if len(worklist) != config.expected_case_count:
        raise HumanReviewError(
            f"review case count mismatch: expected={config.expected_case_count} actual={len(worklist)}"
        )
    if worklist["image_id"].nunique(dropna=False) != config.expected_case_count:
        raise HumanReviewError("worklist image_id values must be unique and non-blank")

    final_root = config.output_base_dir / batch_id
    if final_root.exists():
        raise HumanReviewError(f"review batch already exists; overwrite is forbidden: {final_root}")
    final_root.parent.mkdir(parents=True, exist_ok=True)
    staging_root = final_root.parent / f".{batch_id}.staging-{uuid.uuid4().hex[:8]}"
    if staging_root.exists():
        raise HumanReviewError(f"unexpected staging collision: {staging_root}")

    try:
        original_dir = staging_root / "assets/original"
        overlay_dir = staging_root / "assets/overlay"
        workbook_dir = staging_root / "workbook"
        manifest_dir = staging_root / "manifest"
        original_dir.mkdir(parents=True, exist_ok=False)
        overlay_dir.mkdir(parents=True, exist_ok=False)
        workbook_dir.mkdir(parents=True, exist_ok=False)
        manifest_dir.mkdir(parents=True, exist_ok=False)

        canonical_rows: list[dict[str, str]] = []
        for worklist_row in worklist.to_dict(orient="records"):
            image_id = _normalize_scalar(worklist_row["image_id"])
            _assert_image_id(image_id)
            source_image = validation_images_dir / image_id
            if not source_image.is_file():
                raise HumanReviewError(f"validation image missing for review case: {source_image}")
            copied_image = original_dir / image_id
            copied_image.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source_image, copied_image)

            image_predictions = predictions[predictions["image_id"] == image_id].copy()
            image_gt = ground_truth[ground_truth["image_id"] == image_id].copy()
            active_predictions = image_predictions[
                _numeric_series(image_predictions, "confidence") >= config.threshold_candidate
            ]

            preliminary: dict[str, str] = {
                "schema_version": config.schema_version,
                "review_batch_id": batch_id,
                "source_04_5k_zip_sha256": actual_zip_hash,
                "image_id": image_id,
                "image_filename": image_id,
                "auto_error_category": _normalize_scalar(worklist_row["primary_error_type"]),
                "auto_error_detail_ids": _normalize_scalar(worklist_row["error_types"]),
                "error_case_count": _normalize_scalar(worklist_row["error_case_count"]),
                "ground_truth_error_count": _normalize_scalar(worklist_row["ground_truth_error_count"]),
                "prediction_error_count": _normalize_scalar(worklist_row["prediction_error_count"]),
                "gt_count": str(len(image_gt)),
                "prediction_count": str(len(active_predictions)),
                "max_prediction_confidence": _format_float(
                    _numeric_series(image_predictions, "confidence").max() if not image_predictions.empty else 0.0
                ),
                "best_iou": _format_float(_max_pairwise_iou(active_predictions, image_gt)),
                "threshold_candidate": f"{config.threshold_candidate:.2f}",
                "threshold_designation": config.threshold_designation,
                "original_image_relpath": f"assets/original/{image_id}",
                "overlay_image_relpath": "",
            }
            provisional_fingerprint = source_case_fingerprint(
                {**preliminary, "overlay_image_relpath": "pending"}
            )
            review_case_id = f"l_{_sha256_text(image_id + provisional_fingerprint)[:16].lower()}"
            overlay_relpath = f"assets/overlay/{review_case_id}.jpg"
            preliminary["overlay_image_relpath"] = overlay_relpath
            preliminary["review_case_id"] = review_case_id
            preliminary["source_case_fingerprint"] = source_case_fingerprint(preliminary)
            for column in HUMAN_COLUMNS:
                preliminary[column] = "pending" if column == "review_status" else ""
            canonical_rows.append({column: preliminary[column] for column in CANONICAL_COLUMNS})

            _draw_overlay(
                copied_image,
                staging_root / overlay_relpath,
                image_gt,
                image_predictions,
                config.threshold_candidate,
            )

        source_artifact_hashes = {
            key: {
                "relative_path": path.relative_to(source_root).as_posix(),
                "sha256": sha256_file(path),
                "size_bytes": path.stat().st_size,
            }
            for key, path in source_paths.items()
        }
        source_manifest = {
            "gate_id": "04.5L-PREP",
            "classification": "VALIDATION_ERROR_HUMAN_REVIEW_PACKAGE_PREPARED",
            "created_at_utc": _utc_now(),
            "review_batch_id": batch_id,
            "schema_version": config.schema_version,
            "source_zip_filename": source_zip.name,
            "source_zip_sha256": actual_zip_hash,
            "source_gate_classification": config.expected_gate_classification,
            "case_count": len(canonical_rows),
            "threshold_candidate": f"{config.threshold_candidate:.2f}",
            "threshold_designation": config.threshold_designation,
            "test_split_read": False,
            "model_inference_executed": False,
            "training_started": False,
            "annotation_modified": False,
            "retraining_status": "NOT_YET_APPROVED",
            "deployment_acceptance": "NOT_YET_APPROVED",
            "source_artifacts": source_artifact_hashes,
        }
        _json_dump(manifest_dir / "source_manifest.json", source_manifest)

        workbook_path = workbook_dir / config.workbook_name
        _build_workbook(workbook_path, canonical_rows, staging_root, config, source_manifest)

        artifact_rows = _artifact_manifest_rows(staging_root)
        _write_csv(
            manifest_dir / "asset_manifest.csv",
            artifact_rows,
            ("relative_path", "size_bytes", "sha256"),
        )
        checksum_rows = _artifact_manifest_rows(staging_root)
        checksum_text = "\n".join(
            f"{row['sha256']}  {row['relative_path']}" for row in checksum_rows
        ) + "\n"
        (manifest_dir / "checksums.sha256").write_text(checksum_text, encoding="utf-8")

        staging_root.replace(final_root)
    except Exception:
        shutil.rmtree(staging_root, ignore_errors=True)
        raise

    workbook_path = final_root / "workbook" / config.workbook_name
    return PrepareResult(
        batch_root=final_root,
        workbook_path=workbook_path,
        row_count=config.expected_case_count,
        asset_count=config.expected_case_count * 2,
        source_zip_sha256=actual_zip_hash,
    )


def read_workbook_dataframe(workbook_path: Path) -> pd.DataFrame:
    """Read canonical source and human columns from the protected review sheet."""

    if not workbook_path.is_file():
        raise HumanReviewError(f"review workbook not found: {workbook_path}")
    workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    if tuple(workbook.sheetnames) != WORKBOOK_SHEETS:
        raise HumanReviewError(f"workbook sheet contract mismatch: {workbook.sheetnames}")
    sheet = workbook["Review_Cases"]
    headers = [_normalize_scalar(cell.value) for cell in sheet[1]]
    if headers[:2] != list(PREVIEW_HEADERS):
        raise HumanReviewError(f"workbook preview-column contract mismatch: {headers[:2]}")
    if headers[2:] != list(CANONICAL_COLUMNS):
        raise HumanReviewError("workbook canonical-column order mismatch")

    rows: list[dict[str, str]] = []
    for row_index in range(2, sheet.max_row + 1):
        row: dict[str, str] = {}
        blank = True
        for offset, column in enumerate(CANONICAL_COLUMNS, start=3):
            value = _normalize_scalar(sheet.cell(row=row_index, column=offset).value)
            row[column] = value
            if value:
                blank = False
        if not blank:
            rows.append(row)
    return pd.DataFrame(rows, columns=CANONICAL_COLUMNS).fillna("").astype(str)


def _semantic_issues(
    frame: pd.DataFrame,
    config: ReviewConfig,
    *,
    require_complete: bool,
    batch_root: Path | None = None,
) -> list[dict[str, str]]:
    issues: list[dict[str, str]] = []

    def add(code: str, message: str, row_number: int | str = "") -> None:
        issues.append(
            {
                "row_number": _normalize_scalar(row_number),
                "error_code": code,
                "message": message,
            }
        )

    if frame.columns.tolist() != list(CANONICAL_COLUMNS):
        add("CANONICAL_COLUMN_CONTRACT_MISMATCH", "canonical columns or order do not match Phase 04.5L")
        return issues
    if len(frame) != config.expected_case_count:
        add(
            "ROW_COUNT_MISMATCH",
            f"expected {config.expected_case_count} rows, got {len(frame)}",
        )
    for column in ("review_case_id", "image_id"):
        if frame[column].eq("").any():
            add("BLANK_IDENTITY", f"{column} contains blank values")
        if frame[column].nunique(dropna=False) != len(frame):
            add("DUPLICATE_IDENTITY", f"{column} values must be unique")

    expected_threshold = f"{config.threshold_candidate:.2f}"
    if set(frame["threshold_candidate"]) != {expected_threshold}:
        add("THRESHOLD_MISMATCH", f"threshold_candidate must be {expected_threshold}")
    if set(frame["threshold_designation"]) != {config.threshold_designation}:
        add("THRESHOLD_DESIGNATION_MISMATCH", "threshold designation must remain validation-only")
    if set(value.upper() for value in frame["source_04_5k_zip_sha256"]) != {
        config.expected_source_zip_sha256
    }:
        add("SOURCE_ZIP_IDENTITY_MISMATCH", "source 04.5K ZIP SHA256 mismatch")
    if set(frame["schema_version"]) != {config.schema_version}:
        add("SCHEMA_VERSION_MISMATCH", f"schema_version must be {config.schema_version}")
    if frame["review_batch_id"].nunique(dropna=False) != 1 or frame["review_batch_id"].eq("").any():
        add("BATCH_ID_MISMATCH", "canonical CSV must contain exactly one non-blank review_batch_id")

    for index, row in frame.iterrows():
        row_number = index + 2
        row_dict = {column: _normalize_scalar(row[column]) for column in CANONICAL_COLUMNS}
        if source_case_fingerprint(row_dict) != row_dict["source_case_fingerprint"].upper():
            add("SOURCE_FINGERPRINT_MISMATCH", "immutable source fields were changed", row_number)
        for path_column in ("original_image_relpath", "overlay_image_relpath"):
            value = row_dict[path_column]
            try:
                _assert_validation_only_relpath(value, path_column)
            except HumanReviewError as exc:
                add("FORBIDDEN_PATH", str(exc), row_number)
            if batch_root is not None and not (batch_root / value).is_file():
                add("MISSING_REVIEW_ASSET", f"missing asset: {value}", row_number)

        for column, option_key in (
            ("review_status", "review_status"),
            ("error_disposition", "error_disposition"),
            ("primary_root_cause", "primary_root_cause"),
            ("secondary_root_cause", "secondary_root_cause"),
            ("annotation_quality", "annotation_quality"),
            ("annotation_defect_type", "annotation_defect_type"),
            ("recommended_action", "recommended_action"),
            ("retraining_priority", "retraining_priority"),
            ("correction_proposal_required", "correction_proposal_required"),
        ):
            value = row_dict[column]
            if value and value not in config.options[option_key]:
                add("INVALID_CONTROLLED_VALUE", f"{column}={value!r} is not approved", row_number)

        status = row_dict["review_status"]
        if require_complete and status != "reviewed":
            add("INCOMPLETE_REVIEW", f"review_status must be reviewed, got {status!r}", row_number)
        if status == "reviewed":
            for required in (
                "error_disposition",
                "primary_root_cause",
                "annotation_quality",
                "recommended_action",
                "retraining_priority",
                "correction_proposal_required",
                "reviewer",
                "reviewed_at_utc",
            ):
                if not row_dict[required]:
                    add("MISSING_REVIEW_FIELD", f"{required} is required for reviewed rows", row_number)
            if row_dict["reviewed_at_utc"]:
                try:
                    parsed = datetime.fromisoformat(row_dict["reviewed_at_utc"].replace("Z", "+00:00"))
                    if parsed.tzinfo is None:
                        raise ValueError("timezone required")
                except ValueError:
                    add("INVALID_REVIEW_TIMESTAMP", "reviewed_at_utc must be timezone-aware ISO 8601", row_number)

        annotation_quality = row_dict["annotation_quality"]
        annotation_defect_type = row_dict["annotation_defect_type"]
        correction_required = row_dict["correction_proposal_required"]
        correction_action = row_dict["recommended_action"] == "create_annotation_correction_proposal"

        if annotation_quality == "defect_suspected":
            if correction_required != "yes":
                add("CORRECTION_PROPOSAL_REQUIRED", "defect_suspected requires proposal=yes", row_number)
            if not correction_action:
                add(
                    "CORRECTION_ACTION_REQUIRED",
                    "defect_suspected requires create_annotation_correction_proposal",
                    row_number,
                )
            if annotation_defect_type in {"", "none"} or not row_dict["review_notes"]:
                add(
                    "CORRECTION_EVIDENCE_REQUIRED",
                    "a specific annotation_defect_type and review_notes are required",
                    row_number,
                )
        else:
            if correction_required == "yes":
                add(
                    "CONTRADICTORY_CORRECTION",
                    "correction proposal=yes requires annotation_quality=defect_suspected",
                    row_number,
                )
            if correction_action:
                add(
                    "CONTRADICTORY_CORRECTION_ACTION",
                    "create_annotation_correction_proposal requires annotation_quality=defect_suspected",
                    row_number,
                )
            if annotation_defect_type not in {"", "none"}:
                add(
                    "CONTRADICTORY_DEFECT_TYPE",
                    "annotation_defect_type must be none unless annotation_quality=defect_suspected",
                    row_number,
                )
        if row_dict["retraining_priority"] == "high" and not row_dict["review_notes"]:
            add("HIGH_PRIORITY_NOTES_REQUIRED", "high retraining priority requires review_notes", row_number)

    return issues


def validate_canonical_dataframe(
    frame: pd.DataFrame,
    config: ReviewConfig,
    *,
    require_complete: bool = True,
    batch_root: Path | None = None,
) -> ValidationResult:
    """Validate schema, source identity, review semantics, and safety boundaries."""

    normalized = frame.copy().fillna("").astype(str)
    original_columns = normalized.columns.tolist()
    schema_issues: list[dict[str, str]] = []
    if original_columns != list(CANONICAL_COLUMNS):
        schema_issues.append(
            {
                "row_number": "",
                "error_code": "CANONICAL_COLUMN_CONTRACT_MISMATCH",
                "message": "canonical columns or order do not match Phase 04.5L",
            }
        )
        normalized = normalized.reindex(columns=CANONICAL_COLUMNS, fill_value="")
    issues = schema_issues + _semantic_issues(
        normalized,
        config,
        require_complete=require_complete,
        batch_root=batch_root,
    )
    counts = {
        "reviewed": int((normalized["review_status"] == "reviewed").sum()),
        "pending": int((normalized["review_status"] == "pending").sum()),
        "needs_adjudication": int((normalized["review_status"] == "needs_adjudication").sum()),
        "annotation_defect_suspected": int(
            (normalized["annotation_quality"] == "defect_suspected").sum()
        ),
        "high_retraining_priority": int(
            (normalized["retraining_priority"] == "high").sum()
        ),
    }
    return ValidationResult(
        passed=not issues,
        row_count=len(normalized),
        issue_count=len(issues),
        issues=tuple(issues),
        logical_fingerprint=logical_fingerprint(normalized),
        counts=counts,
    )


def export_review_workbook(
    config: ReviewConfig,
    workbook_path: Path,
    output_csv: Path,
) -> ExportResult:
    """Export a completed workbook transactionally after semantic validation."""

    output_csv = output_csv.resolve()
    if output_csv.exists():
        raise HumanReviewError(f"canonical output already exists; overwrite is forbidden: {output_csv}")
    frame = read_workbook_dataframe(workbook_path)
    validation = validate_canonical_dataframe(frame, config, require_complete=True)
    if not validation.passed:
        preview = "; ".join(issue["error_code"] for issue in validation.issues[:8])
        raise HumanReviewError(
            f"workbook cannot be exported; validation issues={validation.issue_count}: {preview}"
        )

    output_csv.parent.mkdir(parents=True, exist_ok=True)
    handle, temporary_name = tempfile.mkstemp(
        prefix=f".{output_csv.name}.staging-",
        suffix=".csv",
        dir=output_csv.parent,
    )
    os.close(handle)
    temporary = Path(temporary_name)
    try:
        frame.to_csv(temporary, index=False, encoding="utf-8-sig", lineterminator="\n")
        roundtrip = pd.read_csv(
            temporary,
            dtype=str,
            keep_default_na=False,
            encoding="utf-8-sig",
        ).fillna("").astype(str)
        if roundtrip.columns.tolist() != list(CANONICAL_COLUMNS) or not roundtrip.equals(frame):
            raise HumanReviewError("canonical CSV UTF-8-SIG round-trip mismatch")
        temporary.replace(output_csv)
    except Exception:
        temporary.unlink(missing_ok=True)
        raise

    return ExportResult(
        output_csv=output_csv,
        row_count=len(frame),
        logical_fingerprint=validation.logical_fingerprint,
    )


def validate_canonical_csv(
    config: ReviewConfig,
    canonical_csv: Path,
    *,
    workbook_path: Path | None = None,
    batch_root: Path | None = None,
) -> ValidationResult:
    """Validate canonical CSV and optionally require exact workbook mapping."""

    frame = _load_csv(canonical_csv, CANONICAL_COLUMNS)
    if frame.columns.tolist() != list(CANONICAL_COLUMNS):
        issues = (
            {
                "row_number": "",
                "error_code": "CANONICAL_COLUMN_CONTRACT_MISMATCH",
                "message": "canonical columns or order do not match Phase 04.5L",
            },
        )
        return ValidationResult(
            passed=False,
            row_count=len(frame),
            issue_count=1,
            issues=issues,
            logical_fingerprint="",
            counts={},
        )
    result = validate_canonical_dataframe(
        frame,
        config,
        require_complete=True,
        batch_root=batch_root,
    )
    issues = list(result.issues)
    if workbook_path is not None:
        workbook_frame = read_workbook_dataframe(workbook_path)
        if not workbook_frame.equals(frame):
            issues.append(
                {
                    "row_number": "",
                    "error_code": "WORKBOOK_CANONICAL_MAPPING_MISMATCH",
                    "message": "workbook source/human fields do not exactly match canonical CSV",
                }
            )
    return ValidationResult(
        passed=not issues,
        row_count=result.row_count,
        issue_count=len(issues),
        issues=tuple(issues),
        logical_fingerprint=result.logical_fingerprint,
        counts=result.counts,
    )


def _write_validation_outputs(
    result: ValidationResult,
    report_json: Path,
    errors_csv: Path,
) -> None:
    for path in (report_json, errors_csv):
        if path.exists():
            raise HumanReviewError(f"validation output already exists; overwrite is forbidden: {path}")
    report_json.parent.mkdir(parents=True, exist_ok=True)
    errors_csv.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "gate_id": "04.5L-VALIDATE",
        "outcome": "PASS" if result.passed else "BLOCKED",
        "classification": (
            "VALIDATION_ERROR_HUMAN_REVIEW_VERIFIED"
            if result.passed
            else "VALIDATION_ERROR_HUMAN_REVIEW_BLOCKED"
        ),
        "validated_at_utc": _utc_now(),
        "row_count": result.row_count,
        "issue_count": result.issue_count,
        "counts": dict(result.counts),
        "logical_fingerprint": result.logical_fingerprint,
        "test_set_read": False,
        "annotation_modified": False,
        "training_started": False,
        "retraining_status": "NOT_YET_APPROVED",
        "deployment_acceptance": "NOT_YET_APPROVED",
    }
    token = uuid.uuid4().hex[:8]
    staged_report = report_json.parent / f".{report_json.name}.staging-{token}"
    staged_errors = errors_csv.parent / f".{errors_csv.name}.staging-{token}"
    try:
        _json_dump(staged_report, payload)
        _write_csv(
            staged_errors,
            result.issues,
            ("row_number", "error_code", "message"),
        )
        staged_report.replace(report_json)
        staged_errors.replace(errors_csv)
    except Exception:
        staged_report.unlink(missing_ok=True)
        staged_errors.unlink(missing_ok=True)
        report_json.unlink(missing_ok=True)
        errors_csv.unlink(missing_ok=True)
        raise


def summarize_canonical_review(
    config: ReviewConfig,
    canonical_csv: Path,
    batch_root: Path,
) -> SummaryResult:
    """Create summary, action queue, and non-applied annotation proposals."""

    batch_root = batch_root.resolve()
    frame = _load_csv(canonical_csv, CANONICAL_COLUMNS)
    validation = validate_canonical_dataframe(
        frame,
        config,
        require_complete=True,
        batch_root=batch_root,
    )
    if not validation.passed:
        raise HumanReviewError(
            f"summary generation blocked by {validation.issue_count} canonical validation issues"
        )

    canonical_dir = batch_root / "canonical"
    reports_dir = batch_root / "reports"
    proposal_path = canonical_dir / "annotation_correction_proposals.csv"
    summary_json = reports_dir / "review_summary.json"
    summary_md = reports_dir / "review_summary.md"
    action_queue_path = reports_dir / "data_improvement_action_queue.csv"
    action_summary_path = reports_dir / "data_improvement_action_summary.csv"
    output_paths = (
        proposal_path,
        summary_json,
        summary_md,
        action_queue_path,
        action_summary_path,
    )
    existing = [path for path in output_paths if path.exists()]
    if existing:
        raise HumanReviewError(f"summary outputs already exist; overwrite is forbidden: {existing}")

    staging = batch_root / f".summary.staging-{uuid.uuid4().hex[:8]}"
    try:
        staging_canonical = staging / "canonical"
        staging_reports = staging / "reports"
        staging_canonical.mkdir(parents=True, exist_ok=False)
        staging_reports.mkdir(parents=True, exist_ok=False)

        correction_rows: list[dict[str, str]] = []
        action_rows: list[dict[str, str]] = []
        for row in frame.to_dict(orient="records"):
            review_case_id = row["review_case_id"]
            proposal_id = ""
            if row["correction_proposal_required"] == "yes":
                proposal_id = f"proposal_{_sha256_text(review_case_id)[:16].lower()}"
                correction_rows.append(
                    {
                        "proposal_id": proposal_id,
                        "review_case_id": review_case_id,
                        "image_id": row["image_id"],
                        "source_annotation_ids": "",
                        "proposal_type": row["annotation_defect_type"],
                        "proposal_description": row["review_notes"],
                        "evidence_original_relpath": row["original_image_relpath"],
                        "evidence_overlay_relpath": row["overlay_image_relpath"],
                        "reviewer": row["reviewer"],
                        "proposal_status": "PROPOSED_NOT_APPLIED",
                    }
                )
            if row["recommended_action"] != "no_action" or row["retraining_priority"] != "not_applicable":
                action_rows.append(
                    {
                        "action_id": f"action_{_sha256_text(review_case_id + row['recommended_action'])[:16].lower()}",
                        "review_case_id": review_case_id,
                        "image_id": row["image_id"],
                        "error_disposition": row["error_disposition"],
                        "primary_root_cause": row["primary_root_cause"],
                        "recommended_action": row["recommended_action"],
                        "retraining_priority": row["retraining_priority"],
                        "annotation_correction_proposal_id": proposal_id,
                        "action_rationale": row["review_notes"],
                        "source_original_relpath": row["original_image_relpath"],
                        "source_overlay_relpath": row["overlay_image_relpath"],
                        "reviewer": row["reviewer"],
                    }
                )

        proposal_columns = (
            "proposal_id",
            "review_case_id",
            "image_id",
            "source_annotation_ids",
            "proposal_type",
            "proposal_description",
            "evidence_original_relpath",
            "evidence_overlay_relpath",
            "reviewer",
            "proposal_status",
        )
        action_columns = (
            "action_id",
            "review_case_id",
            "image_id",
            "error_disposition",
            "primary_root_cause",
            "recommended_action",
            "retraining_priority",
            "annotation_correction_proposal_id",
            "action_rationale",
            "source_original_relpath",
            "source_overlay_relpath",
            "reviewer",
        )
        _write_csv(staging_canonical / proposal_path.name, correction_rows, proposal_columns)
        _write_csv(staging_reports / action_queue_path.name, action_rows, action_columns)

        action_counter = Counter(
            (row["recommended_action"], row["retraining_priority"]) for row in action_rows
        )
        action_summary_rows = [
            {
                "recommended_action": action,
                "retraining_priority": priority,
                "case_count": str(count),
            }
            for (action, priority), count in sorted(action_counter.items())
        ]
        _write_csv(
            staging_reports / action_summary_path.name,
            action_summary_rows,
            ("recommended_action", "retraining_priority", "case_count"),
        )

        summary = {
            "gate_id": "04.5L-SUMMARY",
            "classification": "VALIDATION_ERROR_HUMAN_REVIEW_SUMMARY_CREATED",
            "created_at_utc": _utc_now(),
            "row_count": len(frame),
            "logical_fingerprint": validation.logical_fingerprint,
            "error_disposition_counts": frame["error_disposition"].value_counts().sort_index().to_dict(),
            "primary_root_cause_counts": frame["primary_root_cause"].value_counts().sort_index().to_dict(),
            "annotation_quality_counts": frame["annotation_quality"].value_counts().sort_index().to_dict(),
            "recommended_action_counts": frame["recommended_action"].value_counts().sort_index().to_dict(),
            "retraining_priority_counts": frame["retraining_priority"].value_counts().sort_index().to_dict(),
            "action_queue_count": len(action_rows),
            "annotation_correction_proposal_count": len(correction_rows),
            "annotation_modified": False,
            "test_set_read": False,
            "training_started": False,
            "retraining_status": "NOT_YET_APPROVED",
            "deployment_acceptance": "NOT_YET_APPROVED",
        }
        _json_dump(staging_reports / summary_json.name, summary)
        markdown = [
            "# FleetVision Phase 04.5L Human Review Summary",
            "",
            f"- Reviewed cases: {len(frame)}",
            f"- Data-improvement actions: {len(action_rows)}",
            f"- Annotation correction proposals: {len(correction_rows)}",
            "- Annotation modified: false",
            "- Test set read: false",
            "- Training started: false",
            "- Retraining status: NOT_YET_APPROVED",
            "- Deployment acceptance: NOT_YET_APPROVED",
            "",
            "## Recommended action counts",
        ]
        for key, value in summary["recommended_action_counts"].items():
            markdown.append(f"- {key}: {value}")
        (staging_reports / summary_md.name).write_text("\n".join(markdown) + "\n", encoding="utf-8")

        canonical_dir.mkdir(parents=True, exist_ok=True)
        reports_dir.mkdir(parents=True, exist_ok=True)
        for staged, final in (
            (staging_canonical / proposal_path.name, proposal_path),
            (staging_reports / summary_json.name, summary_json),
            (staging_reports / summary_md.name, summary_md),
            (staging_reports / action_queue_path.name, action_queue_path),
            (staging_reports / action_summary_path.name, action_summary_path),
        ):
            staged.replace(final)
    except Exception:
        shutil.rmtree(staging, ignore_errors=True)
        for path in output_paths:
            path.unlink(missing_ok=True)
        raise
    finally:
        shutil.rmtree(staging, ignore_errors=True)

    return SummaryResult(
        output_paths=output_paths,
        action_count=len(action_rows),
        correction_proposal_count=len(correction_rows),
    )


def _common_parser(description: str) -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=description)
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG_PATH)
    parser.add_argument("--project-root", type=Path, default=Path.cwd())
    return parser


def main_prepare(argv: Sequence[str] | None = None) -> int:
    parser = _common_parser("Prepare FleetVision Phase 04.5L validation-error review package.")
    parser.add_argument("--source-root", type=Path, required=True)
    parser.add_argument("--source-zip", type=Path, required=True)
    parser.add_argument("--validation-images-dir", type=Path, required=True)
    parser.add_argument("--batch-id", required=True)
    args = parser.parse_args(argv)
    try:
        config = load_config(args.config, args.project_root)
        result = prepare_review_package(
            config,
            args.source_root,
            args.source_zip,
            args.validation_images_dir,
            args.batch_id,
        )
    except Exception as exc:
        print("Gate classification: VALIDATION_ERROR_HUMAN_REVIEW_PACKAGE_BLOCKED")
        print(f"Reason: {exc}")
        return 1
    print("=== FleetVision Phase 04.5L Review Package Preparation ===")
    print("Gate classification: VALIDATION_ERROR_HUMAN_REVIEW_PACKAGE_PREPARED")
    print(f"Batch root: {result.batch_root}")
    print(f"Workbook: {result.workbook_path}")
    print(f"Review cases: {result.row_count}")
    print(f"Review assets: {result.asset_count}")
    print("TEST_SPLIT_READ: NO")
    print("MODEL_INFERENCE_EXECUTED: NO")
    print("TRAINING_STARTED: NO")
    return 0


def main_export(argv: Sequence[str] | None = None) -> int:
    parser = _common_parser("Export completed Phase 04.5L workbook to canonical CSV.")
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--output-csv", type=Path, required=True)
    args = parser.parse_args(argv)
    try:
        config = load_config(args.config, args.project_root)
        result = export_review_workbook(config, args.workbook, args.output_csv)
    except Exception as exc:
        print("Gate classification: VALIDATION_ERROR_HUMAN_REVIEW_EXPORT_BLOCKED")
        print(f"Reason: {exc}")
        return 1
    print("=== FleetVision Phase 04.5L Canonical Export ===")
    print("Gate classification: VALIDATION_ERROR_HUMAN_REVIEW_CANONICAL_EXPORTED")
    print(f"Canonical CSV: {result.output_csv}")
    print(f"Rows: {result.row_count}")
    print(f"Logical fingerprint: {result.logical_fingerprint}")
    return 0


def main_validate(argv: Sequence[str] | None = None) -> int:
    parser = _common_parser("Validate Phase 04.5L canonical human-review CSV.")
    parser.add_argument("--canonical-csv", type=Path, required=True)
    parser.add_argument("--workbook", type=Path)
    parser.add_argument("--batch-root", type=Path)
    parser.add_argument("--report-json", type=Path)
    parser.add_argument("--errors-csv", type=Path)
    args = parser.parse_args(argv)
    if bool(args.report_json) != bool(args.errors_csv):
        parser.error("--report-json and --errors-csv must be supplied together")
    try:
        config = load_config(args.config, args.project_root)
        result = validate_canonical_csv(
            config,
            args.canonical_csv,
            workbook_path=args.workbook,
            batch_root=args.batch_root,
        )
        if args.report_json and args.errors_csv:
            _write_validation_outputs(result, args.report_json, args.errors_csv)
    except Exception as exc:
        print("Gate classification: VALIDATION_ERROR_HUMAN_REVIEW_BLOCKED")
        print(f"Reason: {exc}")
        return 1
    print("=== FleetVision Phase 04.5L Canonical Validation ===")
    print(
        "Gate classification: "
        + (
            "VALIDATION_ERROR_HUMAN_REVIEW_VERIFIED"
            if result.passed
            else "VALIDATION_ERROR_HUMAN_REVIEW_BLOCKED"
        )
    )
    print(f"Rows: {result.row_count}")
    print(f"Validation issues: {result.issue_count}")
    print(f"Logical fingerprint: {result.logical_fingerprint}")
    print("TEST_SET_READ: NO")
    print("ANNOTATION_MODIFIED: NO")
    print("TRAINING_STARTED: NO")
    print("RETRAINING_STATUS: NOT_YET_APPROVED")
    print("DEPLOYMENT_ACCEPTANCE: NOT_YET_APPROVED")
    return 0 if result.passed else 1


def main_summarize(argv: Sequence[str] | None = None) -> int:
    parser = _common_parser("Summarize verified Phase 04.5L review and build proposal queues.")
    parser.add_argument("--canonical-csv", type=Path, required=True)
    parser.add_argument("--batch-root", type=Path, required=True)
    args = parser.parse_args(argv)
    try:
        config = load_config(args.config, args.project_root)
        result = summarize_canonical_review(config, args.canonical_csv, args.batch_root)
    except Exception as exc:
        print("Gate classification: VALIDATION_ERROR_HUMAN_REVIEW_SUMMARY_BLOCKED")
        print(f"Reason: {exc}")
        return 1
    print("=== FleetVision Phase 04.5L Review Summary ===")
    print("Gate classification: VALIDATION_ERROR_HUMAN_REVIEW_SUMMARY_CREATED")
    print(f"Data-improvement actions: {result.action_count}")
    print(f"Annotation correction proposals: {result.correction_proposal_count}")
    for path in result.output_paths:
        print(f"Output: {path}")
    print("ANNOTATION_MODIFIED: NO")
    print("TRAINING_STARTED: NO")
    return 0
