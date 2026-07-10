"""Safely export Pilot 500 human review Excel edits back to canonical CSV."""

from __future__ import annotations

import argparse
import os
import sys
import tempfile
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

import pandas as pd
import yaml

from fleetvision.data.build_pilot_review_excel import WORKLIST_SHEET
from fleetvision.data.build_pilot_review_worklist import HUMAN_COLUMNS, WORKLIST_COLUMNS
from fleetvision.data.validate_pilot_human_review import (
    ERROR_COLUMNS,
    PilotHumanReviewValidationConfig,
    validate_pilot_human_review,
    write_errors_csv,
)


DEFAULT_CONFIG_PATH = Path("configs/data/pilot_review_excel_export_config.yaml")
DEFAULT_INPUT_XLSX = Path("outputs/manual_review/pilot500_human_review_interface.xlsx")
DEFAULT_SOURCE_WORKLIST_CSV = Path("dataset/00_catalog/image_review_labels_pilot500_human_review_worklist.csv")
DEFAULT_OUTPUT_CSV = Path("outputs/manual_review/pilot500_human_review_results.csv")
DEFAULT_SUMMARY_CSV = Path("outputs/metadata/pilot500_human_review_export_summary.csv")
DEFAULT_ERRORS_CSV = Path("outputs/metadata/pilot500_human_review_export_errors.csv")

IDENTITY_CHECK_COLUMNS = ["image_id", "original_path"]
SUMMARY_COLUMNS = [
    "source_rows",
    "workbook_rows",
    "exported_rows",
    "valid_rows",
    "invalid_rows",
    "pending_rows",
    "reviewed_rows",
    "needs_followup_rows",
    "reviewer_filled_rows",
    "reviewed_at_filled_rows",
    "validation_error_count",
]


@dataclass(frozen=True)
class PilotReviewExcelExportConfig:
    """Resolved configuration for exporting Pilot human review Excel results."""

    input_xlsx: Path
    source_worklist_csv: Path
    output_csv: Path
    summary_csv: Path
    errors_csv: Path
    expected_rows: int = 500


@dataclass(frozen=True)
class ExportResult:
    """Result from a Pilot review Excel export attempt."""

    is_valid: bool
    summary: dict[str, int]
    errors: list[dict[str, Any]]
    exported_dataframe: pd.DataFrame


def find_project_root(start: Path | None = None) -> Path:
    """Find the FleetVision project root from a starting path."""
    current = (start or Path.cwd()).resolve()
    markers = ["PROJECT_CONTEXT_BRIEF.md", "src/fleetvision", "configs/data"]
    for path in [current, *current.parents]:
        if all((path / marker).exists() for marker in markers):
            return path
    return current


def resolve_path(path: str | Path | None, project_root: Path, default: Path | None = None) -> Path:
    """Resolve a possibly relative path against the project root."""
    raw = Path(path) if path is not None else default
    if raw is None:
        raise ValueError("path or default must be provided")
    return raw if raw.is_absolute() else project_root / raw


def load_config(config_path: Path, project_root: Path) -> PilotReviewExcelExportConfig:
    """Load and resolve Pilot review Excel export YAML configuration."""
    if not config_path.exists():
        raise FileNotFoundError(f"Pilot review Excel export config not found: {config_path}")

    with config_path.open("r", encoding="utf-8") as file:
        raw_config: dict[str, Any] = yaml.safe_load(file) or {}

    return PilotReviewExcelExportConfig(
        input_xlsx=resolve_path(raw_config.get("input_xlsx"), project_root, DEFAULT_INPUT_XLSX),
        source_worklist_csv=resolve_path(raw_config.get("source_worklist_csv"), project_root, DEFAULT_SOURCE_WORKLIST_CSV),
        output_csv=resolve_path(raw_config.get("output_csv"), project_root, DEFAULT_OUTPUT_CSV),
        summary_csv=resolve_path(raw_config.get("summary_csv"), project_root, DEFAULT_SUMMARY_CSV),
        errors_csv=resolve_path(raw_config.get("errors_csv"), project_root, DEFAULT_ERRORS_CSV),
        expected_rows=int(raw_config.get("expected_rows", 500)),
    )


def read_worklist_csv(path: Path) -> pd.DataFrame:
    """Read the canonical worklist CSV without converting blanks to NaN."""
    if not path.exists():
        raise FileNotFoundError(f"source worklist CSV not found: {path}")
    return pd.read_csv(path, dtype=str, keep_default_na=False, encoding="utf-8-sig")


def read_review_workbook(path: Path) -> pd.DataFrame:
    """Read the review worksheet from an Excel workbook."""
    if not path.exists():
        raise FileNotFoundError(f"input workbook not found: {path}")
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError("openpyxl is required to export the Pilot review Excel workbook") from exc

    workbook = load_workbook(path, data_only=False)
    if WORKLIST_SHEET not in workbook.sheetnames:
        raise ValueError(f"workbook missing required sheet: {WORKLIST_SHEET}")
    worksheet = workbook[WORKLIST_SHEET]
    rows = list(worksheet.iter_rows(values_only=False))
    if not rows:
        raise ValueError(f"workbook sheet is empty: {WORKLIST_SHEET}")

    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in rows[0]]
    records: list[dict[str, Any]] = []
    for excel_row in rows[1:]:
        if all(cell.value is None for cell in excel_row):
            continue
        record: dict[str, Any] = {}
        for header, cell in zip(headers, excel_row):
            if not header:
                continue
            record[header] = _serialize_excel_cell(cell, header)
        records.append(record)
    return pd.DataFrame(records).fillna("")


def export_pilot_review_excel(config: PilotReviewExcelExportConfig) -> ExportResult:
    """Export human edits from Excel into a validated canonical CSV DataFrame."""
    source = read_worklist_csv(config.source_worklist_csv)
    workbook = read_review_workbook(config.input_xlsx)
    exported = merge_excel_human_fields(source, workbook, config.expected_rows)
    validation_config = PilotHumanReviewValidationConfig(
        input_csv=config.output_csv,
        summary_csv=config.summary_csv,
        errors_csv=config.errors_csv,
        expected_rows=config.expected_rows,
    )
    validation = validate_pilot_human_review(exported, validation_config)
    summary = _build_summary(source, workbook, exported, validation)
    return ExportResult(
        is_valid=validation.is_valid,
        summary=summary,
        errors=validation.errors,
        exported_dataframe=exported,
    )


def merge_excel_human_fields(
    source: pd.DataFrame,
    workbook: pd.DataFrame,
    expected_rows: int,
) -> pd.DataFrame:
    """Merge Excel human fields onto canonical source rows using review_id."""
    _validate_source_worklist(source, expected_rows)
    _validate_workbook(workbook, source, expected_rows)

    workbook_by_review_id = workbook.set_index("review_id", drop=False)
    exported = source.copy(deep=True)
    for column in HUMAN_COLUMNS:
        exported[column] = exported["review_id"].map(workbook_by_review_id[column]).fillna("").astype(str)
    exported = exported.loc[:, WORKLIST_COLUMNS].astype(str)
    if any(column.endswith(("_x", "_y")) for column in exported.columns):
        raise ValueError("export produced unexpected merge suffix columns")
    return exported


def write_export_outputs(result: ExportResult, config: PilotReviewExcelExportConfig) -> None:
    """Write summary, validation errors, and atomically replace output CSV when valid."""
    write_export_summary(result.summary, config.summary_csv)
    write_errors_csv(result.errors, config.errors_csv)
    if not result.is_valid:
        return
    _atomic_write_csv(result.exported_dataframe, config.output_csv)


def write_export_summary(summary: dict[str, int], summary_csv: Path) -> None:
    """Write one-row export summary CSV."""
    summary_csv.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame([{column: summary.get(column, 0) for column in SUMMARY_COLUMNS}]).to_csv(
        summary_csv, index=False, encoding="utf-8-sig"
    )


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    """Build the CLI parser."""
    parser = argparse.ArgumentParser(description="Export FleetVision Pilot 500 Excel review results to CSV.")
    parser.add_argument("--project-root", type=Path, default=None, help="FleetVision project root. Default: auto-detect.")
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG_PATH, help="Path to export config YAML.")
    parser.add_argument("--input-xlsx", type=Path, default=None, help="Override input Excel workbook path.")
    parser.add_argument("--source-worklist-csv", type=Path, default=None, help="Override canonical source worklist CSV path.")
    parser.add_argument("--output-csv", type=Path, default=None, help="Override exported review result CSV path.")
    parser.add_argument("--summary-csv", type=Path, default=None, help="Override export summary CSV path.")
    parser.add_argument("--errors-csv", type=Path, default=None, help="Override export errors CSV path.")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    """CLI entry point."""
    args = parse_args(argv)
    project_root = find_project_root(args.project_root)
    config_path = resolve_path(args.config, project_root, DEFAULT_CONFIG_PATH)
    try:
        config = _apply_overrides(load_config(config_path, project_root), args, project_root)
        result = export_pilot_review_excel(config)
        write_export_outputs(result, config)
    except Exception as exc:  # noqa: BLE001 - CLI should return concise errors.
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2

    print("FleetVision Pilot 500 review Excel export summary")
    print(f"output_csv: {config.output_csv}")
    print(f"summary_csv: {config.summary_csv}")
    print(f"errors_csv: {config.errors_csv}")
    for column in SUMMARY_COLUMNS:
        print(f"{column}: {result.summary[column]}")
    return 0 if result.is_valid else 1


def _apply_overrides(
    config: PilotReviewExcelExportConfig,
    args: argparse.Namespace,
    project_root: Path,
) -> PilotReviewExcelExportConfig:
    return PilotReviewExcelExportConfig(
        input_xlsx=resolve_path(args.input_xlsx, project_root, config.input_xlsx),
        source_worklist_csv=resolve_path(args.source_worklist_csv, project_root, config.source_worklist_csv),
        output_csv=resolve_path(args.output_csv, project_root, config.output_csv),
        summary_csv=resolve_path(args.summary_csv, project_root, config.summary_csv),
        errors_csv=resolve_path(args.errors_csv, project_root, config.errors_csv),
        expected_rows=config.expected_rows,
    )


def _validate_source_worklist(source: pd.DataFrame, expected_rows: int) -> None:
    missing = [column for column in WORKLIST_COLUMNS if column not in source.columns]
    if missing:
        raise ValueError("source worklist missing required column(s): " + ", ".join(missing))
    if len(source) != expected_rows:
        raise ValueError(f"source worklist expected {expected_rows} row(s), found {len(source)}")
    _validate_unique_values(source, "review_id", "source worklist")
    _validate_unique_values(source, "image_id", "source worklist")


def _validate_workbook(workbook: pd.DataFrame, source: pd.DataFrame, expected_rows: int) -> None:
    required = ["review_id", *IDENTITY_CHECK_COLUMNS, *HUMAN_COLUMNS]
    missing = [column for column in required if column not in workbook.columns]
    if missing:
        raise ValueError("workbook missing required column(s): " + ", ".join(missing))
    if len(workbook) != expected_rows:
        raise ValueError(f"workbook expected {expected_rows} row(s), found {len(workbook)}")
    _validate_unique_values(workbook, "review_id", "workbook")

    source_ids = set(source["review_id"].astype(str))
    workbook_ids = set(workbook["review_id"].astype(str))
    missing_ids = sorted(source_ids - workbook_ids)
    unknown_ids = sorted(workbook_ids - source_ids)
    if missing_ids:
        raise ValueError("workbook missing review_id value(s): " + ", ".join(missing_ids[:10]))
    if unknown_ids:
        raise ValueError("workbook contains unknown review_id value(s): " + ", ".join(unknown_ids[:10]))

    source_by_id = source.set_index("review_id", drop=False)
    workbook_by_id = workbook.set_index("review_id", drop=False)
    for review_id in source["review_id"].astype(str).tolist():
        for column in IDENTITY_CHECK_COLUMNS:
            source_value = str(source_by_id.at[review_id, column])
            workbook_value = str(workbook_by_id.at[review_id, column])
            if source_value != workbook_value:
                raise ValueError(f"workbook {column} mismatch for review_id {review_id}")
        for column in HUMAN_COLUMNS:
            value = workbook_by_id.at[review_id, column]
            if isinstance(value, str) and value.startswith("="):
                raise ValueError(f"workbook human field contains formula for review_id {review_id}: {column}")


def _validate_unique_values(dataframe: pd.DataFrame, column: str, source_name: str) -> None:
    values = dataframe[column].astype(str).str.strip()
    if (values == "").any():
        raise ValueError(f"{source_name} contains blank {column}")
    duplicate_count = int(values.duplicated(keep=False).sum())
    if duplicate_count:
        duplicate_values = sorted(values[values.duplicated(keep=False)].unique().tolist())
        raise ValueError(
            f"{source_name} contains duplicate {column}: {duplicate_count} duplicate row(s); "
            f"examples: {', '.join(duplicate_values[:5])}"
        )


def _serialize_excel_cell(cell, header: str) -> str:
    if header in HUMAN_COLUMNS and cell.data_type == "f":
        return "=" + str(cell.value).lstrip("=")
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return datetime.combine(value, time.min).isoformat(sep=" ", timespec="seconds")
    return str(value)


def _build_summary(source, workbook, exported, validation) -> dict[str, int]:
    return {
        "source_rows": int(len(source)),
        "workbook_rows": int(len(workbook)),
        "exported_rows": int(len(exported)) if validation.is_valid else 0,
        "valid_rows": int(validation.valid_rows),
        "invalid_rows": int(validation.invalid_rows),
        "pending_rows": int(validation.pending_rows),
        "reviewed_rows": int(validation.reviewed_rows),
        "needs_followup_rows": int(validation.needs_followup_rows),
        "reviewer_filled_rows": int((exported["human_reviewer"].astype(str).str.strip() != "").sum()),
        "reviewed_at_filled_rows": int(validation.reviewed_at_filled_rows),
        "validation_error_count": int(validation.error_count),
    }


def _atomic_write_csv(dataframe: pd.DataFrame, output_csv: Path) -> None:
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    temp_path: Path | None = None
    with tempfile.NamedTemporaryFile(
        mode="w",
        newline="",
        encoding="utf-8-sig",
        suffix=".csv",
        delete=False,
        dir=output_csv.parent,
    ) as temp_file:
        temp_path = Path(temp_file.name)
        dataframe.to_csv(temp_file, index=False)
    os.replace(temp_path, output_csv)


if __name__ == "__main__":
    raise SystemExit(main())
