"""Build a Pilot 500 Excel human review interface workbook."""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
import yaml

from fleetvision.data.build_pilot_review_worklist import WORKLIST_COLUMNS
from fleetvision.data.validate_pilot_human_review import ALLOWED_VALUES


DEFAULT_CONFIG_PATH = Path("configs/data/pilot_review_excel_config.yaml")
DEFAULT_INPUT_CSV = Path("dataset/00_catalog/image_review_labels_pilot500_human_review_worklist.csv")
DEFAULT_OUTPUT_XLSX = Path("outputs/manual_review/pilot500_human_review_interface.xlsx")

INSTRUCTIONS_SHEET = "使用說明"
WORKLIST_SHEET = "覆核工作表"
SUMMARY_SHEET = "進度摘要"
OPTIONS_SHEET = "選項清單"
OPEN_IMAGE_COLUMN = "open_image"
REVIEW_STATUS_COLUMN = "human_review_status"
REVIEWER_COLUMN = "human_reviewer"
REVIEWED_AT_COLUMN = "human_reviewed_at"
PHOTO_TYPE_COLUMN = "human_photo_type_review"
ANGLE_COLUMN = "human_angle_review"
IS_EXTERIOR_COLUMN = "human_is_exterior_review"
DAMAGE_COLUMN = "human_has_visible_damage_review"
SEVERITY_COLUMN = "human_severity_review"

EXCEL_COLUMNS = [OPEN_IMAGE_COLUMN, *WORKLIST_COLUMNS]
DROPDOWN_COLUMNS = [
    PHOTO_TYPE_COLUMN,
    ANGLE_COLUMN,
    IS_EXTERIOR_COLUMN,
    DAMAGE_COLUMN,
    SEVERITY_COLUMN,
    REVIEW_STATUS_COLUMN,
]

GROUP_FILLS = {
    "open_image": "D9EAD3",
    "identity": "D9D9D9",
    "suggestion": "CFE2F3",
    "seed": "FFF2CC",
    "human": "D9EAD3",
}

STATUS_FILLS = {
    "pending": "FFF2CC",
    "reviewed": "D9EAD3",
    "needs_followup": "FCE5CD",
    "skipped": "D9D9D9",
}


@dataclass(frozen=True)
class PilotReviewExcelConfig:
    """Resolved configuration for building the Pilot 500 review Excel workbook."""

    input_csv: Path
    output_xlsx: Path
    project_root: Path
    expected_rows: int = 500


def find_project_root(start: Path | None = None) -> Path:
    """Find the FleetVision project root from a starting path."""
    current = (start or Path.cwd()).resolve()
    markers = ["PROJECT_CONTEXT_BRIEF.md", "src/fleetvision", "configs/data"]
    for path in [current, *current.parents]:
        if all((path / marker).exists() for marker in markers):
            return path
    return current


def resolve_path(path: str | Path | None, project_root: Path, default: Path | None = None) -> Path:
    """Resolve a possibly relative path against the project root."""
    raw = Path(path) if path is not None else default
    if raw is None:
        raise ValueError("path or default must be provided")
    return raw if raw.is_absolute() else project_root / raw


def load_config(config_path: Path, project_root: Path) -> PilotReviewExcelConfig:
    """Load and resolve Pilot review Excel YAML configuration."""
    if not config_path.exists():
        raise FileNotFoundError(f"Pilot review Excel config not found: {config_path}")

    with config_path.open("r", encoding="utf-8") as file:
        raw_config: dict[str, Any] = yaml.safe_load(file) or {}

    return PilotReviewExcelConfig(
        input_csv=resolve_path(raw_config.get("input_csv"), project_root, DEFAULT_INPUT_CSV),
        output_xlsx=resolve_path(raw_config.get("output_xlsx"), project_root, DEFAULT_OUTPUT_XLSX),
        project_root=project_root,
        expected_rows=int(raw_config.get("expected_rows", 500)),
    )


def read_csv(path: Path) -> pd.DataFrame:
    """Read a CSV with UTF-8 BOM support while preserving blank strings."""
    if not path.exists():
        raise FileNotFoundError(f"CSV not found: {path}")
    return pd.read_csv(path, dtype=str, keep_default_na=False, encoding="utf-8-sig")


def build_pilot_review_excel(dataframe: pd.DataFrame, config: PilotReviewExcelConfig):
    """Build an openpyxl workbook from a Pilot 500 human review worklist."""
    _validate_input_dataframe(dataframe, config)
    openpyxl = _require_openpyxl()
    Workbook = openpyxl.Workbook

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    ws_instructions = workbook.create_sheet(INSTRUCTIONS_SHEET)
    ws_worklist = workbook.create_sheet(WORKLIST_SHEET)
    ws_summary = workbook.create_sheet(SUMMARY_SHEET)
    ws_options = workbook.create_sheet(OPTIONS_SHEET)
    ws_options.sheet_state = "hidden"

    _write_instructions(ws_instructions)
    _write_options(ws_options)
    _write_worklist(ws_worklist, dataframe, config.project_root)
    _write_summary(ws_summary, len(dataframe))
    _style_workbook(workbook)
    return workbook


def write_workbook(workbook, output_path: Path) -> None:
    """Write the Excel workbook to disk."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    """Build the CLI parser."""
    parser = argparse.ArgumentParser(description="Build the FleetVision Pilot 500 human review Excel interface.")
    parser.add_argument("--project-root", type=Path, default=None, help="FleetVision project root. Default: auto-detect.")
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG_PATH, help="Path to Excel config YAML.")
    parser.add_argument("--input", type=Path, default=None, help="Override input worklist CSV.")
    parser.add_argument("--output", type=Path, default=None, help="Override output Excel workbook path.")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    """CLI entry point."""
    args = parse_args(argv)
    project_root = find_project_root(args.project_root)
    config_path = resolve_path(args.config, project_root, DEFAULT_CONFIG_PATH)

    try:
        config = load_config(config_path, project_root)
        config = _apply_overrides(config, args, project_root)
        dataframe = read_csv(config.input_csv)
        workbook = build_pilot_review_excel(dataframe, config)
        write_workbook(workbook, config.output_xlsx)
    except Exception as exc:  # noqa: BLE001 - CLI should return concise errors.
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    print("FleetVision Pilot 500 human review Excel interface")
    print(f"input_csv: {config.input_csv}")
    print(f"output_xlsx: {config.output_xlsx}")
    print(f"rows: {len(dataframe)}")
    return 0


def _apply_overrides(
    config: PilotReviewExcelConfig,
    args: argparse.Namespace,
    project_root: Path,
) -> PilotReviewExcelConfig:
    return PilotReviewExcelConfig(
        input_csv=resolve_path(args.input, project_root, config.input_csv),
        output_xlsx=resolve_path(args.output, project_root, config.output_xlsx),
        project_root=project_root,
        expected_rows=config.expected_rows,
    )


def _validate_input_dataframe(dataframe: pd.DataFrame, config: PilotReviewExcelConfig) -> None:
    missing = [column for column in WORKLIST_COLUMNS if column not in dataframe.columns]
    if missing:
        raise ValueError("input worklist CSV missing required column(s): " + ", ".join(missing))
    if len(dataframe) != config.expected_rows:
        raise ValueError(f"input worklist CSV expected {config.expected_rows} row(s), found {len(dataframe)}")
    duplicate_count = int(dataframe["image_id"].duplicated(keep=False).sum())
    if duplicate_count:
        duplicate_values = sorted(dataframe.loc[dataframe["image_id"].duplicated(keep=False), "image_id"].astype(str).unique())
        raise ValueError(
            f"duplicate image_id in input worklist CSV: {duplicate_count} duplicate row(s); "
            f"examples: {', '.join(duplicate_values[:5])}"
        )


def _write_instructions(worksheet) -> None:
    instructions = [
        "FleetVision Pilot 500 人工覆核介面",
        "黃／藍區為參考值。",
        "綠色欄位才是人工編輯區。",
        "預填值不代表已覆核。",
        "完成時需設定 human_review_status=reviewed。",
        "reviewed 必須填 reviewer 與 reviewed_at。",
        "needs_followup／skipped 必須填 notes。",
        "不得依 source_bucket 直接判定 damage 或 severity。",
        "編輯完成後需執行 Validator。",
    ]
    for row_index, text in enumerate(instructions, start=1):
        worksheet.cell(row=row_index, column=1, value=text)
    worksheet.column_dimensions["A"].width = 90


def _write_options(worksheet) -> None:
    for column_index, column_name in enumerate(DROPDOWN_COLUMNS, start=1):
        values = sorted(ALLOWED_VALUES[column_name])
        worksheet.cell(row=1, column=column_index, value=column_name)
        for row_index, value in enumerate(values, start=2):
            worksheet.cell(row=row_index, column=column_index, value=value)


def _write_worklist(worksheet, dataframe: pd.DataFrame, project_root: Path) -> None:
    for column_index, column_name in enumerate(EXCEL_COLUMNS, start=1):
        worksheet.cell(row=1, column=column_index, value=column_name)

    for row_index, (_, row) in enumerate(dataframe.iterrows(), start=2):
        original_path = str(row["original_path"])
        open_cell = worksheet.cell(row=row_index, column=1, value="開啟圖片")
        open_cell.hyperlink = _resolve_image_uri(original_path, project_root)
        open_cell.style = "Hyperlink"
        for column_index, column_name in enumerate(WORKLIST_COLUMNS, start=2):
            cell = worksheet.cell(row=row_index, column=column_index)
            _set_literal_cell_value(cell, str(row[column_name]))

    worksheet.freeze_panes = "G2"
    worksheet.auto_filter.ref = worksheet.dimensions


def _write_summary(worksheet, total_rows: int) -> None:
    last_row = total_rows + 1
    status_column = _column_letter(EXCEL_COLUMNS.index(REVIEW_STATUS_COLUMN) + 1)
    reviewer_column = _column_letter(EXCEL_COLUMNS.index(REVIEWER_COLUMN) + 1)
    reviewed_at_column = _column_letter(EXCEL_COLUMNS.index(REVIEWED_AT_COLUMN) + 1)
    worksheet.append(["metric", "value"])
    rows = [
        ("總筆數", f"=COUNTA('{WORKLIST_SHEET}'!B2:B{last_row})"),
        ("pending", f'=COUNTIF(\'{WORKLIST_SHEET}\'!{status_column}2:{status_column}{last_row},"pending")'),
        ("reviewed", f'=COUNTIF(\'{WORKLIST_SHEET}\'!{status_column}2:{status_column}{last_row},"reviewed")'),
        ("needs_followup", f'=COUNTIF(\'{WORKLIST_SHEET}\'!{status_column}2:{status_column}{last_row},"needs_followup")'),
        ("skipped", f'=COUNTIF(\'{WORKLIST_SHEET}\'!{status_column}2:{status_column}{last_row},"skipped")'),
        ("完成率", "=IF(B2=0,0,B4/B2)"),
        ("reviewed 但缺 reviewer", f'=COUNTIFS(\'{WORKLIST_SHEET}\'!{status_column}2:{status_column}{last_row},"reviewed",\'{WORKLIST_SHEET}\'!{reviewer_column}2:{reviewer_column}{last_row},"")'),
        ("reviewed 但缺 reviewed_at", f'=COUNTIFS(\'{WORKLIST_SHEET}\'!{status_column}2:{status_column}{last_row},"reviewed",\'{WORKLIST_SHEET}\'!{reviewed_at_column}2:{reviewed_at_column}{last_row},"")'),
    ]
    for row in rows:
        worksheet.append(row)
    worksheet.column_dimensions["A"].width = 28
    worksheet.column_dimensions["B"].width = 20


def _style_workbook(workbook) -> None:
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    ws = workbook[WORKLIST_SHEET]
    options = workbook[OPTIONS_SHEET]
    max_row = ws.max_row
    max_col = ws.max_column
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill("solid", fgColor=_fill_for_column(cell.value))
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor=_fill_for_column(ws.cell(row=1, column=cell.column).value))

    for column_index, column_name in enumerate(EXCEL_COLUMNS, start=1):
        ws.column_dimensions[_column_letter(column_index)].width = _width_for_column(column_name)

    reviewed_at_col = _column_letter(EXCEL_COLUMNS.index(REVIEWED_AT_COLUMN) + 1)
    for cell in ws[f"{reviewed_at_col}2:{reviewed_at_col}{max_row}"]:
        for item in cell:
            item.number_format = "yyyy-mm-dd hh:mm:ss"

    _add_dropdowns(ws, options, max_row)
    _add_conditional_formatting(ws, max_row)


def _add_dropdowns(ws, options, max_row: int) -> None:
    from openpyxl.worksheet.datavalidation import DataValidation

    for option_col_index, column_name in enumerate(DROPDOWN_COLUMNS, start=1):
        values_count = len(ALLOWED_VALUES[column_name])
        formula = f"'{OPTIONS_SHEET}'!${_column_letter(option_col_index)}$2:${_column_letter(option_col_index)}${values_count + 1}"
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        validation.error = "請選擇清單中的值。"
        validation.errorTitle = "無效選項"
        ws.add_data_validation(validation)
        target_col = _column_letter(EXCEL_COLUMNS.index(column_name) + 1)
        validation.add(f"{target_col}2:{target_col}{max_row}")


def _add_conditional_formatting(ws, max_row: int) -> None:
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import PatternFill

    status_col = _column_letter(EXCEL_COLUMNS.index(REVIEW_STATUS_COLUMN) + 1)
    reviewer_col = _column_letter(EXCEL_COLUMNS.index(REVIEWER_COLUMN) + 1)
    reviewed_at_col = _column_letter(EXCEL_COLUMNS.index(REVIEWED_AT_COLUMN) + 1)
    photo_type_col = _column_letter(EXCEL_COLUMNS.index(PHOTO_TYPE_COLUMN) + 1)
    is_exterior_col = _column_letter(EXCEL_COLUMNS.index(IS_EXTERIOR_COLUMN) + 1)
    damage_col = _column_letter(EXCEL_COLUMNS.index(DAMAGE_COLUMN) + 1)
    severity_col = _column_letter(EXCEL_COLUMNS.index(SEVERITY_COLUMN) + 1)
    review_range = f"A2:{_column_letter(ws.max_column)}{max_row}"

    for status, color in STATUS_FILLS.items():
        ws.conditional_formatting.add(
            review_range,
            FormulaRule(formula=[f'${status_col}2="{status}"'], fill=PatternFill("solid", fgColor=color)),
        )

    red_fill = PatternFill("solid", fgColor="F4CCCC")
    highlight_fill = PatternFill("solid", fgColor="D9EAD3")
    formulas = [
        f'AND(${status_col}2="reviewed",${reviewer_col}2="")',
        f'AND(${status_col}2="reviewed",${reviewed_at_col}2="")',
        f'AND(${photo_type_col}2="exterior",${is_exterior_col}2<>"1")',
        f'AND(OR(${photo_type_col}2="interior",${photo_type_col}2="low_quality",${photo_type_col}2="irrelevant"),${is_exterior_col}2<>"0")',
        f'AND(${damage_col}2="0",${severity_col}2<>"none")',
        f'AND(${damage_col}2="1",${severity_col}2="none")',
        f'AND(${damage_col}2="unknown",${severity_col}2<>"unknown")',
    ]
    for formula in formulas:
        ws.conditional_formatting.add(review_range, FormulaRule(formula=[formula], fill=red_fill))

    ws.conditional_formatting.add(
        review_range,
        FormulaRule(
            formula=[f'OR(${photo_type_col}2="unknown",${damage_col}2="unknown",${status_col}2="needs_followup")'],
            fill=highlight_fill,
        ),
    )


def _fill_for_column(column_name: str) -> str:
    if column_name == OPEN_IMAGE_COLUMN:
        return GROUP_FILLS["open_image"]
    if column_name in {"review_id", "image_id", "source_bucket", "original_path", "filename"}:
        return GROUP_FILLS["identity"]
    if column_name.startswith("suggested_") or column_name in {"photo_type_confidence", "angle_confidence", "auto_review_notes"}:
        return GROUP_FILLS["suggestion"]
    if column_name.startswith("seed_"):
        return GROUP_FILLS["seed"]
    if column_name.startswith("human_"):
        return GROUP_FILLS["human"]
    return "FFFFFF"


def _width_for_column(column_name: str) -> float:
    if column_name in {"original_path", OPEN_IMAGE_COLUMN}:
        return 36
    if column_name.endswith("notes"):
        return 34
    if column_name in {"review_id", "image_id", "source_bucket", "filename"}:
        return 22
    if column_name.endswith("confidence"):
        return 14
    return 18


def _set_literal_cell_value(cell, value: str) -> None:
    cell.value = value
    if value.startswith(("=", "+", "-", "@")):
        cell.data_type = "s"


def _resolve_image_uri(original_path: str, project_root: Path) -> str:
    path = Path(original_path)
    resolved_path = path if path.is_absolute() else project_root / path
    return resolved_path.resolve().as_uri()


def _column_letter(column_index: int) -> str:
    from openpyxl.utils import get_column_letter

    return get_column_letter(column_index)


def _require_openpyxl():
    try:
        import openpyxl
    except ModuleNotFoundError as exc:
        raise RuntimeError("openpyxl is required to build the Pilot review Excel workbook, but it is not installed") from exc
    return openpyxl


if __name__ == "__main__":
    raise SystemExit(main())

