"""Read-only Phase 04.5L completed-review findings analysis.

This module implements two explicit operational sub-gates:

* F1 validates the frozen completed review and creates a severity-scope review
  package in a timestamped, no-overwrite workspace outside the repository.
* F2 validates the completed scope review and creates deterministic findings,
  evidence, and an advisory recommendation.

It never reads the test split, runs model inference, mutates annotations or
fixed datasets, starts training, or grants retraining/deployment approval.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import os
import shutil
import subprocess
import tempfile
import uuid
import zipfile
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path, PurePosixPath
from typing import Any, Callable, Iterable, Mapping, Sequence

import pandas as pd
import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from fleetvision.data.validation_error_human_review import (
    CANONICAL_COLUMNS,
    HumanReviewError,
    ReviewConfig,
    export_review_workbook,
    load_config as load_existing_review_config,
    logical_fingerprint,
    read_workbook_dataframe,
    sha256_file,
    summarize_canonical_review,
    validate_canonical_csv,
    write_validation_outputs,
)

DEFAULT_CONFIG_PATH = Path(
    "configs/data/phase04_5l_completed_review_findings_config.yaml"
)

SCOPE_COLUMNS = (
    "scope_review_status",
    "scope_group",
    "scope_reason",
    "operability",
    "scope_confidence",
    "scope_reviewer_notes",
    "scope_reviewer",
    "scope_reviewed_at_utc",
)
SCOPE_EXPORT_COLUMNS = CANONICAL_COLUMNS + SCOPE_COLUMNS
SCOPE_WORKBOOK_SHEETS = (
    "Instructions",
    "Scope_Review",
    "Option_Lists",
    "Manifest",
    "Progress_Summary",
)
PRIMARY_RECOMMENDATIONS = (
    "NO_RETRAINING_RECOMMENDED",
    "DATA_CORRECTION_REQUIRED_BEFORE_RETRAINING",
    "SCOPE_REBALANCING_REQUIRED_BEFORE_RETRAINING",
    "RETRAINING_PROPOSAL_JUSTIFIED",
    "ADDITIONAL_REVIEW_REQUIRED",
)


class FindingsAnalysisError(RuntimeError):
    """Raised when a findings-analysis safety or data contract fails."""


@dataclass(frozen=True)
class FindingsConfig:
    project_root: Path
    existing_review_config_path: Path
    expected_validator_classification: str
    expected_branch: str
    protected_untracked_path: str
    completed_workbook: Path
    completed_workbook_size: int
    completed_workbook_sha256: str
    completed_logical_fingerprint: str
    source_workbook: Path
    source_workbook_sha256: str
    frozen_package: Path
    frozen_package_sha256: str
    expected_case_count: int
    expected_reviewer: str
    expected_export_classification: str
    workspace_parent: Path
    workspace_prefix: str
    canonical_filename: str
    scope_workbook_filename: str
    scope_completed_workbook_filename: str
    scope_source_filename: str
    scope_asset_manifest_filename: str
    scope_export_filename: str
    scope_options: Mapping[str, tuple[str, ...]]
    recommendation_rules: Mapping[str, float]
    retraining_status: str
    deployment_acceptance: str


@dataclass(frozen=True)
class RepositoryState:
    branch: str
    local_head: str
    origin_main: str
    remote_head: str
    unexpected_status: tuple[str, ...]


@dataclass(frozen=True)
class WorkspacePaths:
    root: Path
    completed_snapshot: Path
    source_workbook_snapshot: Path
    package_snapshot: Path
    extracted_package: Path
    canonical_csv: Path
    validation_report: Path
    validation_errors: Path
    scope_workbook: Path
    scope_completed_workbook: Path
    scope_source_csv: Path
    scope_asset_manifest: Path
    scope_export_csv: Path


@dataclass(frozen=True)
class ScopeValidationResult:
    passed: bool
    row_count: int
    issue_count: int
    issues: tuple[dict[str, str], ...]
    counts: Mapping[str, int]


@dataclass(frozen=True)
class RecommendationResult:
    primary: str
    reasons: tuple[str, ...]
    metrics: Mapping[str, float]


GitRunner = Callable[[Sequence[str], Path], str]


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _normalize(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.tzinfo is None:
            value = value.replace(tzinfo=timezone.utc)
        return value.astimezone(timezone.utc).isoformat()
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def _resolve_path(project_root: Path, value: str) -> Path:
    path = Path(value)
    return path if path.is_absolute() else project_root / path


def _require_hash(name: str, value: Any) -> str:
    digest = _normalize(value).upper()
    if len(digest) != 64 or any(c not in "0123456789ABCDEF" for c in digest):
        raise FindingsAnalysisError(f"{name} must be 64 uppercase hexadecimal characters")
    return digest


def _require_exact_options(
    raw: Mapping[str, Any],
    key: str,
    expected: tuple[str, ...],
) -> tuple[str, ...]:
    values_raw = raw.get(key)
    if not isinstance(values_raw, list):
        raise FindingsAnalysisError(f"scope_options.{key} must be a list")
    values = tuple(_normalize(value) for value in values_raw)
    if values != expected or any(not value for value in values) or len(set(values)) != len(values):
        raise FindingsAnalysisError(
            f"scope_options.{key} must exactly equal the approved controlled values"
        )
    return values


def load_findings_config(config_path: Path, project_root: Path) -> FindingsConfig:
    """Load and strictly validate the immutable findings-analysis contract."""

    project_root = project_root.resolve()
    path = _resolve_path(project_root, str(config_path))
    if not path.is_file():
        raise FileNotFoundError(f"findings config not found: {path}")
    raw = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    completed = raw.get("completed_review", {})
    repository = raw.get("repository", {})
    workspace = raw.get("workspace", {})
    scope_raw = raw.get("scope_options", {})
    rules_raw = raw.get("recommendation_rules", {})
    safety = raw.get("safety", {})

    expected_scope_options = {
        "scope_review_status": ("pending", "reviewed", "needs_adjudication"),
        "scope_group": (
            "IN_SCOPE_LIGHT_MODERATE",
            "BOUNDARY_HEAVY_DAMAGE",
            "OUT_OF_SCOPE_CATASTROPHIC",
        ),
        "scope_reason": (
            "light_surface_damage",
            "moderate_external_damage",
            "heavy_external_damage",
            "structural_damage",
            "catastrophic_collision",
            "extensive_multi_panel_damage",
            "vehicle_integrity_compromised",
            "insufficient_visual_evidence",
            "other",
        ),
        "operability": (
            "drivable_or_likely_drivable",
            "uncertain",
            "non_drivable_or_likely_non_drivable",
        ),
        "scope_confidence": ("high", "medium", "low"),
    }
    scope_options = {
        key: _require_exact_options(scope_raw, key, expected)
        for key, expected in expected_scope_options.items()
    }

    required_rule_keys = (
        "additional_review_low_confidence_share",
        "scope_rebalancing_non_scope_share",
        "distribution_total_variation_threshold",
        "threshold_tradeoff_share_delta",
        "retraining_in_scope_priority_share",
        "retraining_in_scope_confirmed_error_share",
    )
    recommendation_rules: dict[str, float] = {}
    for key in required_rule_keys:
        try:
            value = float(rules_raw[key])
        except (KeyError, TypeError, ValueError) as exc:
            raise FindingsAnalysisError(f"recommendation_rules.{key} is required") from exc
        if not 0.0 <= value <= 1.0:
            raise FindingsAnalysisError(f"recommendation_rules.{key} must be in [0, 1]")
        recommendation_rules[key] = value

    for key in (
        "test_split_read",
        "model_inference_executed",
        "annotation_modified",
        "training_started",
    ):
        if safety.get(key) is not False:
            raise FindingsAnalysisError(f"safety.{key} must remain false")

    retraining_status = _normalize(safety.get("retraining_status"))
    deployment_acceptance = _normalize(safety.get("deployment_acceptance"))
    if retraining_status != "NOT_YET_APPROVED":
        raise FindingsAnalysisError("retraining_status must remain NOT_YET_APPROVED")
    if deployment_acceptance != "NOT_YET_APPROVED":
        raise FindingsAnalysisError("deployment_acceptance must remain NOT_YET_APPROVED")

    expected_validator = _normalize(raw.get("expected_validator_classification"))
    if expected_validator != "VALIDATION_ERROR_HUMAN_REVIEW_VERIFIED":
        raise FindingsAnalysisError(
            "expected_validator_classification must be VALIDATION_ERROR_HUMAN_REVIEW_VERIFIED"
        )
    expected_case_count = int(completed.get("review_cases", 0))
    if expected_case_count != 130:
        raise FindingsAnalysisError("completed_review.review_cases must equal 130")
    if int(completed.get("reviewed", -1)) != 130:
        raise FindingsAnalysisError("completed_review.reviewed must equal 130")
    if int(completed.get("pending", -1)) != 0:
        raise FindingsAnalysisError("completed_review.pending must equal 0")
    if int(completed.get("needs_adjudication", -1)) != 0:
        raise FindingsAnalysisError("completed_review.needs_adjudication must equal 0")

    config = FindingsConfig(
        project_root=project_root,
        existing_review_config_path=_resolve_path(
            project_root, _normalize(raw.get("existing_review_config"))
        ),
        expected_validator_classification=expected_validator,
        expected_branch=_normalize(repository.get("branch")),
        protected_untracked_path=_normalize(
            repository.get("protected_untracked_path")
        ).replace("\\", "/").rstrip("/")
        + "/",
        completed_workbook=_resolve_path(
            project_root, _normalize(completed.get("workbook_path"))
        ),
        completed_workbook_size=int(completed.get("workbook_size_bytes", 0)),
        completed_workbook_sha256=_require_hash(
            "completed_review.workbook_sha256", completed.get("workbook_sha256")
        ),
        completed_logical_fingerprint=_require_hash(
            "completed_review.logical_fingerprint", completed.get("logical_fingerprint")
        ),
        source_workbook=_resolve_path(
            project_root, _normalize(completed.get("source_workbook_path"))
        ),
        source_workbook_sha256=_require_hash(
            "completed_review.source_workbook_sha256",
            completed.get("source_workbook_sha256"),
        ),
        frozen_package=_resolve_path(
            project_root, _normalize(completed.get("frozen_package_path"))
        ),
        frozen_package_sha256=_require_hash(
            "completed_review.frozen_package_sha256",
            completed.get("frozen_package_sha256"),
        ),
        expected_case_count=expected_case_count,
        expected_reviewer=_normalize(completed.get("reviewer")),
        expected_export_classification=_normalize(
            completed.get("export_classification")
        ),
        workspace_parent=_resolve_path(
            project_root, _normalize(workspace.get("parent_dir"))
        ),
        workspace_prefix=_normalize(workspace.get("prefix")),
        canonical_filename=_normalize(workspace.get("canonical_filename")),
        scope_workbook_filename=_normalize(workspace.get("scope_workbook_filename")),
        scope_completed_workbook_filename=_normalize(
            workspace.get("scope_completed_workbook_filename")
        ),
        scope_source_filename=_normalize(workspace.get("scope_source_filename")),
        scope_asset_manifest_filename=_normalize(
            workspace.get("scope_asset_manifest_filename")
        ),
        scope_export_filename=_normalize(workspace.get("scope_export_filename")),
        scope_options=scope_options,
        recommendation_rules=recommendation_rules,
        retraining_status=retraining_status,
        deployment_acceptance=deployment_acceptance,
    )
    if config.expected_branch != "main":
        raise FindingsAnalysisError("repository.branch must remain main")
    if not config.protected_untracked_path:
        raise FindingsAnalysisError("repository.protected_untracked_path is required")
    if config.completed_workbook_size <= 0:
        raise FindingsAnalysisError("completed Workbook size must be positive")
    for name in (
        config.workspace_prefix,
        config.canonical_filename,
        config.scope_workbook_filename,
        config.scope_completed_workbook_filename,
        config.scope_source_filename,
        config.scope_asset_manifest_filename,
        config.scope_export_filename,
    ):
        if not name:
            raise FindingsAnalysisError("workspace names must be nonblank")
    return config


def _run_git(args: Sequence[str], cwd: Path) -> str:
    completed = subprocess.run(
        ["git", *args],
        cwd=cwd,
        check=False,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    if completed.returncode != 0:
        raise FindingsAnalysisError(
            f"git {' '.join(args)} failed: {completed.stderr.strip()}"
        )
    return completed.stdout


def inspect_repository_state(
    project_root: Path,
    expected_head: str,
    *,
    runner: GitRunner | None = None,
) -> RepositoryState:
    """Fail closed unless branch, refs, and worktree match the Gate contract."""

    resolved = project_root.resolve()
    command = runner or _run_git
    branch = command(("branch", "--show-current"), resolved).strip()
    local_head = command(("rev-parse", "HEAD"), resolved).strip()
    origin_main = command(("rev-parse", "origin/main"), resolved).strip()
    remote_line = command(("ls-remote", "origin", "refs/heads/main"), resolved).strip()
    remote_head = remote_line.split()[0] if remote_line else ""
    status_lines = tuple(
        line
        for line in command(
            ("status", "--porcelain=v1", "--untracked-files=all"), resolved
        ).splitlines()
        if line.strip()
    )
    allowed_prefix = "outputs/metadata/external_assets/"
    unexpected = tuple(
        line
        for line in status_lines
        if not (
            line.startswith("?? ")
            and line[3:].replace("\\", "/").startswith(allowed_prefix)
        )
    )
    if branch != "main":
        raise FindingsAnalysisError(f"expected branch main; got {branch!r}")
    if len(expected_head) != 40 or any(c not in "0123456789abcdefABCDEF" for c in expected_head):
        raise FindingsAnalysisError("expected_head must be a 40-character commit SHA")
    if {local_head, origin_main, remote_head} != {expected_head}:
        raise FindingsAnalysisError(
            "repository HEAD mismatch: "
            f"expected={expected_head} local={local_head} "
            f"origin/main={origin_main} remote={remote_head}"
        )
    if unexpected:
        raise FindingsAnalysisError(
            f"unexpected worktree entries block the Gate: {unexpected}"
        )
    return RepositoryState(
        branch=branch,
        local_head=local_head,
        origin_main=origin_main,
        remote_head=remote_head,
        unexpected_status=unexpected,
    )


def verify_authoritative_inputs(
    config: FindingsConfig,
) -> dict[str, dict[str, str]]:
    """Verify immutable input existence, size, and SHA256 before any workspace write."""

    specs = {
        "completed_workbook": (
            config.completed_workbook,
            config.completed_workbook_sha256,
            config.completed_workbook_size,
        ),
        "source_workbook": (
            config.source_workbook,
            config.source_workbook_sha256,
            None,
        ),
        "frozen_package": (
            config.frozen_package,
            config.frozen_package_sha256,
            None,
        ),
    }
    rows: dict[str, dict[str, str]] = {}
    for name, (path, expected_hash, expected_size) in specs.items():
        if not path.is_file():
            raise FindingsAnalysisError(f"authoritative input missing ({name}): {path}")
        actual_size = path.stat().st_size
        if expected_size is not None and actual_size != expected_size:
            raise FindingsAnalysisError(
                f"authoritative input size mismatch ({name}): "
                f"expected={expected_size} actual={actual_size}"
            )
        actual_hash = sha256_file(path)
        if actual_hash != expected_hash:
            raise FindingsAnalysisError(
                f"authoritative input SHA256 mismatch ({name}): "
                f"expected={expected_hash} actual={actual_hash}"
            )
        rows[name] = {
            "path": str(path),
            "expected_size_bytes": "" if expected_size is None else str(expected_size),
            "actual_size_bytes": str(actual_size),
            "expected_sha256": expected_hash,
            "actual_sha256": actual_hash,
            "match": "true",
        }
    return rows


def _timestamp_token(timestamp: str | None = None) -> str:
    if timestamp:
        token = timestamp.strip()
    else:
        token = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    if not token or any(c not in "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz_-" for c in token):
        raise FindingsAnalysisError("timestamp contains unsupported characters")
    return token


def create_workspace(
    config: FindingsConfig,
    *,
    timestamp: str | None = None,
) -> WorkspacePaths:
    """Create the approved timestamped directory structure without outputs."""

    token = _timestamp_token(timestamp)
    root = config.workspace_parent / f"{config.workspace_prefix}_{token}"
    if root.exists():
        raise FindingsAnalysisError(f"analysis workspace already exists: {root}")
    directories = (
        root / "input_snapshot/completed_workbook",
        root / "input_snapshot/source_workbook",
        root / "input_snapshot/formal_package",
        root / "input_snapshot/extracted_package",
        root / "canonical",
        root / "reports",
        root / "scope_review",
        root / "final_findings",
        root / "evidence",
    )
    try:
        for directory in directories:
            directory.mkdir(parents=True, exist_ok=False)
    except Exception:
        shutil.rmtree(root, ignore_errors=True)
        raise
    return WorkspacePaths(
        root=root,
        completed_snapshot=directories[0] / config.completed_workbook.name,
        source_workbook_snapshot=directories[1] / config.source_workbook.name,
        package_snapshot=directories[2] / config.frozen_package.name,
        extracted_package=directories[3],
        canonical_csv=root / "canonical" / config.canonical_filename,
        validation_report=root / "reports/validation_report.json",
        validation_errors=root / "reports/validation_errors.csv",
        scope_workbook=root / "scope_review" / config.scope_workbook_filename,
        scope_completed_workbook=root
        / "scope_review_app/exports"
        / config.scope_completed_workbook_filename,
        scope_source_csv=root / "scope_review" / config.scope_source_filename,
        scope_asset_manifest=root
        / "scope_review"
        / config.scope_asset_manifest_filename,
        scope_export_csv=root / "final_findings" / config.scope_export_filename,
    )


def _safe_extract_zip(source_zip: Path, destination: Path) -> None:
    if destination.exists() and any(destination.iterdir()):
        raise FindingsAnalysisError(f"ZIP destination is not empty: {destination}")
    destination.mkdir(parents=True, exist_ok=True)
    root = destination.resolve()
    with zipfile.ZipFile(source_zip, "r") as archive:
        bad_member = archive.testzip()
        if bad_member is not None:
            raise FindingsAnalysisError(f"frozen package CRC failure: {bad_member}")
        for member in archive.infolist():
            pure = PurePosixPath(member.filename.replace("\\", "/"))
            unix_mode = member.external_attr >> 16
            is_symlink = (unix_mode & 0o170000) == 0o120000
            if pure.is_absolute() or ".." in pure.parts or is_symlink:
                raise FindingsAnalysisError(
                    f"unsafe frozen-package member: {member.filename}"
                )
            target = (destination / Path(*pure.parts)).resolve()
            try:
                target.relative_to(root)
            except ValueError as exc:
                raise FindingsAnalysisError(
                    f"frozen-package member escapes snapshot: {member.filename}"
                ) from exc
        archive.extractall(destination)


def _copy_verified(source: Path, destination: Path, expected_hash: str) -> None:
    if destination.exists():
        raise FindingsAnalysisError(f"snapshot output already exists: {destination}")
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, destination)
    if destination.stat().st_size != source.stat().st_size:
        destination.unlink(missing_ok=True)
        raise FindingsAnalysisError(f"snapshot size mismatch: {destination}")
    copied_hash = sha256_file(destination)
    if copied_hash != expected_hash or copied_hash != sha256_file(source):
        destination.unlink(missing_ok=True)
        raise FindingsAnalysisError(f"snapshot SHA256 mismatch: {destination}")


def snapshot_authoritative_inputs(
    config: FindingsConfig,
    paths: WorkspacePaths,
) -> Path:
    """Copy verified inputs and safely extract only the copied package."""

    _copy_verified(
        config.completed_workbook,
        paths.completed_snapshot,
        config.completed_workbook_sha256,
    )
    _copy_verified(
        config.source_workbook,
        paths.source_workbook_snapshot,
        config.source_workbook_sha256,
    )
    _copy_verified(
        config.frozen_package,
        paths.package_snapshot,
        config.frozen_package_sha256,
    )
    _safe_extract_zip(paths.package_snapshot, paths.extracted_package)
    return paths.extracted_package


def _write_csv(
    path: Path,
    rows: Iterable[Mapping[str, Any]],
    columns: Sequence[str],
) -> None:
    if path.exists():
        raise FindingsAnalysisError(f"output already exists; overwrite is forbidden: {path}")
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(columns), extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({column: _normalize(row.get(column, "")) for column in columns})


def _write_json(path: Path, payload: Any) -> None:
    if path.exists():
        raise FindingsAnalysisError(f"output already exists; overwrite is forbidden: {path}")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def _write_text(path: Path, text: str) -> None:
    if path.exists():
        raise FindingsAnalysisError(f"output already exists; overwrite is forbidden: {path}")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def _manifest_rows(root: Path, *, exclude: set[str] | None = None) -> list[dict[str, str]]:
    excluded = exclude or set()
    rows: list[dict[str, str]] = []
    for path in sorted(p for p in root.rglob("*") if p.is_file()):
        relative = path.relative_to(root).as_posix()
        if relative in excluded or any(part.startswith(".staging-") for part in path.parts):
            continue
        rows.append(
            {
                "relative_path": relative,
                "size_bytes": str(path.stat().st_size),
                "sha256": sha256_file(path),
            }
        )
    return rows


def _find_asset_root(extracted_root: Path, canonical_frame: pd.DataFrame) -> Path:
    candidates: list[Path] = []
    first = canonical_frame.iloc[0]
    relpaths = (
        _normalize(first["original_image_relpath"]),
        _normalize(first["overlay_image_relpath"]),
    )
    for directory in [extracted_root, *[p for p in extracted_root.rglob("*") if p.is_dir()]]:
        if all((directory / rel).is_file() for rel in relpaths):
            candidates.append(directory)
    if len(candidates) != 1:
        raise FindingsAnalysisError(
            f"expected exactly one extracted asset root; matched={len(candidates)}"
        )
    return candidates[0]


def _fit_image(path: Path, max_width: int = 280, max_height: int = 180) -> XLImage:
    image = XLImage(str(path))
    scale = min(max_width / image.width, max_height / image.height, 1.0)
    image.width = int(image.width * scale)
    image.height = int(image.height * scale)
    return image


def _add_named_validation(
    sheet: Any,
    column_letter: str,
    defined_name: str,
    row_end: int,
) -> None:
    validation = DataValidation(
        type="list", formula1=f"={defined_name}", allow_blank=True
    )
    validation.error = "Select an approved FleetVision scope value."
    validation.errorTitle = "Invalid scope value"
    sheet.add_data_validation(validation)
    validation.add(f"{column_letter}2:{column_letter}{row_end}")


def build_scope_review_package(
    canonical_frame: pd.DataFrame,
    asset_root: Path,
    config: FindingsConfig,
    paths: WorkspacePaths,
) -> None:
    """Create deterministic scope source CSV, workbook, and asset manifest."""

    if len(canonical_frame) != config.expected_case_count:
        raise FindingsAnalysisError(
            f"canonical row count mismatch: {len(canonical_frame)}"
        )
    scope_frame = canonical_frame.loc[:, list(CANONICAL_COLUMNS)].copy()
    for column in SCOPE_COLUMNS:
        scope_frame[column] = "pending" if column == "scope_review_status" else ""
    scope_frame = scope_frame.loc[:, list(SCOPE_EXPORT_COLUMNS)].fillna("").astype(str)
    if paths.scope_source_csv.exists():
        raise FindingsAnalysisError("scope source output already exists")
    scope_frame.to_csv(
        paths.scope_source_csv,
        index=False,
        encoding="utf-8-sig",
        lineterminator="\n",
    )
    roundtrip = pd.read_csv(
        paths.scope_source_csv,
        dtype=str,
        keep_default_na=False,
        encoding="utf-8-sig",
    ).fillna("").astype(str)
    if roundtrip.columns.tolist() != list(SCOPE_EXPORT_COLUMNS) or not roundtrip.equals(
        scope_frame
    ):
        paths.scope_source_csv.unlink(missing_ok=True)
        raise FindingsAnalysisError("scope source CSV round-trip mismatch")

    asset_rows: list[dict[str, str]] = []
    for row in canonical_frame.to_dict(orient="records"):
        for asset_type, column in (
            ("original", "original_image_relpath"),
            ("overlay", "overlay_image_relpath"),
        ):
            relpath = _normalize(row[column]).replace("\\", "/")
            pure = PurePosixPath(relpath)
            if pure.is_absolute() or ".." in pure.parts or any(
                part.lower() == "test" for part in pure.parts
            ):
                raise FindingsAnalysisError(f"forbidden asset path: {relpath}")
            asset = asset_root / relpath
            if not asset.is_file():
                raise FindingsAnalysisError(f"missing scope asset: {asset}")
            asset_rows.append(
                {
                    "review_case_id": _normalize(row["review_case_id"]),
                    "asset_type": asset_type,
                    "relative_path": relpath,
                    "size_bytes": str(asset.stat().st_size),
                    "sha256": sha256_file(asset),
                }
            )
    _write_csv(
        paths.scope_asset_manifest,
        asset_rows,
        ("review_case_id", "asset_type", "relative_path", "size_bytes", "sha256"),
    )

    if paths.scope_workbook.exists():
        raise FindingsAnalysisError("scope Workbook already exists")
    workbook = Workbook()
    workbook.remove(workbook.active)
    instructions = workbook.create_sheet("Instructions")
    instructions["A1"] = "FleetVision Phase 04.5L — Severity Scope Review"
    instructions["A1"].font = Font(size=16, bold=True)
    instructions.append([])
    instructions.append(["Boundary", "Validation-only review. Do not modify GT, data, splits, Registry, or models."])
    instructions.append(["Scope", "Classify all 130 cases as in-scope, boundary-heavy, or catastrophic out-of-scope."])
    instructions.append(["Status", "F2 PASS remains advisory; retraining and deployment stay NOT_YET_APPROVED."])
    instructions.column_dimensions["A"].width = 24
    instructions.column_dimensions["B"].width = 105
    instructions.protection.sheet = True

    review = workbook.create_sheet("Scope_Review")
    headers = ("Original Preview", "Overlay Preview") + SCOPE_EXPORT_COLUMNS
    for index, header in enumerate(headers, start=1):
        cell = review.cell(row=1, column=index, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_index = {header: index for index, header in enumerate(headers, start=1)}
    row_end = len(scope_frame) + 1
    for row_index, row in enumerate(scope_frame.to_dict(orient="records"), start=2):
        review.row_dimensions[row_index].height = 145
        for column in SCOPE_EXPORT_COLUMNS:
            cell = review.cell(row=row_index, column=header_index[column], value=row[column])
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.protection = Protection(locked=column not in SCOPE_COLUMNS)
        for preview_column, rel_column in (
            (1, "original_image_relpath"),
            (2, "overlay_image_relpath"),
        ):
            asset = asset_root / row[rel_column]
            try:
                image = _fit_image(asset)
                review.add_image(image, f"{get_column_letter(preview_column)}{row_index}")
            except Exception:
                review.cell(
                    row=row_index,
                    column=preview_column,
                    value=f'=HYPERLINK("{row[rel_column]}","Open asset")',
                )
                review.cell(row=row_index, column=preview_column).protection = Protection(
                    locked=True
                )
    review.freeze_panes = "A2"
    review.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_end}"
    review.column_dimensions["A"].width = 42
    review.column_dimensions["B"].width = 42
    for column in range(3, len(headers) + 1):
        review.column_dimensions[get_column_letter(column)].width = 23

    options = workbook.create_sheet("Option_Lists")
    option_map = {
        "scope_review_status": config.scope_options["scope_review_status"],
        "scope_group": config.scope_options["scope_group"],
        "scope_reason": config.scope_options["scope_reason"],
        "operability": config.scope_options["operability"],
        "scope_confidence": config.scope_options["scope_confidence"],
    }
    for column_index, (key, values) in enumerate(option_map.items(), start=1):
        options.cell(row=1, column=column_index, value=key)
        for row_index, value in enumerate(values, start=2):
            options.cell(row=row_index, column=column_index, value=value)
        name = f"FV_SCOPE_{key.upper()}"
        ref = (
            f"'{options.title}'!${get_column_letter(column_index)}$2:"
            f"${get_column_letter(column_index)}${len(values) + 1}"
        )
        workbook.defined_names.add(DefinedName(name, attr_text=ref))
        _add_named_validation(
            review,
            get_column_letter(header_index[key]),
            name,
            row_end,
        )
    options.sheet_state = "hidden"
    options.protection.sheet = True

    manifest = workbook.create_sheet("Manifest")
    manifest.append(["key", "value"])
    manifest_rows = {
        "schema_version": "1",
        "case_count": len(scope_frame),
        "canonical_logical_fingerprint": logical_fingerprint(canonical_frame),
        "scope_source_csv_sha256": sha256_file(paths.scope_source_csv),
        "asset_manifest_sha256": sha256_file(paths.scope_asset_manifest),
        "test_split_read": False,
        "model_inference_executed": False,
        "annotation_modified": False,
        "training_started": False,
        "retraining_status": config.retraining_status,
        "deployment_acceptance": config.deployment_acceptance,
    }
    for key, value in sorted(manifest_rows.items()):
        manifest.append([key, _normalize(value)])
    manifest.protection.sheet = True

    progress = workbook.create_sheet("Progress_Summary")
    status_letter = get_column_letter(header_index["scope_review_status"])
    progress.append(["metric", "value"])
    progress.append(["total_cases", len(scope_frame)])
    progress.append(
        ["reviewed", f'=COUNTIF(Scope_Review!${status_letter}$2:${status_letter}${row_end},"reviewed")']
    )
    progress.append(
        ["pending", f'=COUNTIF(Scope_Review!${status_letter}$2:${status_letter}${row_end},"pending")']
    )
    progress.append(
        [
            "needs_adjudication",
            f'=COUNTIF(Scope_Review!${status_letter}$2:${status_letter}${row_end},"needs_adjudication")',
        ]
    )
    progress.append(["completion_rate", "=IF(B2=0,0,B3/B2)"])
    progress["B6"].number_format = "0.0%"
    progress.protection.sheet = True

    review.protection.sheet = True
    review.protection.password = "FleetVision"
    review.protection.autoFilter = False
    review.protection.sort = False
    if tuple(workbook.sheetnames) != SCOPE_WORKBOOK_SHEETS:
        raise FindingsAnalysisError(
            f"unexpected scope workbook sheets: {workbook.sheetnames}"
        )
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcOnSave = True
    workbook.save(paths.scope_workbook)


def read_scope_workbook(path: Path) -> pd.DataFrame:
    """Read source and scope columns, always releasing the Workbook handle."""

    if not path.is_file():
        raise FindingsAnalysisError(f"scope Workbook not found: {path}")
    workbook = load_workbook(path, data_only=False, read_only=True)
    try:
        if tuple(workbook.sheetnames) != SCOPE_WORKBOOK_SHEETS:
            raise FindingsAnalysisError(
                f"scope Workbook sheet contract mismatch: {workbook.sheetnames}"
            )
        sheet = workbook["Scope_Review"]
        headers = [
            _normalize(cell.value)
            for cell in next(sheet.iter_rows(min_row=1, max_row=1))
        ]
        if headers[:2] != ["Original Preview", "Overlay Preview"]:
            raise FindingsAnalysisError(
                "scope Workbook preview-column contract mismatch"
            )
        if headers[2:] != list(SCOPE_EXPORT_COLUMNS):
            raise FindingsAnalysisError(
                "scope Workbook column contract mismatch"
            )
        rows: list[dict[str, str]] = []
        for cells in sheet.iter_rows(min_row=2):
            row = {
                column: _normalize(cells[offset + 2].value)
                for offset, column in enumerate(SCOPE_EXPORT_COLUMNS)
            }
            if any(row.values()):
                rows.append(row)
        return (
            pd.DataFrame(rows, columns=SCOPE_EXPORT_COLUMNS)
            .fillna("")
            .astype(str)
        )
    finally:
        workbook.close()


def _parse_timestamp(value: str) -> bool:
    try:
        parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return False
    return parsed.tzinfo is not None


def validate_scope_dataframe(
    frame: pd.DataFrame,
    source_frame: pd.DataFrame,
    config: FindingsConfig,
) -> ScopeValidationResult:
    """Validate complete scope semantics and exact source/order immutability."""

    normalized = frame.copy().fillna("").astype(str)
    source = source_frame.copy().fillna("").astype(str)
    issues: list[dict[str, str]] = []

    def add(code: str, message: str, row_number: int | str = "") -> None:
        issues.append(
            {
                "row_number": _normalize(row_number),
                "error_code": code,
                "message": message,
            }
        )

    if normalized.columns.tolist() != list(SCOPE_EXPORT_COLUMNS):
        add("SCOPE_COLUMN_CONTRACT_MISMATCH", "scope columns or order changed")
        normalized = normalized.reindex(columns=SCOPE_EXPORT_COLUMNS, fill_value="")
    if source.columns.tolist() != list(SCOPE_EXPORT_COLUMNS):
        add("SCOPE_SOURCE_COLUMN_CONTRACT_MISMATCH", "scope source columns changed")
        source = source.reindex(columns=SCOPE_EXPORT_COLUMNS, fill_value="")
    if len(normalized) != config.expected_case_count:
        add(
            "SCOPE_ROW_COUNT_MISMATCH",
            f"expected {config.expected_case_count} rows, got {len(normalized)}",
        )
    if len(source) != config.expected_case_count:
        add("SCOPE_SOURCE_ROW_COUNT_MISMATCH", "F1 source row count changed")

    ids = normalized["review_case_id"] if "review_case_id" in normalized else pd.Series(dtype=str)
    if len(ids) and (ids.eq("").any() or ids.nunique(dropna=False) != len(ids)):
        add("SCOPE_IDENTITY_INVALID", "review_case_id must be unique and nonblank")
    if normalized["review_case_id"].tolist() != source["review_case_id"].tolist():
        add("SCOPE_ROW_ORDER_CHANGED", "scope Workbook identity order changed")

    source_columns = list(CANONICAL_COLUMNS)
    min_rows = min(len(normalized), len(source))
    if min_rows:
        left = normalized.iloc[:min_rows].loc[:, source_columns].reset_index(drop=True)
        right = source.iloc[:min_rows].loc[:, source_columns].reset_index(drop=True)
        if not left.equals(right):
            for index in range(min_rows):
                for column in source_columns:
                    if left.at[index, column] != right.at[index, column]:
                        add(
                            "SCOPE_SOURCE_FIELD_CHANGED",
                            f"immutable source field changed: {column}",
                            index + 2,
                        )

    catastrophic_reasons = {
        "structural_damage",
        "catastrophic_collision",
        "extensive_multi_panel_damage",
        "vehicle_integrity_compromised",
    }
    counts = Counter()
    for index, row in normalized.iterrows():
        row_number = index + 2
        values = {column: _normalize(row[column]) for column in SCOPE_EXPORT_COLUMNS}
        status = values["scope_review_status"]
        counts[status] += 1
        for column in (
            "scope_review_status",
            "scope_group",
            "scope_reason",
            "operability",
            "scope_confidence",
        ):
            if values[column] not in config.scope_options[column]:
                add(
                    "INVALID_SCOPE_CONTROLLED_VALUE",
                    f"{column}={values[column]!r} is not approved",
                    row_number,
                )
        if status != "reviewed":
            add("INCOMPLETE_SCOPE_REVIEW", f"scope_review_status={status!r}", row_number)
        if status == "reviewed":
            for required in (
                "scope_group",
                "scope_reason",
                "operability",
                "scope_confidence",
                "scope_reviewer",
                "scope_reviewed_at_utc",
            ):
                if not values[required]:
                    add("MISSING_SCOPE_FIELD", f"{required} is required", row_number)
            if values["scope_reviewed_at_utc"] and not _parse_timestamp(
                values["scope_reviewed_at_utc"]
            ):
                add(
                    "INVALID_SCOPE_TIMESTAMP",
                    "scope_reviewed_at_utc must be timezone-aware ISO 8601",
                    row_number,
                )
        notes = values["scope_reviewer_notes"]
        if values["scope_confidence"] == "low" and not notes:
            add("LOW_SCOPE_CONFIDENCE_NOTES_REQUIRED", "low confidence requires notes", row_number)
        if values["scope_reason"] == "other" and not notes:
            add("OTHER_SCOPE_REASON_NOTES_REQUIRED", "other reason requires notes", row_number)
        if values["scope_group"] == "OUT_OF_SCOPE_CATASTROPHIC":
            if values["scope_reason"] not in catastrophic_reasons:
                add(
                    "CATASTROPHIC_REASON_REQUIRED",
                    "catastrophic group requires an approved catastrophic reason",
                    row_number,
                )
            if values["operability"] == "drivable_or_likely_drivable" and not notes:
                add(
                    "CATASTROPHIC_DRIVABLE_NOTES_REQUIRED",
                    "catastrophic + likely drivable requires notes",
                    row_number,
                )
        if values["scope_group"] == "IN_SCOPE_LIGHT_MODERATE" and values[
            "scope_reason"
        ] in {"catastrophic_collision", "vehicle_integrity_compromised"}:
            add(
                "IN_SCOPE_CATASTROPHIC_REASON_FORBIDDEN",
                "in-scope rows cannot use catastrophic reasons",
                row_number,
            )
        if values["scope_reason"] == "insufficient_visual_evidence":
            if values["scope_confidence"] != "low" or not notes:
                add(
                    "INSUFFICIENT_EVIDENCE_CONTRACT_VIOLATION",
                    "insufficient evidence requires low confidence and notes",
                    row_number,
                )

    result_counts = {
        "reviewed": int((normalized["scope_review_status"] == "reviewed").sum()),
        "pending": int((normalized["scope_review_status"] == "pending").sum()),
        "needs_adjudication": int(
            (normalized["scope_review_status"] == "needs_adjudication").sum()
        ),
    }
    return ScopeValidationResult(
        passed=not issues,
        row_count=len(normalized),
        issue_count=len(issues),
        issues=tuple(issues),
        counts=result_counts,
    )


def export_scope_classification(
    config: FindingsConfig,
    workbook_path: Path,
    source_csv: Path,
    output_csv: Path,
) -> ScopeValidationResult:
    """Validate and transactionally export deterministic UTF-8-SIG scope CSV."""

    if output_csv.exists():
        raise FindingsAnalysisError(
            f"scope export already exists; overwrite is forbidden: {output_csv}"
        )
    frame = read_scope_workbook(workbook_path)
    source = pd.read_csv(
        source_csv,
        dtype=str,
        keep_default_na=False,
        encoding="utf-8-sig",
    ).fillna("").astype(str)
    result = validate_scope_dataframe(frame, source, config)
    if not result.passed:
        preview = "; ".join(issue["error_code"] for issue in result.issues[:8])
        raise FindingsAnalysisError(
            f"scope export blocked by {result.issue_count} issues: {preview}"
        )
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    handle, temporary_name = tempfile.mkstemp(
        prefix=f".{output_csv.name}.staging-",
        suffix=".csv",
        dir=output_csv.parent,
    )
    os.close(handle)
    temporary = Path(temporary_name)
    try:
        frame.to_csv(
            temporary,
            index=False,
            encoding="utf-8-sig",
            lineterminator="\n",
        )
        roundtrip = pd.read_csv(
            temporary,
            dtype=str,
            keep_default_na=False,
            encoding="utf-8-sig",
        ).fillna("").astype(str)
        if roundtrip.columns.tolist() != list(SCOPE_EXPORT_COLUMNS) or not roundtrip.equals(
            frame
        ):
            raise FindingsAnalysisError("scope CSV UTF-8-SIG round-trip mismatch")
        temporary.replace(output_csv)
    except Exception:
        temporary.unlink(missing_ok=True)
        raise
    return result


def _distribution(frame: pd.DataFrame, column: str) -> list[dict[str, Any]]:
    counts = frame[column].value_counts(dropna=False).sort_index()
    total = len(frame)
    return [
        {
            "value": _normalize(value),
            "count": int(count),
            "percentage": 0.0 if total == 0 else round(int(count) * 100.0 / total, 6),
        }
        for value, count in counts.items()
    ]


def _cross_tab(frame: pd.DataFrame, row: str, column: str) -> list[dict[str, Any]]:
    table = pd.crosstab(frame[row], frame[column], dropna=False)
    row_values = sorted(_normalize(value) for value in table.index)
    column_values = sorted(_normalize(value) for value in table.columns)
    rows: list[dict[str, Any]] = []
    for row_value in row_values:
        record: dict[str, Any] = {row: row_value}
        for column_value in column_values:
            record[column_value] = int(table.loc[row_value, column_value])
        rows.append(record)
    return rows


def _normalized_counts(counts: Mapping[str, int]) -> dict[str, float]:
    total = sum(int(value) for value in counts.values())
    if total == 0:
        return {key: 0.0 for key in counts}
    return {key: int(value) / total for key, value in counts.items()}


def total_variation_distance(
    left_counts: Mapping[str, int],
    right_counts: Mapping[str, int],
) -> float:
    """Return half of the L1 distance between normalized categorical counts."""

    left = _normalized_counts(left_counts)
    right = _normalized_counts(right_counts)
    keys = set(left) | set(right)
    return 0.5 * sum(abs(left.get(key, 0.0) - right.get(key, 0.0)) for key in keys)


def _value_counts(frame: pd.DataFrame, column: str) -> dict[str, int]:
    return {
        _normalize(key): int(value)
        for key, value in frame[column].value_counts(dropna=False).sort_index().items()
    }


def _share(mask: pd.Series) -> float:
    return 0.0 if len(mask) == 0 else float(mask.sum()) / float(len(mask))


def build_findings_payload(
    combined: pd.DataFrame,
    config: FindingsConfig,
) -> dict[str, Any]:
    """Build deterministic distributions, cross-tabs, and composition analysis."""

    dimensions = (
        "error_disposition",
        "primary_root_cause",
        "secondary_root_cause",
        "annotation_quality",
        "annotation_defect_type",
        "recommended_action",
        "retraining_priority",
        "scope_group",
        "operability",
        "scope_reason",
        "scope_confidence",
    )
    for column in dimensions:
        if column not in combined.columns:
            raise FindingsAnalysisError(f"combined findings missing column: {column}")
    distributions = {column: _distribution(combined, column) for column in dimensions}
    cross_tabs = {
        "scope_group_by_error_disposition": _cross_tab(
            combined, "scope_group", "error_disposition"
        ),
        "scope_group_by_primary_root_cause": _cross_tab(
            combined, "scope_group", "primary_root_cause"
        ),
        "scope_group_by_annotation_quality": _cross_tab(
            combined, "scope_group", "annotation_quality"
        ),
        "scope_group_by_recommended_action": _cross_tab(
            combined, "scope_group", "recommended_action"
        ),
        "scope_group_by_retraining_priority": _cross_tab(
            combined, "scope_group", "retraining_priority"
        ),
    }
    in_scope = combined[combined["scope_group"] == "IN_SCOPE_LIGHT_MODERATE"].copy()
    comparison_dimensions = (
        "error_disposition",
        "primary_root_cause",
        "annotation_quality",
        "recommended_action",
        "retraining_priority",
    )
    tvd = {
        column: round(
            total_variation_distance(
                _value_counts(combined, column), _value_counts(in_scope, column)
            ),
            12,
        )
        for column in comparison_dimensions
    }
    non_scope_share = _share(
        combined["scope_group"] != "IN_SCOPE_LIGHT_MODERATE"
    )
    max_tvd = max(tvd.values(), default=0.0)
    rules = config.recommendation_rules
    composition_distortion = (
        non_scope_share >= rules["scope_rebalancing_non_scope_share"]
        or max_tvd >= rules["distribution_total_variation_threshold"]
    )
    all_tradeoff_share = _share(
        combined["error_disposition"] == "expected_threshold_tradeoff"
    )
    in_scope_tradeoff_share = _share(
        in_scope["error_disposition"] == "expected_threshold_tradeoff"
    )
    threshold_sensitive = (
        abs(all_tradeoff_share - in_scope_tradeoff_share)
        >= rules["threshold_tradeoff_share_delta"]
    )
    return {
        "row_count": len(combined),
        "in_scope_row_count": len(in_scope),
        "distributions": distributions,
        "cross_tabs": cross_tabs,
        "all_vs_in_scope": {
            "total_variation_distance": tvd,
            "max_total_variation_distance": round(max_tvd, 12),
            "non_scope_share": round(non_scope_share, 12),
            "composition_distortion_detected": composition_distortion,
        },
        "threshold_candidate_0_20": {
            "designation": "BALANCED_VALIDATION_THRESHOLD_CANDIDATE",
            "all_tradeoff_share": round(all_tradeoff_share, 12),
            "in_scope_tradeoff_share": round(in_scope_tradeoff_share, 12),
            "composition_sensitive": threshold_sensitive,
            "deployment_threshold": False,
        },
    }


def choose_primary_recommendation(
    *,
    additional_review: bool,
    data_correction: bool,
    scope_rebalancing: bool,
    retraining_justified: bool,
) -> str:
    if additional_review:
        return "ADDITIONAL_REVIEW_REQUIRED"
    if data_correction:
        return "DATA_CORRECTION_REQUIRED_BEFORE_RETRAINING"
    if scope_rebalancing:
        return "SCOPE_REBALANCING_REQUIRED_BEFORE_RETRAINING"
    if retraining_justified:
        return "RETRAINING_PROPOSAL_JUSTIFIED"
    return "NO_RETRAINING_RECOMMENDED"


def classify_recommendation(
    combined: pd.DataFrame,
    payload: Mapping[str, Any],
    config: FindingsConfig,
) -> RecommendationResult:
    """Issue exactly one advisory recommendation using fail-closed precedence."""

    rules = config.recommendation_rules
    low_confidence_share = _share(combined["scope_confidence"] == "low")
    insufficient = bool((combined["scope_reason"] == "insufficient_visual_evidence").any())
    correction_count = int((combined["correction_proposal_required"] == "yes").sum())
    defect_count = int((combined["annotation_quality"] == "defect_suspected").sum())
    in_scope = combined[combined["scope_group"] == "IN_SCOPE_LIGHT_MODERATE"]
    in_scope_priority_share = _share(
        in_scope["retraining_priority"].isin(["medium", "high"])
    )
    in_scope_confirmed_error_share = _share(
        in_scope["error_disposition"] == "confirmed_model_error"
    )
    additional_review = insufficient or (
        low_confidence_share > rules["additional_review_low_confidence_share"]
    )
    data_correction = correction_count > 0 or defect_count > 0
    scope_rebalancing = bool(
        payload["all_vs_in_scope"]["composition_distortion_detected"]
    )
    retraining_justified = (
        len(in_scope) > 0
        and in_scope_priority_share
        >= rules["retraining_in_scope_priority_share"]
        and in_scope_confirmed_error_share
        >= rules["retraining_in_scope_confirmed_error_share"]
    )
    primary = choose_primary_recommendation(
        additional_review=additional_review,
        data_correction=data_correction,
        scope_rebalancing=scope_rebalancing,
        retraining_justified=retraining_justified,
    )
    reasons: list[str] = []
    if additional_review:
        reasons.append(
            "Low-confidence or insufficient-visual-evidence scope review requires additional adjudication."
        )
    if data_correction:
        reasons.append(
            "Annotation correction proposals or suspected annotation defects must be resolved before retraining."
        )
    if scope_rebalancing:
        reasons.append(
            "Heavy/catastrophic composition materially changes the validation-error profile."
        )
    if retraining_justified:
        reasons.append(
            "The in-scope subset contains sufficient confirmed model errors and medium/high retraining priority."
        )
    if not reasons:
        reasons.append("Current reviewed evidence does not justify a retraining proposal.")
    metrics = {
        "low_confidence_share": round(low_confidence_share, 12),
        "insufficient_visual_evidence_count": float(
            (combined["scope_reason"] == "insufficient_visual_evidence").sum()
        ),
        "annotation_correction_proposal_count": float(correction_count),
        "annotation_defect_suspected_count": float(defect_count),
        "non_scope_share": float(payload["all_vs_in_scope"]["non_scope_share"]),
        "max_total_variation_distance": float(
            payload["all_vs_in_scope"]["max_total_variation_distance"]
        ),
        "in_scope_priority_share": round(in_scope_priority_share, 12),
        "in_scope_confirmed_model_error_share": round(
            in_scope_confirmed_error_share, 12
        ),
    }
    return RecommendationResult(primary, tuple(reasons), metrics)


def _recommendation_payload(
    result: RecommendationResult,
    config: FindingsConfig,
) -> dict[str, Any]:
    return {
        "primary_recommendation": result.primary,
        "advisory_only": True,
        "retraining_status": config.retraining_status,
        "deployment_acceptance": config.deployment_acceptance,
        "reasons": list(result.reasons),
        "metrics": dict(result.metrics),
    }


def _render_report_markdown(
    payload: Mapping[str, Any],
    recommendation: RecommendationResult,
    config: FindingsConfig,
) -> str:
    lines = [
        "# FleetVision Phase 04.5L Completed Review Findings",
        "",
        f"- Reviewed cases: {payload['row_count']}",
        f"- In-scope light/moderate cases: {payload['in_scope_row_count']}",
        f"- Primary advisory recommendation: `{recommendation.primary}`",
        "- Recommendation is advisory only.",
        f"- Retraining status: `{config.retraining_status}`",
        f"- Deployment acceptance: `{config.deployment_acceptance}`",
        "- Test split read: false",
        "- Model inference executed: false",
        "- Annotation modified: false",
        "- Training started: false",
        "",
        "## Required analysis outputs",
        "",
        "1. Error-disposition distribution is recorded in the JSON report.",
        "2. Primary-root-cause distribution is recorded.",
        "3. Secondary-root-cause distribution is recorded.",
        "4. Annotation-quality distribution is recorded.",
        "5. Annotation-defect-type distribution is recorded.",
        "6. Recommended-action distribution is recorded.",
        "7. Retraining-priority distribution is recorded.",
        "8. Annotation correction proposal count remains proposal-only and unapplied.",
        "9. Severity-scope counts and percentages are recorded.",
        "10. Operability and scope-reason distributions are recorded.",
        "11. Scope group × error disposition cross-tabulation is recorded.",
        "12. Scope group × root cause cross-tabulation is recorded.",
        "13. Scope group × annotation quality cross-tabulation is recorded.",
        "14. Scope group × recommended action cross-tabulation is recorded.",
        "15. Scope group × retraining priority cross-tabulation is recorded.",
        "16. All-130 versus in-scope comparison is recorded.",
        "17. Heavy/catastrophic composition distortion is assessed.",
        "18. Threshold candidate 0.20 composition sensitivity is assessed without tuning.",
        "19. Ranked recommendations are advisory and require a later governance Gate.",
        "",
        "## Recommendation rationale",
    ]
    lines.extend(f"- {reason}" for reason in recommendation.reasons)
    lines.extend(
        [
            "",
            "## Composition assessment",
            f"- Non-scope share: {payload['all_vs_in_scope']['non_scope_share']:.6f}",
            f"- Maximum total variation distance: {payload['all_vs_in_scope']['max_total_variation_distance']:.6f}",
            f"- Composition distortion detected: {str(payload['all_vs_in_scope']['composition_distortion_detected']).lower()}",
            f"- Threshold candidate 0.20 composition sensitive: {str(payload['threshold_candidate_0_20']['composition_sensitive']).lower()}",
            "",
        ]
    )
    return "\n".join(lines)


def _write_blocked_gate(root: Path, stage: str, error: Exception) -> None:
    evidence = root / "evidence"
    evidence.mkdir(parents=True, exist_ok=True)
    path = evidence / f"{stage.lower()}_blocked_gate_result.json"
    if path.exists():
        return
    path.write_text(
        json.dumps(
            {
                "gate_id": stage,
                "outcome": "BLOCKED",
                "classification": f"{stage}_BLOCKED",
                "reason": str(error),
                "test_split_read": False,
                "model_inference_executed": False,
                "annotation_modified": False,
                "training_started": False,
                "retraining_status": "NOT_YET_APPROVED",
                "deployment_acceptance": "NOT_YET_APPROVED",
            },
            ensure_ascii=False,
            indent=2,
            sort_keys=True,
        )
        + "\n",
        encoding="utf-8",
    )


def _source_hash_rows(
    config: FindingsConfig,
    paths: WorkspacePaths,
) -> list[dict[str, str]]:
    specs = (
        (
            "completed_workbook",
            config.completed_workbook,
            paths.completed_snapshot,
            config.completed_workbook_size,
            config.completed_workbook_sha256,
        ),
        (
            "source_workbook",
            config.source_workbook,
            paths.source_workbook_snapshot,
            "",
            config.source_workbook_sha256,
        ),
        (
            "frozen_package",
            config.frozen_package,
            paths.package_snapshot,
            "",
            config.frozen_package_sha256,
        ),
    )
    rows: list[dict[str, str]] = []
    for artifact, original, snapshot, expected_size, expected_hash in specs:
        rows.append(
            {
                "artifact": artifact,
                "original_path": str(original),
                "snapshot_path": str(snapshot),
                "expected_size_bytes": str(expected_size),
                "actual_size_bytes": str(original.stat().st_size),
                "expected_sha256": expected_hash,
                "actual_sha256": sha256_file(original),
                "snapshot_sha256": sha256_file(snapshot),
                "match": str(
                    sha256_file(original) == expected_hash
                    and sha256_file(snapshot) == expected_hash
                ).lower(),
            }
        )
    return rows


def run_f1(
    config: FindingsConfig,
    expected_head: str,
    *,
    timestamp: str | None = None,
    git_runner: GitRunner | None = None,
) -> WorkspacePaths:
    """Run F1 only: validate completed review and create scope package."""

    inspect_repository_state(config.project_root, expected_head, runner=git_runner)
    verify_authoritative_inputs(config)
    token = _timestamp_token(timestamp)
    target_name = f"{config.workspace_prefix}_{token}"
    parent_inventory: list[dict[str, str]] = []
    if config.workspace_parent.exists():
        for entry in sorted(config.workspace_parent.iterdir(), key=lambda item: item.name):
            parent_inventory.append(
                {
                    "entry_name": entry.name,
                    "entry_type": "directory" if entry.is_dir() else "file",
                    "target_name_collision": str(entry.name == target_name).lower(),
                }
            )
    paths = create_workspace(config, timestamp=token)
    try:
        _write_csv(
            paths.root / "evidence/workspace_before.csv",
            parent_inventory,
            ("entry_name", "entry_type", "target_name_collision"),
        )
        asset_snapshot = snapshot_authoritative_inputs(config, paths)
        existing_config: ReviewConfig = load_existing_review_config(
            config.existing_review_config_path, config.project_root
        )
        completed_frame = read_workbook_dataframe(paths.completed_snapshot)
        if len(completed_frame) != config.expected_case_count:
            raise FindingsAnalysisError(
                f"completed Workbook row count mismatch: {len(completed_frame)}"
            )
        actual_fingerprint = logical_fingerprint(completed_frame)
        if actual_fingerprint != config.completed_logical_fingerprint:
            raise FindingsAnalysisError(
                "completed Workbook logical fingerprint mismatch: "
                f"expected={config.completed_logical_fingerprint} actual={actual_fingerprint}"
            )
        asset_root = _find_asset_root(asset_snapshot, completed_frame)
        export_result = export_review_workbook(
            existing_config, paths.completed_snapshot, paths.canonical_csv
        )
        if export_result.logical_fingerprint != config.completed_logical_fingerprint:
            raise FindingsAnalysisError("canonical export logical fingerprint mismatch")
        validation = validate_canonical_csv(
            existing_config,
            paths.canonical_csv,
            workbook_path=paths.completed_snapshot,
            batch_root=asset_root,
        )
        if not validation.passed:
            raise FindingsAnalysisError(
                f"existing completed-review validator blocked F1: {validation.issue_count} issues"
            )
        write_validation_outputs(
            validation, paths.validation_report, paths.validation_errors
        )
        report = json.loads(paths.validation_report.read_text(encoding="utf-8"))
        if report.get("classification") != config.expected_validator_classification:
            raise FindingsAnalysisError("unexpected existing validator classification")
        summarize_canonical_review(
            existing_config,
            paths.canonical_csv,
            paths.root,
            asset_root=asset_root,
        )
        canonical_frame = pd.read_csv(
            paths.canonical_csv,
            dtype=str,
            keep_default_na=False,
            encoding="utf-8-sig",
        ).fillna("").astype(str)
        build_scope_review_package(canonical_frame, asset_root, config, paths)
        source_hashes = paths.root / "evidence/source_hashes.csv"
        _write_csv(
            source_hashes,
            _source_hash_rows(config, paths),
            (
                "artifact",
                "original_path",
                "snapshot_path",
                "expected_size_bytes",
                "actual_size_bytes",
                "expected_sha256",
                "actual_sha256",
                "snapshot_sha256",
                "match",
            ),
        )
        gate = {
            "gate_id": "PHASE_04_5L_COMPLETED_REVIEW_VALIDATION_AND_FINDINGS_ANALYSIS_F1",
            "outcome": "PASS",
            "classification": "PHASE_04_5L_COMPLETED_REVIEW_VALIDATED_AND_SCOPE_REVIEW_PACKAGE_CREATED",
            "review_cases": len(canonical_frame),
            "scope_reviewed": 0,
            "pending": len(canonical_frame),
            "needs_adjudication": 0,
            "completed_logical_fingerprint": actual_fingerprint,
            "scope_source_csv_sha256": sha256_file(paths.scope_source_csv),
            "scope_workbook_sha256": sha256_file(paths.scope_workbook),
            "scope_asset_manifest_sha256": sha256_file(paths.scope_asset_manifest),
            "human_review_interface": "LOCAL_STREAMLIT_TRADITIONAL_CHINESE",
            "live_review_state": "SQLITE",
            "scope_workbook_role": "READ_ONLY_EXPORT_TEMPLATE",
            "test_split_read": False,
            "model_inference_executed": False,
            "annotation_modified": False,
            "training_started": False,
            "retraining_status": config.retraining_status,
            "deployment_acceptance": config.deployment_acceptance,
        }
        _write_json(paths.root / "evidence/f1_gate_result.json", gate)
        _write_csv(
            paths.root / "evidence/F1_SHA256SUMS.csv",
            _manifest_rows(
                paths.root,
                exclude={"evidence/F1_SHA256SUMS.csv"},
            ),
            ("relative_path", "size_bytes", "sha256"),
        )
        return paths
    except Exception as exc:
        _write_blocked_gate(
            paths.root,
            "PHASE_04_5L_COMPLETED_REVIEW_VALIDATION_AND_FINDINGS_ANALYSIS_F1",
            exc,
        )
        for staging in paths.root.rglob(".staging-*"):
            if staging.is_dir():
                shutil.rmtree(staging, ignore_errors=True)
            else:
                staging.unlink(missing_ok=True)
        raise


def _verify_f1_checksum_manifest(root: Path) -> None:
    """Verify every immutable F1 output before reading app-completed results."""

    manifest_path = root / "evidence/F1_SHA256SUMS.csv"
    if not manifest_path.is_file():
        raise FindingsAnalysisError("F1 checksum manifest is missing")
    manifest = pd.read_csv(
        manifest_path,
        dtype=str,
        keep_default_na=False,
        encoding="utf-8-sig",
    ).fillna("").astype(str)
    if manifest.columns.tolist() != ["relative_path", "size_bytes", "sha256"]:
        raise FindingsAnalysisError("F1 checksum manifest schema changed")
    if manifest["relative_path"].eq("").any() or manifest["relative_path"].duplicated().any():
        raise FindingsAnalysisError("F1 checksum manifest contains blank or duplicate paths")
    for row in manifest.to_dict(orient="records"):
        relative = row["relative_path"].replace("\\", "/")
        pure = PurePosixPath(relative)
        if pure.is_absolute() or ".." in pure.parts:
            raise FindingsAnalysisError(f"unsafe F1 checksum path: {relative}")
        path = root / Path(*pure.parts)
        if not path.is_file():
            raise FindingsAnalysisError(f"F1 checksummed output missing: {relative}")
        if str(path.stat().st_size) != row["size_bytes"]:
            raise FindingsAnalysisError(f"F1 checksummed output size changed: {relative}")
        if sha256_file(path) != row["sha256"].upper():
            raise FindingsAnalysisError(f"F1 checksummed output hash changed: {relative}")


def _verify_f1_workspace(config: FindingsConfig, root: Path) -> WorkspacePaths:
    paths = WorkspacePaths(
        root=root,
        completed_snapshot=root
        / "input_snapshot/completed_workbook"
        / config.completed_workbook.name,
        source_workbook_snapshot=root
        / "input_snapshot/source_workbook"
        / config.source_workbook.name,
        package_snapshot=root
        / "input_snapshot/formal_package"
        / config.frozen_package.name,
        extracted_package=root / "input_snapshot/extracted_package",
        canonical_csv=root / "canonical" / config.canonical_filename,
        validation_report=root / "reports/validation_report.json",
        validation_errors=root / "reports/validation_errors.csv",
        scope_workbook=root / "scope_review" / config.scope_workbook_filename,
        scope_completed_workbook=root
        / "scope_review_app/exports"
        / config.scope_completed_workbook_filename,
        scope_source_csv=root / "scope_review" / config.scope_source_filename,
        scope_asset_manifest=root
        / "scope_review"
        / config.scope_asset_manifest_filename,
        scope_export_csv=root / "final_findings" / config.scope_export_filename,
    )
    f1_gate = root / "evidence/f1_gate_result.json"
    required = (
        f1_gate,
        paths.completed_snapshot,
        paths.source_workbook_snapshot,
        paths.package_snapshot,
        paths.canonical_csv,
        paths.scope_workbook,
        paths.scope_source_csv,
        paths.scope_asset_manifest,
    )
    missing = [str(path) for path in required if not path.is_file()]
    if missing:
        raise FindingsAnalysisError(f"F1 workspace missing required files: {missing}")
    gate = json.loads(f1_gate.read_text(encoding="utf-8"))
    if gate.get("outcome") != "PASS" or gate.get("classification") != (
        "PHASE_04_5L_COMPLETED_REVIEW_VALIDATED_AND_SCOPE_REVIEW_PACKAGE_CREATED"
    ):
        raise FindingsAnalysisError("F1 workspace does not contain the expected PASS gate")
    if sha256_file(paths.completed_snapshot) != config.completed_workbook_sha256:
        raise FindingsAnalysisError("F1 completed Workbook snapshot changed")
    if sha256_file(paths.source_workbook_snapshot) != config.source_workbook_sha256:
        raise FindingsAnalysisError("F1 source Workbook snapshot changed")
    if sha256_file(paths.package_snapshot) != config.frozen_package_sha256:
        raise FindingsAnalysisError("F1 package snapshot changed")
    if gate.get("scope_source_csv_sha256") != sha256_file(paths.scope_source_csv):
        raise FindingsAnalysisError("F1 scope source CSV changed")
    if gate.get("scope_asset_manifest_sha256") != sha256_file(paths.scope_asset_manifest):
        raise FindingsAnalysisError("F1 scope asset manifest changed")
    _verify_f1_checksum_manifest(root)
    return paths



def verify_completed_scope_review_export(
    config: FindingsConfig,
    paths: WorkspacePaths,
) -> Path:
    """Verify the app-completed scope Workbook and its export evidence."""

    completed = paths.scope_completed_workbook
    evidence = completed.parent / "scope_review_export_result.json"
    if not completed.is_file():
        raise FindingsAnalysisError(
            f"completed scope Workbook is missing: {completed}"
        )
    if not evidence.is_file():
        raise FindingsAnalysisError(
            f"scope review export evidence is missing: {evidence}"
        )
    payload = json.loads(evidence.read_text(encoding="utf-8"))
    if payload.get("outcome") != "PASS" or payload.get("classification") != (
        "LOCAL_SCOPE_REVIEW_APP_COMPLETED_WORKBOOK_EXPORTED"
    ):
        raise FindingsAnalysisError(
            "scope review export evidence does not contain the expected PASS classification"
        )
    if (
        int(payload.get("review_cases", -1)) != config.expected_case_count
        or int(payload.get("reviewed", -1)) != config.expected_case_count
        or int(payload.get("pending", -1)) != 0
        or int(payload.get("needs_adjudication", -1)) != 0
    ):
        raise FindingsAnalysisError("scope review export completion counts are invalid")

    declared_path = Path(_normalize(payload.get("completed_scope_workbook"))).resolve()
    if declared_path != completed.resolve():
        raise FindingsAnalysisError("scope review export path does not match F2 workspace")
    declared_hash = _require_hash(
        "scope review completed Workbook SHA256",
        payload.get("completed_scope_workbook_sha256"),
    )
    if sha256_file(completed) != declared_hash:
        raise FindingsAnalysisError("completed scope Workbook SHA256 mismatch")

    source_hash_contract = {
        "source_scope_csv_sha256": sha256_file(paths.scope_source_csv),
        "source_scope_template_sha256": sha256_file(paths.scope_workbook),
        "source_scope_asset_manifest_sha256": sha256_file(paths.scope_asset_manifest),
    }
    for key, actual in source_hash_contract.items():
        if _normalize(payload.get(key)).upper() != actual:
            raise FindingsAnalysisError(f"scope review export source hash mismatch: {key}")

    for key in (
        "test_split_read",
        "model_inference_executed",
        "annotation_modified",
        "training_started",
    ):
        if payload.get(key) is not False:
            raise FindingsAnalysisError(f"scope review export safety declaration changed: {key}")
    if payload.get("retraining_status") != config.retraining_status:
        raise FindingsAnalysisError("scope review export retraining status changed")
    if payload.get("deployment_acceptance") != config.deployment_acceptance:
        raise FindingsAnalysisError("scope review export deployment status changed")
    return completed


def _ranked_action_recommendations(combined: pd.DataFrame) -> list[dict[str, Any]]:
    """Rank non-no-action recommendations by priority-weighted reviewed count."""

    priority_weight = {"high": 3, "medium": 2, "low": 1, "not_applicable": 0}
    rows: list[dict[str, Any]] = []
    for action, group in combined.groupby("recommended_action", sort=True):
        if action == "no_action":
            continue
        counts = Counter(group["retraining_priority"].astype(str))
        score = sum(priority_weight.get(priority, 0) * int(count) for priority, count in counts.items())
        rows.append(
            {
                "recommended_action": _normalize(action),
                "case_count": int(len(group)),
                "priority_score": int(score),
                "high_count": int(counts.get("high", 0)),
                "medium_count": int(counts.get("medium", 0)),
                "low_count": int(counts.get("low", 0)),
            }
        )
    return sorted(
        rows,
        key=lambda row: (
            -row["priority_score"],
            -row["case_count"],
            row["recommended_action"],
        ),
    )


def run_f2(
    config: FindingsConfig,
    workspace_root: Path,
    expected_head: str,
    *,
    git_runner: GitRunner | None = None,
) -> RecommendationResult:
    """Run F2 only after all scope-review rows are complete."""

    inspect_repository_state(config.project_root, expected_head, runner=git_runner)
    verify_authoritative_inputs(config)
    root = workspace_root.resolve()
    paths = _verify_f1_workspace(config, root)
    completed_scope_workbook = verify_completed_scope_review_export(config, paths)
    final_outputs = (
        paths.scope_export_csv,
        root / "final_findings/severity_scope_summary.json",
        root / "final_findings/severity_scope_summary.md",
        root / "final_findings/phase04_5l_findings_report.json",
        root / "final_findings/phase04_5l_findings_report.md",
        root / "final_findings/retraining_recommendation.json",
        root / "evidence/gate_result.json",
        root / "evidence/SHA256SUMS.csv",
    )
    if any(path.exists() for path in final_outputs):
        raise FindingsAnalysisError("one or more final F2 outputs already exist")
    created_outputs: list[Path] = []
    try:
        scope_result = export_scope_classification(
            config,
            completed_scope_workbook,
            paths.scope_source_csv,
            paths.scope_export_csv,
        )
        created_outputs.append(paths.scope_export_csv)
        canonical = pd.read_csv(
            paths.canonical_csv,
            dtype=str,
            keep_default_na=False,
            encoding="utf-8-sig",
        ).fillna("").astype(str)
        scope = pd.read_csv(
            paths.scope_export_csv,
            dtype=str,
            keep_default_na=False,
            encoding="utf-8-sig",
        ).fillna("").astype(str)
        if canonical["review_case_id"].tolist() != scope["review_case_id"].tolist():
            raise FindingsAnalysisError("canonical and scope identity order mismatch")
        combined = scope.copy()
        payload = build_findings_payload(combined, config)
        recommendation = classify_recommendation(combined, payload, config)
        recommendation_json = _recommendation_payload(recommendation, config)
        scope_summary = {
            "gate_id": "PHASE_04_5L_SCOPE_SUMMARY",
            "row_count": len(combined),
            "scope_group_distribution": payload["distributions"]["scope_group"],
            "operability_distribution": payload["distributions"]["operability"],
            "scope_reason_distribution": payload["distributions"]["scope_reason"],
            "scope_confidence_distribution": payload["distributions"]["scope_confidence"],
            "test_split_read": False,
            "annotation_modified": False,
            "training_started": False,
        }
        severity_scope_summary_json = root / "final_findings/severity_scope_summary.json"
        _write_json(severity_scope_summary_json, scope_summary)
        created_outputs.append(severity_scope_summary_json)
        severity_scope_summary_md = root / "final_findings/severity_scope_summary.md"
        _write_text(
            severity_scope_summary_md,
            "# FleetVision Severity Scope Summary\n\n"
            + f"- Reviewed: {scope_result.counts['reviewed']}\n"
            + f"- Pending: {scope_result.counts['pending']}\n"
            + f"- Needs adjudication: {scope_result.counts['needs_adjudication']}\n",
        )
        created_outputs.append(severity_scope_summary_md)
        correction_mask = combined["correction_proposal_required"] == "yes"
        correction_cases = (
            combined.loc[
                correction_mask,
                [
                    "review_case_id",
                    "image_id",
                    "annotation_defect_type",
                    "review_notes",
                ],
            ]
            .sort_values(["review_case_id", "image_id"])
            .to_dict(orient="records")
        )
        report = {
            "gate_id": "PHASE_04_5L_FINDINGS_REPORT",
            "created_at_utc": _utc_now(),
            **payload,
            "annotation_correction_proposal_count": int(correction_mask.sum()),
            "annotation_correction_proposal_cases": correction_cases,
            "ranked_action_recommendations": _ranked_action_recommendations(combined),
            "recommendation": recommendation_json,
            "test_split_read": False,
            "model_inference_executed": False,
            "annotation_modified": False,
            "training_started": False,
        }
        findings_report_json = root / "final_findings/phase04_5l_findings_report.json"
        _write_json(findings_report_json, report)
        created_outputs.append(findings_report_json)
        findings_report_md = root / "final_findings/phase04_5l_findings_report.md"
        _write_text(
            findings_report_md,
            _render_report_markdown(payload, recommendation, config),
        )
        created_outputs.append(findings_report_md)
        recommendation_path = root / "final_findings/retraining_recommendation.json"
        _write_json(recommendation_path, recommendation_json)
        created_outputs.append(recommendation_path)
        gate = {
            "gate_id": "PHASE_04_5L_COMPLETED_REVIEW_VALIDATION_AND_FINDINGS_ANALYSIS",
            "outcome": "PASS",
            "classification": "PHASE_04_5L_COMPLETED_REVIEW_VALIDATION_AND_FINDINGS_ANALYSIS_COMPLETED",
            "review_cases": len(combined),
            "scope_reviewed": scope_result.counts["reviewed"],
            "pending": scope_result.counts["pending"],
            "needs_adjudication": scope_result.counts["needs_adjudication"],
            "primary_recommendation": recommendation.primary,
            "test_split_read": False,
            "model_inference_executed": False,
            "annotation_modified": False,
            "training_started": False,
            "retraining_status": config.retraining_status,
            "deployment_acceptance": config.deployment_acceptance,
        }
        gate_path = root / "evidence/gate_result.json"
        _write_json(gate_path, gate)
        created_outputs.append(gate_path)
        workspace_after_path = root / "evidence/workspace_after.csv"
        _write_csv(
            workspace_after_path,
            _manifest_rows(
                root,
                exclude={
                    "evidence/workspace_after.csv",
                    "evidence/SHA256SUMS.csv",
                },
            ),
            ("relative_path", "size_bytes", "sha256"),
        )
        created_outputs.append(workspace_after_path)
        checksum_path = root / "evidence/SHA256SUMS.csv"
        _write_csv(
            checksum_path,
            _manifest_rows(root, exclude={"evidence/SHA256SUMS.csv"}),
            ("relative_path", "size_bytes", "sha256"),
        )
        created_outputs.append(checksum_path)
        return recommendation
    except Exception as exc:
        for output in reversed(created_outputs):
            output.unlink(missing_ok=True)
        _write_blocked_gate(
            root,
            "PHASE_04_5L_COMPLETED_REVIEW_VALIDATION_AND_FINDINGS_ANALYSIS_F2",
            exc,
        )
        for staging in root.rglob(".staging-*"):
            if staging.is_dir():
                shutil.rmtree(staging, ignore_errors=True)
            else:
                staging.unlink(missing_ok=True)
        raise


def _common_parser(description: str) -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=description)
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG_PATH)
    parser.add_argument("--project-root", type=Path, default=Path.cwd())
    parser.add_argument("--expected-head", required=True)
    return parser


def main_f1(argv: Sequence[str] | None = None) -> int:
    parser = _common_parser(
        "Run FleetVision Phase 04.5L F1 completed-review validation and scope package creation."
    )
    parser.add_argument("--timestamp")
    args = parser.parse_args(argv)
    try:
        config = load_findings_config(args.config, args.project_root)
        result = run_f1(
            config,
            args.expected_head,
            timestamp=args.timestamp,
        )
    except Exception as exc:
        print("=== FleetVision Phase 04.5L Findings F1 ===")
        print("Gate classification: PHASE_04_5L_COMPLETED_REVIEW_FINDINGS_F1_BLOCKED")
        print(f"Reason: {exc}")
        print("F1_EXECUTED: NO_OR_BLOCKED")
        print("F2_EXECUTED: NO")
        print("TEST_SPLIT_READ: NO")
        print("MODEL_INFERENCE_EXECUTED: NO")
        print("ANNOTATION_MODIFIED: NO")
        print("TRAINING_STARTED: NO")
        print("RETRAINING_STATUS: NOT_YET_APPROVED")
        print("DEPLOYMENT_ACCEPTANCE: NOT_YET_APPROVED")
        return 1
    print("=== FleetVision Phase 04.5L Findings F1 ===")
    print(
        "Gate classification: "
        "PHASE_04_5L_COMPLETED_REVIEW_VALIDATED_AND_SCOPE_REVIEW_PACKAGE_CREATED"
    )
    print(f"Workspace: {result.root}")
    print(f"Scope Workbook: {result.scope_workbook}")
    print("F1_EXECUTED: YES")
    print("F2_EXECUTED: NO")
    print("TEST_SPLIT_READ: NO")
    print("MODEL_INFERENCE_EXECUTED: NO")
    print("ANNOTATION_MODIFIED: NO")
    print("TRAINING_STARTED: NO")
    print("RETRAINING_STATUS: NOT_YET_APPROVED")
    print("DEPLOYMENT_ACCEPTANCE: NOT_YET_APPROVED")
    return 0


def main_f2(argv: Sequence[str] | None = None) -> int:
    parser = _common_parser(
        "Run FleetVision Phase 04.5L F2 severity-scope validation and final findings analysis."
    )
    parser.add_argument("--workspace-root", type=Path, required=True)
    args = parser.parse_args(argv)
    try:
        config = load_findings_config(args.config, args.project_root)
        recommendation = run_f2(
            config,
            args.workspace_root,
            args.expected_head,
        )
    except Exception as exc:
        print("=== FleetVision Phase 04.5L Findings F2 ===")
        print("Gate classification: PHASE_04_5L_COMPLETED_REVIEW_FINDINGS_F2_BLOCKED")
        print(f"Reason: {exc}")
        print("TEST_SPLIT_READ: NO")
        print("MODEL_INFERENCE_EXECUTED: NO")
        print("ANNOTATION_MODIFIED: NO")
        print("TRAINING_STARTED: NO")
        print("RETRAINING_STATUS: NOT_YET_APPROVED")
        print("DEPLOYMENT_ACCEPTANCE: NOT_YET_APPROVED")
        return 1
    print("=== FleetVision Phase 04.5L Findings F2 ===")
    print(
        "Gate classification: "
        "PHASE_04_5L_COMPLETED_REVIEW_VALIDATION_AND_FINDINGS_ANALYSIS_COMPLETED"
    )
    print(f"Primary advisory recommendation: {recommendation.primary}")
    print("TEST_SPLIT_READ: NO")
    print("MODEL_INFERENCE_EXECUTED: NO")
    print("ANNOTATION_MODIFIED: NO")
    print("TRAINING_STARTED: NO")
    print("RETRAINING_STATUS: NOT_YET_APPROVED")
    print("DEPLOYMENT_ACCEPTANCE: NOT_YET_APPROVED")
    return 0


if __name__ == "__main__":
    raise SystemExit(main_f1())
